import math
from decimal import Decimal, ROUND_FLOOR, ROUND_HALF_UP, getcontext
import logging
import os
import tempfile
import sys
import configparser
import subprocess
import re
import warnings
from pathlib import Path
import platform
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
import psycopg2
from psycopg2 import OperationalError
from psycopg2.extras import execute_values
from io import StringIO
from dotenv import load_dotenv

load_dotenv()

import pandas as pd
import openpyxl
from openpyxl import Workbook
from datetime import datetime, date, timedelta




WORK_MONTH = None
current_treeview_label = None  # 전역 라벨 변수

# logging.basicConfig(
#     level=logging.DEBUG,  # DEBUG 레벨로 설정하여 상세한 로그를 출력
#     format='%(asctime)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.StreamHandler()  # 콘솔(터미널)에 로그를 출력합니다.
#     ]
# )

logging.basicConfig(
    level=logging.ERROR,  # ERROR 이상만 출력
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logging.getLogger().setLevel(logging.ERROR)

# 현재 실행 중인 EXE 파일의 디렉터리 경로를 계산
if getattr(sys, 'frozen', False):  # PyInstaller로 빌드된 경우
    base_path = os.path.dirname(sys.executable)
else:  # 개발 중인 경우
    base_path = os.path.dirname(os.path.abspath(__file__))

# 설정 파일 경로 생성 (EXE와 동일한 경로)
config_path = os.path.join(base_path, 'config_msd.ini')

# 글로벌 설정 객체 생성
config = configparser.ConfigParser()

# 설정 파일 읽기 및 초기화 (인코딩 명시적으로 지정)
if os.path.exists(config_path):
    print(f"{config_path} 파일을 로드합니다.")
    config.read(config_path, encoding='utf-8')  # 인코딩 지정
else:
    print(f"{config_path} 파일을 찾을 수 없습니다. 새로 생성합니다.")
    # 기본 설정 생성
    config['WINDOW'] = {'position': '1200x800'}
    with open(config_path, 'w', encoding='utf-8') as configfile:
        config.write(configfile)  # 새 설정 파일 생성
    print(f'새 설정 파일 생성 완료: {config_path}')

# DB 연결 설정
DB_NAME = os.environ.get('DB_NAME', 'ecommerce')
USER = os.environ.get('DB_USER', 'postgres')
PASSWORD = os.environ.get('DB_PASSWORD', '')
HOST = os.environ.get('DB_HOST', 'foodall.co.kr')
PORT = os.environ.get('DB_PORT', '5432')

def get_postgres_connection():
    try:
        return psycopg2.connect(
            dbname=DB_NAME,
            user=USER,
            password=PASSWORD,
            host=HOST,
            port=PORT
        )
    except OperationalError as e:
        logging.error(f"DB 연결 오류: {e}")
        return None
    
def reset_no_column(treeview):
    """
    정렬 후 'NO' 열의 번호를 다시 설정합니다.
    'NO' 열의 위치를 동적으로 찾아서 업데이트합니다.
    """
    columns = treeview['columns']
    # 'no' 열이 없으면 그냥 반환
    if 'no' not in columns:
        return

    no_column_index = columns.index('no')

    for index, item in enumerate(treeview.get_children('')):
        values = list(treeview.item(item, 'values'))
        if len(values) > no_column_index:
            values[no_column_index] = index + 1  # 'NO' 열의 위치에 인덱스 설정
            treeview.item(item, values=values)


def reapply_row_tags(treeview):
    for index, item in enumerate(treeview.get_children('')):
        tags = list(treeview.item(item, 'tags'))

        # 합계(totalrow)는 건너뜀
        if 'totalrow' in tags:
            continue

        # pinned 행은 even/odd를 덮어씌우지 않음
        if 'pinned' in tags:
            continue

        # 그 외 행: 기존 evenrow/oddrow 제거 후 새로 지정
        tags = [tag for tag in tags if tag not in ('evenrow', 'oddrow')]
        row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'
        tags.append(row_tag)
        treeview.item(item, tags=tags)


# 라인별 음영처리를 위한 함수
def tag_alternate_rows(treeview):
    for index, item in enumerate(treeview.get_children()):
        tags = list(treeview.item(item, 'tags'))
        # pinned 행이면 건너뛰어 evenrow/oddrow를 덮어씌우지 않음
        if 'pinned' in tags:
            continue

        # 기존 evenrow/oddrow 제거
        tags = [tag for tag in tags if tag not in ('evenrow','oddrow')]
        # 짝수/홀수
        row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'
        tags.append(row_tag)
        treeview.item(item, tags=tags)
        


def safe_float_from_string(value):
    """
    문자열을 부동 소수점 숫자로 변환하는 함수입니다.
    변환이 불가능한 경우 None을 반환합니다.
    """
    if value is None or value == '':
        return None  # None이나 빈 문자열은 None으로 반환
    value = re.sub(r'[^\d.-]', '', str(value))  # 숫자, 마이너스, 소수점만 남김
    try:
        return float(value)
    except ValueError:
        return None  # 변환이 불가능한 경우 None 반환


def format_numeric_columns(treeview, numeric_columns):
    """
    트리뷰의 숫자 컬럼에 대해 포맷팅을 적용.
    - 'amount_columns_treeview0'에 속하는 컬럼은 정수(반올림) + 천단위 쉼표 처리
    - 그 외 숫자 컬럼은 기존 로직(소수점 5자리 등)
    """
    for child in treeview.get_children():
        item = treeview.item(child)
        values = item['values']
        formatted_values = []
        
        for col_id, value in zip(treeview['columns'], values):
            # numeric_columns에 포함된 컬럼이면 처리
            if col_id in numeric_columns:
                numeric_value = safe_float_from_string(value)
                
                if col_id in amount_columns_treeview0:
                    # 금액 컬럼: 정수 처리
                    formatted_value = format_amount_value(numeric_value)
                else:
                    # 기타(단가, 수량 등)는 기존 방식
                    formatted_value = format_numeric_value(numeric_value)
                    
                formatted_values.append(formatted_value)
            else:
                # 숫자 컬럼이 아니라면 그대로 둠
                formatted_values.append(value)

        treeview.item(child, values=formatted_values)


def format_amount_value(value):
    """
    금액만 정수로 표시하는 포맷팅 함수.
    """
    if value is None:
        return ''
    try:
        val_rounded = round(float(value))
        return f"{int(val_rounded):,}"  # 천단위 쉼표
    except:
        return ''


from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

def format_numeric_value(value, decimals: int = 5) -> str:
    """
    숫자를 6째 자리에서 반올림 → 5째 자리까지 표시.
    정수면 천단위 쉼표, 소수면 최대 5자리(뒤쪽 0·점 제거).
    """
    if value in (None, ''):
        return ''

    try:
        # ① Decimal 변환 후 6째 자리에서 반올림
        quant = Decimal('1.' + '0' * decimals)          # 0.00001
        num_q = Decimal(str(value)).quantize(
            quant, rounding=ROUND_HALF_UP
        )

        # ② 정수 여부 판별
        if num_q == num_q.to_integral_value():
            return f"{int(num_q):,}"
        else:
            return f"{num_q:,.{decimals}f}".rstrip('0').rstrip('.')

    except (InvalidOperation, ValueError):
        return ''




# 숫자 컬럼 정의

# ====================================================================
# 2. numeric_columns_treeview0 업데이트 (새로운 컬럼들 추가)
# ====================================================================

numeric_columns_treeview0 = [
    # 기존 컬럼들
    'beginning_unit_price', 'beginning_quantity', 'beginning_amount',
    'incoming_unit_price', 'incoming_quantity', 'incoming_amount',
    'misc_profit_amount', 'incentive_amount',
    'transfer_in_free_quantity', 'transfer_in_code_change_quantity', 'transfer_in_code_change_amount',
    'outgoing_unit_price', 'outgoing_quantity', 'outgoing_amount',
    'transfer_out_donation_quantity', 'transfer_out_donation_amount',
    'transfer_out_free_quantity', 'transfer_out_free_amount',
    'transfer_out_internal_use_quantity', 'transfer_out_internal_use_amount',
    'transfer_out_sample_quantity', 'transfer_out_sample_amount',
    'transfer_out_employee_gift_quantity', 'transfer_out_employee_gift_amount',
    'transfer_out_code_change_quantity', 'transfer_out_code_change_amount',
    'transfer_out_loss_quantity', 'transfer_out_loss_amount',
    'transfer_out_account_substitution_quantity', 'transfer_out_account_substitution_amount',
    'transfer_out_accident_compensation_quantity', 'transfer_out_accident_compensation_amount',
    'transfer_out_expired_quantity', 'transfer_out_expired_amount',
    'transfer_out_inventory_adjustment_quantity', 'transfer_out_inventory_adjustment_amount',
    'transfer_out_regular_inventory_check_quantity', 'transfer_out_regular_inventory_check_amount',
    'transfer_out_claim_processing_quantity', 'transfer_out_claim_processing_amount',
    'current_unit_price', 'current_quantity', 'current_amount',
    'verification_quantity', 'verification_amount',
    
    # 기존 재고실사 컬럼들
    'inventory_inspection_chasan', 'inventory_inspection_icheon', 'inventory_inspection_hanam',
    'inventory_inspection_cheongnyangni', 'inventory_inspection_total_quantity',
    'inventory_inspection_difference_quantity',
    'inventory_inspection_chasan_amount', 'inventory_inspection_icheon_amount',
    'inventory_inspection_hanam_amount', 'inventory_inspection_cheongnyangni_amount',
    'inventory_inspection_total_amount', 'inventory_inspection_difference_amount',
    
    # === 새로운 재고실사 컬럼들 ===
    # 차산점 관련
    'inventory_inspection_chasan_sum_quantity', 'inventory_inspection_chasan_sum_amount',
    'inventory_inspection_chasan_a_quantity', 'inventory_inspection_chasan_a_amount',
    'inventory_inspection_import_warehouse_chasan_quantity', 'inventory_inspection_import_warehouse_chasan_amount',
    'inventory_inspection_import_warehouse_storage_quantity', 'inventory_inspection_import_warehouse_storage_amount',
    'inventory_inspection_chasan_return_quantity', 'inventory_inspection_chasan_return_amount',
    
    # 청량리점 관련
    'inventory_inspection_cheongnyangni_sum_quantity', 'inventory_inspection_cheongnyangni_sum_amount',
    'inventory_inspection_cheongnyangni_return_quantity', 'inventory_inspection_cheongnyangni_return_amount',
    
    # 이천점 관련
    'inventory_inspection_icheon_sum_quantity', 'inventory_inspection_icheon_sum_amount',
    'inventory_inspection_catering_quantity', 'inventory_inspection_catering_amount',
    'inventory_inspection_ecommerce_quantity', 'inventory_inspection_ecommerce_amount',
    'inventory_inspection_icheon_return_quantity', 'inventory_inspection_icheon_return_amount',
    
    # 기타
    'inventory_inspection_prepurchase_quantity', 'inventory_inspection_prepurchase_amount',
]

# ====================================================================
# 3. amount_columns_treeview0 업데이트 (금액 컬럼들에 새로운 컬럼들 추가)
# ====================================================================

amount_columns_treeview0 = [
    # 기존 금액 컬럼들
    'beginning_amount', 'incoming_amount', 'misc_profit_amount', 'incentive_amount',
    'transfer_in_code_change_amount', 'outgoing_amount', 'transfer_out_donation_amount',
    'transfer_out_free_amount', 'transfer_out_internal_use_amount', 'transfer_out_sample_amount',
    'transfer_out_employee_gift_amount', 'transfer_out_code_change_amount', 'transfer_out_loss_amount',
    'transfer_out_account_substitution_amount', 'transfer_out_accident_compensation_amount',
    'transfer_out_expired_amount', 'transfer_out_inventory_adjustment_amount', 
    'transfer_out_regular_inventory_check_amount', 'transfer_out_claim_processing_amount',
    'current_amount',
    'inventory_inspection_chasan_amount', 'inventory_inspection_icheon_amount',
    'inventory_inspection_hanam_amount', 'inventory_inspection_cheongnyangni_amount',
    'inventory_inspection_total_amount', 'inventory_inspection_difference_amount',
    
    # === 새로운 금액 컬럼들 ===
    'inventory_inspection_chasan_sum_amount',
    'inventory_inspection_chasan_a_amount',
    'inventory_inspection_import_warehouse_chasan_amount',
    'inventory_inspection_import_warehouse_storage_amount',
    'inventory_inspection_chasan_return_amount',
    'inventory_inspection_cheongnyangni_sum_amount',
    'inventory_inspection_cheongnyangni_return_amount',
    'inventory_inspection_icheon_sum_amount',
    'inventory_inspection_catering_amount',
    'inventory_inspection_ecommerce_amount',
    'inventory_inspection_icheon_return_amount',
    'inventory_inspection_prepurchase_amount',
]

numeric_columns_treeview1 = [
    '차산점', '차산점A', '수입창고', '청량리점', '이천점',
    '케이터링', '하남점', '이커머스', '선매입창고', '합계수량',
    '차산점반품', '청량리반품', '이천점반품', '하남점반품', '반품합계수량',
    '차산점폐기', '이천점폐기', '폐기합계수량'
]
numeric_columns_treeview2 = ['양품출고량', '단가', '금액', '외화금액', '단위중량']
numeric_columns_treeview3 = ['입고량', '단가', '금액']
numeric_columns_treeview4 = [
    '출하수량', '단가', '금액', '원화금액(출하)', '부가세(출하)',
    '원화금액(매출)', '부가세(매출)', '총금액(출하)', '총금액(매출)', '중량'
]
numeric_columns_treeview5 = ['부가세', '총금액', '입고수량', '입고금액', '외화금액']
numeric_columns_treeview6 = [
    '기초수량', '기초단가', '기초금액',
    '입고수량', '입고금액', '대체수량', '대체금액',
    '출고수량', '출고금액', '재고수량', '재고단가', '재고금액'
]
numeric_columns_treeview8 = ['sum_won_amount', 'sum_vat_amount', 'sum_total_amount', 'ratio', 'incentive']

# 정렬 함수 정의 (기본 기능 유지, 필요 변경사항 주석 표시)
def sort_treeview_column(treeview, col, numeric_columns, reverse=None):
    """
    트리뷰의 특정 열(col)을 클릭하면 정렬 방향(오름차순/내림차순)을 번갈아가며 적용합니다.
    + totalrow를 최상단, pinned를 그 다음, 일반 행을 정렬 후 아래쪽에 배치
    """
    # 'NO' 열 등 특정 열은 정렬하지 않음
    if col.lower() == 'no':
        return

    # 정렬 상태 초기화
    if not hasattr(treeview, '_sort_states'):
        treeview._sort_states = {}

    # 이전 정렬 상태 확인 및 토글
    if reverse is None:  # reverse가 None이면, 클릭에 따라 방향을 토글
        reverse = not treeview._sort_states.get(col, False)
    treeview._sort_states[col] = reverse

    # 1) totalrow, pinned, 일반 행을 분리
    totalrow_items = []
    pinned_items = []
    normal_items = []
    for child in treeview.get_children(''):
        tags = treeview.item(child, 'tags')
        if 'totalrow' in tags:
            totalrow_items.append(child)
        elif 'pinned' in tags:
            pinned_items.append(child)
        else:
            normal_items.append(child)

    # 2) 일반 행만 정렬 대상
    data = []
    for item in normal_items:
        cell_value = treeview.set(item, col)  # 해당 열의 값
        data.append((cell_value, item))

    # 숫자 컬럼인지 확인
    is_numeric = (col in numeric_columns)

    # 3) 정렬 로직
    try:
        if is_numeric:
            # 숫자 열 정렬
            processed_data = []
            for value, item in data:
                cleaned_value = re.sub(r'[^\d.-]', '', value)
                try:
                    numeric_value = float(cleaned_value) if cleaned_value else 0
                except ValueError:
                    numeric_value = 0
                processed_data.append((numeric_value, item))
            processed_data.sort(reverse=reverse)
            data = processed_data
        else:
            # 텍스트 열 정렬
            data.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)
    except Exception as e:
        logging.error(f"열 '{col}' 정렬 중 오류 발생: {e}")
        # 오류 발생 시에도 텍스트 기준 정렬 시도
        data.sort(key=lambda t: str(t[0]).lower(), reverse=reverse)

    # 4) 정렬된 데이터를 트리뷰에 재배치
    #    => "totalrow" 먼저, "pinned" 다음, 정렬된 일반 행 마지막
    current_index = 0

    # (A) totalrow 행들을 최상단
    for t_item in totalrow_items:
        treeview.move(t_item, '', current_index)
        current_index += 1

    # (B) pinned 행들을 그 다음
    for p_item in pinned_items:
        treeview.move(p_item, '', current_index)
        current_index += 1

    # (C) 정렬된 일반 행들
    for val, item in data:
        treeview.move(item, '', current_index)
        current_index += 1

    # 5) 행의 음영 처리 재적용
    reapply_row_tags(treeview)

    # 6) 다음 클릭 시 정렬 방향을 반대로 설정
    treeview.heading(col, command=lambda: sort_treeview_column(treeview, col, numeric_columns, None))

    # 7) 숫자 열 천단위 포맷팅 등 적용
    format_numeric_columns(treeview, numeric_columns)

    # 8) 정렬 후 NO 열 번호 재설정
    reset_no_column(treeview)



def upload_warehouse_inventory():
    """
    창고별 재고 현황 엑셀 파일을 업로드하여 master 테이블의 수량을 업데이트하는 함수
    """
    # 파일 경로 설정
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,  # 기본 경로 설정
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
    )
    if not file_path:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")
        return

    confirm = messagebox.askyesno("확인", "창고 데이터 업로드를 진행하면 기존 데이터를 일괄 업데이트합니다.\n업로드를 계속하시겠습니까?")
    if not confirm:
        return

    try:
        print("엑셀 파일을 읽는 중...")
        # 파일 읽기 (헤더가 2라인인 경우 처리)
        data = pd.read_excel(file_path, header=[0, 1], dtype=str)
        print("엑셀 파일 로드 완료.")

        print("창고 컬럼을 처리 중...")
        # 첫 번째 헤더에서 창고명을 추출하고 두 번째 헤더와 조합하여 컬럼명 생성
        warehouse_columns = []
        for col in data.columns:
            if col[0] in ['차산점', '차산점A', '수입창고', '청량리점', '이천점', '케이터링', '하남점', '이커머스', '선매입창고', '차산점반품', '차산점폐기', '청량리반품', '이천점반품', '이천점폐기', '하남점반품']:
                combined_column_name = f"{col[0]}_{col[1].replace('(', '').replace(')', '').replace(' ', '_').lower()}"
                combined_column_name = combined_column_name.replace('box', '_box').replace('낱개', '_낱개')
                warehouse_columns.append((combined_column_name, col))
        print(f"처리된 창고 컬럼 수: {len(warehouse_columns)}")

        print("임시 데이터프레임 생성 중...")
        # 필요한 데이터만 추출하여 임시 DataFrame 생성
        temp_df = pd.DataFrame()
        temp_df['item_code'] = data[('품목', '품목')].astype(str).str.strip()

        for combined_column_name, warehouse in warehouse_columns:
            temp_df[combined_column_name] = data[warehouse].fillna('0').astype(int)
        print(f"임시 데이터프레임 생성 완료. 총 품목 수: {len(temp_df)}")

        # DB 연결
        print("데이터베이스에 연결 중...")
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        print("데이터베이스 연결 성공.")

        try:
            # 트랜잭션 시작 전에 autocommit 설정
            conn.autocommit = False
            cursor = conn.cursor()
            print("트랜잭션 시작.")

            # 임시 테이블 생성
            print("임시 테이블 생성 중...")
            create_temp_table_query = f"""
                CREATE TEMP TABLE temp_master_update (
                    item_code VARCHAR PRIMARY KEY,
                    {', '.join([f"{col} INTEGER" for col, _ in warehouse_columns])}
                ) ON COMMIT DROP;
            """
            cursor.execute(create_temp_table_query)
            print("임시 테이블 생성 완료.")

            # 임시 테이블에 데이터 삽입 (배치 삽입)
            print("임시 테이블에 데이터 삽입 중...")
            tuples = [tuple(x) for x in temp_df.to_numpy()]
            cols = ','.join(temp_df.columns)
            values_placeholder = ','.join(['%s'] * len(temp_df.columns))
            insert_query = f"INSERT INTO temp_master_update ({cols}) VALUES ({values_placeholder})"
            cursor.executemany(insert_query, tuples)
            print("임시 테이블에 데이터 삽입 완료.")

            # 기존 창고 데이터 일괄 초기화
            print("기존 창고 데이터 초기화 중...")
            reset_columns = ", ".join([f"{col} = 0" for col, _ in warehouse_columns])
            reset_query = f"UPDATE master SET {reset_columns};"
            cursor.execute(reset_query)
            print("기존 창고 데이터 초기화 완료.")

            # master 테이블과 임시 테이블을 조인하여 업데이트
            print("master 테이블 업데이트 중...")
            set_statements = ", ".join([f"{col} = t.{col}" for col, _ in warehouse_columns])  # 'm.' 제거
            update_query = f"""
                UPDATE master m
                SET {set_statements}
                FROM temp_master_update t
                WHERE m.item_code = t.item_code;
            """
            cursor.execute(update_query)
            print("master 테이블 업데이트 완료.")

            # master 테이블에 임시 테이블에 없는 item_code의 창고 데이터를 0으로 설정
            print("존재하지 않는 item_code의 창고 데이터를 0으로 설정 중...")
            zero_out_query = f"""
                UPDATE master m
                SET {reset_columns}
                WHERE NOT EXISTS (
                    SELECT 1 FROM temp_master_update t WHERE m.item_code = t.item_code
                );
            """
            cursor.execute(zero_out_query)
            print("존재하지 않는 item_code의 창고 데이터 0 설정 완료.")

            # 마지막 업데이트 시간 기록
            print("마지막 업데이트 시간 기록 중...")
            cursor.execute("UPDATE master SET last_updated = %s WHERE TRUE", (datetime.now(),))
            print("마지막 업데이트 시간 기록 완료.")

            # 트랜잭션 커밋
            print("트랜잭션을 커밋 중...")
            conn.commit()
            print("트랜잭션 커밋 완료.")

            # 업데이트된 품목 수 계산
            updated_count = len(temp_df)
            print(f"{updated_count}개의 품목이 업데이트되었습니다.")

            messagebox.showinfo("성공", f"창고별 재고 현황이 성공적으로 업데이트되었습니다.\n업데이트된 품목 수: {updated_count}")

        except Exception as e:
            conn.rollback()
            logging.error(f"트랜잭션 중 오류 발생: {e}")
            print("트랜잭션 롤백 중...")
            messagebox.showerror("오류", f"데이터베이스 업데이트 중 오류가 발생했습니다.\n{e}")
        finally:
            cursor.close()
            conn.autocommit = True
            print("데이터베이스 연결 종료.")

    except Exception as e:
        messagebox.showerror("오류", f"오류 발생: {e}")


def delete_old_files(directory):
    today_str = datetime.now().strftime('%Y-%m-%d')
    for filename in os.listdir(directory):
        filepath = os.path.join(directory, filename)
        if os.path.isfile(filepath):
            file_date_str = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime('%Y-%m-%d')
            if file_date_str != today_str:
                try:
                    os.remove(filepath)
                    logging.info(f"이전 파일 삭제됨: {filepath}")
                except Exception as e:
                    logging.warning(f"파일 삭제 실패: {filepath}. 에러: {e}")


def truncate_string(value, max_length):
    """지정된 최대 길이로 문자열을 자릅니다."""
    if isinstance(value, str) and len(value) > max_length:
        return value[:max_length]
    return value

from decimal import Decimal
from psycopg2.extras import execute_values
import pandas as pd
from tkinter import messagebox, simpledialog, filedialog
import tempfile, io, os
from datetime import datetime

# --------------------------------------------------
# 공통 유틸리티
# --------------------------------------------------

def truncate_string(value: str, max_len: int) -> str:
    """지정 길이 초과 시 말줄임표 없이 자릅니다."""
    return value[:max_len]

# --------------------------------------------------
# 1) upload_basic_data  ── execute_values 버전 (이전 내용과 동일)
# --------------------------------------------------

def upload_basic_data():
    """엑셀 기초데이터를 PostgreSQL로 고속 업로드 (execute_values 사용)"""
    global root

    # 1. 기준년 입력
    year = simpledialog.askstring("기준년 입력", "기준년을 입력하세요 (예: 2024):", parent=root)
    if not year:
        messagebox.showwarning("경고", "기준년을 입력해야 합니다.")
        return

    # 2. 파일 선택
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return

    # 3. DB 연결
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 4. 기존 연도 데이터 삭제 여부
        cursor.execute("SELECT COUNT(*) FROM mds_basic_data WHERE reference_year = %s", (year,))
        if cursor.fetchone()[0] > 0:
            if not messagebox.askyesno("확인", f"{year}년도 데이터를 삭제하고 다시 업로드하시겠습니까?"):
                messagebox.showinfo("정보", "업로드를 취소합니다.")
                return
            cursor.execute("DELETE FROM mds_basic_data WHERE reference_year = %s", (year,))
            conn.commit()
            messagebox.showinfo("정보", f"{year}년도 기존 데이터를 삭제했습니다.")

        # 5. 엑셀 로드 & 전처리
        df = pd.read_excel(file_path, header=[0, 1])
        df.columns = pd.MultiIndex.from_tuples([
            (str(c[0]).strip() if c[0] else "", str(c[1]).strip() if "Unnamed" not in str(c[1]) else "")
            for c in df.columns
        ])
        df.columns = ["_".join(filter(None, c)) for c in df.columns]

        rename_map = {
            "품목코드": "item_code",
            "품명": "item_name",
            "규격": "specification",
            "단위": "unit",
            "분류": "category",
            "이월재고_단가": "beginning_unit_price",
            "이월재고_수량": "beginning_quantity",
            "이월재고_금액": "beginning_amount",
        }
        df = df.rename(columns=rename_map)

        required = [
            "item_code", "item_name", "specification", "unit", "category",
            "beginning_unit_price", "beginning_quantity", "beginning_amount",
        ]
        miss = [c for c in required if c not in df.columns]
        if miss:
            messagebox.showerror("에러", f"다음 컬럼이 누락되었습니다: {', '.join(miss)}")
            return

        # "합계" 이후 제거
        df = df.reset_index(drop=True)
        sum_idx = df.index[df["item_code"].astype(str).str.strip() == "합계"].tolist()
        if sum_idx:
            df = df.loc[: sum_idx[0] - 1]

        # 숫자 컬럼 Decimal 변환(5자리)
        for col in ["beginning_unit_price", "beginning_quantity", "beginning_amount"]:
            df[col] = (
                pd.to_numeric(df[col], errors="coerce")
                .fillna(0)
                .round(5)
                .apply(lambda x: Decimal(f"{x:.5f}"))
            )

        df["reference_year"] = year

        # execute_values
        data_cols = [
            "item_code", "item_name", "specification", "unit", "category",
            "beginning_unit_price", "beginning_quantity", "beginning_amount", "reference_year",
        ]
        data = list(df[data_cols].itertuples(index=False, name=None))
        insert_query = f"INSERT INTO mds_basic_data ({', '.join(data_cols)}) VALUES %s"
        execute_values(cursor, insert_query, data)

        conn.commit()
        messagebox.showinfo("성공", f"업로드 완료: {len(df)}개 행이 업로드되었습니다.")

    except Exception as e:
        conn.rollback()
        messagebox.showerror("에러", f"업로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

# --------------------------------------------------
# 2) upload_master_data  ── COPY + 임시테이블 개선 버전
# --------------------------------------------------

def upload_master_data():
    """master 테이블 대용량 업로드 (임시테이블 + COPY)
    - 엑셀/CSV → DataFrame
    - 행 수 검증
    - TEMP TABLE + copy_expert(StringIO) → UPDATE / INSERT
    """

    directory = "C:/ERPUExport"
    delete_old_files(directory)  # 기존 정리 함수가 있으면 사용

    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("CSV files", "*.csv"), ("All files", "*.*")],
    )
    if not file_path:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")
        return

    # -------------------------------- DB 연결 --------------------------------
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("오류", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM master")
        db_count = cursor.fetchone()[0] or 0
        conn.commit()  # 자동 트랜잭션 종료

        # -------------------------------- 파일 로드 --------------------------------
        if file_path.endswith(".csv"):
            df = pd.read_csv(file_path, dtype=str)
        else:
            df = pd.read_excel(file_path, dtype=str)
        file_count = len(df)

        # 행 수 차이 검증
        if (db_count - file_count) >= 500:
            messagebox.showwarning("경고", "사용유무 Y/N 필터링 된 전체데이터를 다운로드 후 진행해주세요")
            return

        # -------------------------------- 컬럼 매핑 --------------------------------
        rename_map = {
            "품목코드": "item_code",
            "품목명": "item_name",
            "규격": "specification",
            "대구분": "major_category",
            "중구분": "minor_category",
            "주거래처명": "main_supplier_name",
            "매입부가세": "purchase_vat",
            "등록자명": "registrant_name",
            "등록일": "registration_date",
            "수정자명": "modifier_name",
            "수정일": "modification_date",
            "상품바코드": "product_barcode",
            "박스바코드": "box_barcode",
            "유통기한": "expiration_period",
            "보관방법": "storage_method",
            "원산지": "origin",
            "정사파레트적재수량": "standard_pallet_qty",
            "직파레트적재수량": "straight_pallet_qty",
            "입수수량": "unit_quantity",
            "출하단위": "unit",
            "계정구분": "category",
        }
        df = df.rename(columns=rename_map)

        insert_cols = [
            "item_code", "item_name", "specification", "major_category", "minor_category",
            "main_supplier_name", "purchase_vat", "registrant_name", "registration_date",
            "modifier_name", "modification_date", "product_barcode", "box_barcode",
            "expiration_period", "storage_method", "origin", "standard_pallet_qty",
            "straight_pallet_qty", "unit_quantity", "unit", "category",
        ]
        for col in insert_cols:
            if col not in df.columns:
                df[col] = None

        # 문자열 컬럼 정제/길이 제한
        str_cols = [
            "item_code", "item_name", "specification", "major_category", "minor_category",
            "main_supplier_name", "purchase_vat", "registrant_name", "modifier_name",
            "product_barcode", "box_barcode", "expiration_period", "storage_method",
            "origin", "unit", "category",
        ]
        max_len = {
            "item_code": 255,
            "item_name": 255,
            "specification": 100,
            "major_category": 50,
            "minor_category": 50,
            "main_supplier_name": 100,
            "purchase_vat": 100,
            "registrant_name": 50,
            "modifier_name": 50,
            "product_barcode": 255,
            "box_barcode": 255,
            "expiration_period": 20,
            "storage_method": 50,
            "origin": 50,
            "unit": 50,
            "category": 50,
        }
        for col in str_cols:
            df[col] = (
                df[col]
                .astype(str)
                .str.strip()
                .replace({"nan": None, "NaN": None, "None": None, "": None})
                .apply(lambda x: truncate_string(x, max_len[col]) if x else None)
            )

        # 날짜/정수 컬럼 처리
        df["registration_date"] = pd.to_datetime(df["registration_date"], format="%Y/%m/%d", errors="coerce").dt.date
        df["modification_date"] = pd.to_datetime(df["modification_date"], format="%Y/%m/%d", errors="coerce").dt.date
        df[["standard_pallet_qty", "straight_pallet_qty", "unit_quantity"]] = (
            df[["standard_pallet_qty", "straight_pallet_qty", "unit_quantity"]]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .astype(int)
        )
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        df["last_updated"] = now_str

        # -------------------------------- 임시테이블 + COPY --------------------------------
        conn.autocommit = False
        tmp_table = "temp_master_upload"
        cursor.execute(f"""
            CREATE TEMP TABLE {tmp_table} (LIKE master INCLUDING ALL) ON COMMIT DROP;
        """)

        # copy_expert via StringIO (메모리 내 스트림)
        buffer = io.StringIO()
        df[insert_cols + ["last_updated"]].to_csv(buffer, index=False, header=False)
        buffer.seek(0)
        cursor.copy_expert(
            f"COPY {tmp_table} ({', '.join(insert_cols + ['last_updated'])}) FROM STDIN WITH (FORMAT CSV, DELIMITER ',', NULL '')",
            buffer,
        )

        # -------------------------------- UPDATE --------------------------------
        cursor.execute(
            f"""
            UPDATE master m SET
                item_name = t.item_name,
                specification = t.specification,
                major_category = t.major_category,
                minor_category = t.minor_category,
                main_supplier_name = t.main_supplier_name,
                purchase_vat = t.purchase_vat,
                registrant_name = t.registrant_name,
                registration_date = t.registration_date,
                modifier_name = t.modifier_name,
                modification_date = t.modification_date,
                product_barcode = t.product_barcode,
                box_barcode = t.box_barcode,
                expiration_period = t.expiration_period,
                storage_method = t.storage_method,
                origin = t.origin,
                standard_pallet_qty = t.standard_pallet_qty,
                straight_pallet_qty = t.straight_pallet_qty,
                unit_quantity = t.unit_quantity,
                unit = t.unit,
                category = t.category,
                last_updated = t.last_updated
            FROM {tmp_table} t
            WHERE m.item_code = t.item_code
              AND (m.* IS DISTINCT FROM t.*)
            """
        )
        updated = cursor.rowcount

        # -------------------------------- INSERT (신규) --------------------------------
        cursor.execute(
            f"""
            INSERT INTO master ({', '.join(insert_cols + ['last_updated'])})
            SELECT {', '.join(['t.' + c for c in insert_cols])}, t.last_updated
            FROM {tmp_table} t
            LEFT JOIN master m ON m.item_code = t.item_code
            WHERE m.item_code IS NULL
            """
        )
        inserted = cursor.rowcount

        conn.commit()
        messagebox.showinfo(
            "성공",
            f"마스터 데이터 등록 완료\n전체: {len(df)}개 | 업데이트: {updated} | 신규: {inserted}",
        )
        load_last_updated_info()

    except Exception as ex:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"오류 발생: {ex}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()





def quote_identifier(identifier: str) -> str:
    """
    PostgreSQL에서 안전하게 식별자를 감싸기 위해,
    내부의 쌍따옴표는 ""로 이스케이프하고, 전체를 "..."로 감싸줍니다.
    예: 수입창고(차산)_수량 → "수입창고(차산)_수량"
    """
    escaped = identifier.replace('"', '""')
    return f'"{escaped}"'

def upload_monthly_inventory_status():
    """
    월별 기말재고등록 엑셀 파일을 업로드하여 mds_monthly_inventory_status 테이블에 데이터를 삽입하는 함수.
    (특수문자 '(' ')'가 들어간 컬럼명도 quote_identifier를 통해 안전하게 COPY 및 INSERT)
    개선 버전:
      1) 엑셀 → DataFrame → CSV(메모리) → PostgreSQL 임시테이블(COPY) → 최종테이블(INSERT SELECT)
      2) 기존 데이터 중복 시, 사용자에게 초기화 여부 확인
      3) exclude_item_codes 테이블에 포함된 품목은 제거
    """
    # 1) 기준월 입력
    current_year_month = datetime.now().strftime("%Y/%m")  # 기본값(현재 년/월)
    year_month = simpledialog.askstring("기준월 선택", "년/월을 입력하세요 :", initialvalue=current_year_month)
    if not year_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    # 2) 파일 선택
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
    )
    if not file_path:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")
        return

    try:
        logging.info("엑셀 파일을 읽는 중...")
        # header=0, skiprows=[1,2]: 첫 행을 헤더로, 2~3행 스킵
        data = pd.read_excel(file_path, header=0, skiprows=[1, 2], dtype=str)
        logging.info("엑셀 파일 로드 완료.")

        # ───────────── [A] 엑셀 컬럼 전처리 ─────────────
        # 1) 'Unnamed' 컬럼 제거
        data = data.loc[:, ~data.columns.str.contains('^Unnamed')]

        # 2) 엑셀 → DB 컬럼명 매핑
        rename_dict = {
            '품목': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '차산점': '차산점_수량',
            '차산점A': '차산점a_수량',
            '수입창고(차산)': '수입창고(차산)_수량',
            '수입창고(보관)': '수입창고(보관)_수량',
            '청량리점': '청량리점_수량',
            '이천점': '이천점_수량',
            '케이터링': '케이터링_수량',
            '하남점': '하남점_수량',
            '이커머스': '이커머스_수량',
            '선매입창고': '선매입창고_수량',
            '차산점반품': '차산점반품_수량',
            '차산점폐기': '차산점폐기_수량',
            '청량리반품': '청량리반품_수량',
            '이천점반품': '이천점반품_수량',
            '이천점폐기': '이천점폐기_수량',
            '하남점반품': '하남점반품_수량'
        }
        # 존재하지 않는 엑셀 컬럼은 rename_dict에서 제거
        for old_col in list(rename_dict.keys()):
            if old_col not in data.columns:
                logging.info(f"엑셀에 '{old_col}' 컬럼이 없음 → 건너뜁니다.")
                del rename_dict[old_col]

        # rename
        data = data.rename(columns=rename_dict)

        # ───────────── [B] 숫자 변환 대상 컬럼 ─────────────
        integer_columns = [
            '차산점_수량', '차산점a_수량',
            '수입창고(차산)_수량', '수입창고(보관)_수량',
            '청량리점_수량', '이천점_수량',
            '케이터링_수량', '하남점_수량',
            '이커머스_수량', '선매입창고_수량',
            '차산점반품_수량', '차산점폐기_수량',
            '청량리반품_수량', '이천점반품_수량',
            '이천점폐기_수량', '하남점반품_수량'
        ]
        for col in integer_columns:
            if col not in data.columns:
                logging.info(f"'{col}' 컬럼이 없어 정수 변환을 건너뜁니다.")
                continue
            data[col] = data[col].astype(str).str.replace(',', '').replace('', '0')
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0).astype(int)

        # 문자열 컬럼 처리
        for col in ['item_code', 'item_name', 'specification']:
            if col in data.columns:
                data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})

        # 기준월 컬럼 추가
        data['reference_month'] = year_month

        # ───────────── [C] DB 연결 및 기존 데이터 처리 ─────────────
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # 기존 데이터 존재 여부
        cursor.execute('SELECT COUNT(*) FROM mds_monthly_inventory_status WHERE reference_month = %s', (year_month,))
        existing_data_count = cursor.fetchone()[0]

        if existing_data_count > 0:
            if messagebox.askyesno("기존 데이터 확인", f"기준월 {year_month}의 데이터가 이미 존재합니다. 초기화하고 업로드 하시겠습니까?"):
                cursor.execute('DELETE FROM mds_monthly_inventory_status WHERE reference_month = %s', (year_month,))
                conn.commit()
                logging.info(f"{year_month} 기존 데이터 초기화 완료.")
            else:
                logging.info("업로드가 취소되었습니다.")
                cursor.close()
                conn.close()
                return

        # ───────────── [D] 실제 삽입할 컬럼 결정 ─────────────
        # data.columns 예: ['item_code', 'item_name', '차산점_수량', ... , 'reference_month']
        db_cols = [
            'item_code', 'item_name', 'specification',
            '차산점_수량', '차산점a_수량',
            '수입창고(차산)_수량', '수입창고(보관)_수량',
            '청량리점_수량', '이천점_수량',
            '케이터링_수량', '하남점_수량', '이커머스_수량',
            '선매입창고_수량', '차산점반품_수량', '차산점폐기_수량',
            '청량리반품_수량', '이천점반품_수량', '이천점폐기_수량',
            '하남점반품_수량',
            'reference_month'
        ]
        # 교집합
        final_cols = [c for c in db_cols if c in data.columns]
        logging.info(f"실제 삽입할 컬럼: {final_cols}")

        insert_data = data[final_cols]

        # ───────────── [E] 임시테이블 + COPY 방식을 위한 사전 작업 ─────────────
        # 1) 임시테이블 만들기
        #    (모든 컬럼을 TEXT로 받아두고, 최종테이블로 INSERT SELECT 시 필요한 CAST나 변환을 진행)
        temp_table = "mds_monthly_inventory_status_temp"
        cursor.execute(f"DROP TABLE IF EXISTS {temp_table}")
        # 각 컬럼을 모두 TEXT로 만듦 (여기서는 간단히 TEXT. 실제론 타입을 정확히 지정해도 됨)
        create_cols_sql = []
        for col in final_cols:
            create_cols_sql.append(f'{quote_identifier(col)} TEXT')
        create_table_sql = f"CREATE TEMP TABLE {temp_table} ({', '.join(create_cols_sql)}) ON COMMIT DROP;"
        cursor.execute(create_table_sql)
        logging.info("임시테이블 생성 완료")

        # 2) DataFrame → CSV (메모리상 StringIO 사용)
        #    header=False, index=False 로 생성 (COPY 시 테이블 컬럼 순서대로 매핑)
        csv_buffer = StringIO()
        insert_data.to_csv(csv_buffer, index=False, header=False)  
        csv_buffer.seek(0)

        # 3) COPY FROM STDIN (CSV 형식)
        #    컬럼 구분자는 기본 쉼표, 문자열 내의 따옴표/쉼표는 to_csv에서 자동 처리됨
        copy_sql = f"COPY {temp_table} FROM STDIN WITH CSV"
        cursor.copy_expert(copy_sql, csv_buffer)
        logging.info(f"COPY 완료 (행 수: {len(insert_data)})")

        # ───────────── [F] 임시테이블 → 최종테이블 INSERT ─────────────
        #    최종테이블에는 실제 타입에 맞춰 CAST가 필요할 수 있음.
        #    예) 수량 컬럼들은 INT로, 문자열은 TEXT로, etc.
        #    여기서는 간단히 ( TEXT → INT )가 필요한 것들만 CAST 처리 예시
        cast_list = []
        for col in final_cols:
            if col in integer_columns:
                cast_list.append(f"CAST({quote_identifier(col)} AS INT)")
            else:
                cast_list.append(f"{quote_identifier(col)}")

        # INSERT 쿼리: INSERT INTO mds_monthly_inventory_status(...) SELECT ... FROM 임시테이블
        quoted_cols = [quote_identifier(c) for c in final_cols]
        insert_cols_str = ",".join(quoted_cols)
        select_cols_str = ",".join(cast_list)

        final_insert_sql = f"""
            INSERT INTO mds_monthly_inventory_status ({insert_cols_str})
            SELECT {select_cols_str}
            FROM {temp_table}
        """
        cursor.execute(final_insert_sql)
        conn.commit()
        logging.info("임시테이블 → 최종테이블 삽입 완료")

        # ───────────── [G] exclude_item_codes 제거 ─────────────
        delete_excluded_query = """
            DELETE FROM mds_monthly_inventory_status
            WHERE item_code IN (SELECT item_code FROM exclude_item_codes)
              AND reference_month = %s
        """
        cursor.execute(delete_excluded_query, (year_month,))
        excluded_deleted_count = cursor.rowcount
        conn.commit()

        cursor.close()
        conn.close()

        uploaded_lines = len(insert_data)
        messagebox.showinfo(
            "업로드 완료",
            f"{year_month} 데이터 업로드 완료.\n"
            f"업로드된 행 수: {uploaded_lines}\n"
            f"제외된 품목코드 행 수: {excluded_deleted_count}"
        )
        logging.info(f"업로드 완료: {year_month} 데이터 업로드. 업로드된 행 수: {uploaded_lines}, 제외된 품목코드 행 수: {excluded_deleted_count}")

    except Exception as e:
        logging.error(f"오류 발생: {e}")
        if 'conn' in locals() and conn:
            conn.rollback()
            conn.close()
        messagebox.showerror("오류", f"데이터 업로드 중 오류가 발생했습니다:\n{e}")






def upload_account_substitution_data():
    """
    계정대체출고현황 데이터를 업로드하는 함수
    (특정 '연월' 접두어(TGOYYYYMM)에 해당하는 모든 출고번호를 초기화).
    """
    # 파일 경로 설정
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
    )
    if not file_path:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")
        return

    try:
        print("엑셀 파일을 읽는 중...")
        data = pd.read_excel(file_path, header=0, skiprows=[1], dtype=str)
        print("엑셀 파일 로드 완료.")

        # 필요한 컬럼만 선택
        required_columns = [
            '품목코드', '품목명', '규격', '양품출고량', '대체유형', '창고', '출고번호', '의뢰번호',
            '담당부서', '담당자', '거래처코드', '거래처명', '단가', '금액', '외화금액', '중량단위',
            '계정구분', '요청부서명', '헤더비고', '라인비고', '단위중량'
        ]
        data = data[required_columns]

        # 컬럼명 매핑
        data = data.rename(columns={
            '품목코드': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '양품출고량': 'quantity',
            '대체유형': 'substitution_type',
            '창고': 'warehouse',
            '출고번호': 'output_number',
            '의뢰번호': 'request_number',
            '담당부서': 'department',
            '담당자': 'manager',
            '거래처코드': 'customer_code',
            '거래처명': 'customer_name',
            '단가': 'unit_price',
            '금액': 'amount',
            '외화금액': 'foreign_currency_amount',
            '중량단위': 'weight_unit',
            '계정구분': 'account_type',
            '요청부서명': 'requesting_department',
            '헤더비고': 'header_note',
            '라인비고': 'line_note',
            '단위중량': 'unit_weight'
        })

        # 숫자형 컬럼 변환
        integer_columns = ['quantity', 'unit_price', 'amount', 'foreign_currency_amount', 'unit_weight']
        for col in integer_columns:
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0).astype(int)

        # 문자열 컬럼 변환
        string_columns = [col for col in data.columns if col not in integer_columns]
        for col in string_columns:
            data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})

        # ────────────────
        # ① 업로드 엑셀에서 'output_number'의 앞 9자리 추출 (TGOYYYYMM) → set으로 중복 제거
        # ────────────────
        monthly_prefixes = set()
        for val in data['output_number']:
            if val and len(val) >= 9:
                prefix = val[:9]  # 예: "TGO202501"
                monthly_prefixes.add(prefix)

        if not monthly_prefixes:
            raise ValueError("출고번호가 9자리 이상인 값이 없습니다. (TGOYYYYMM)")

        # ────────────────
        # ② 팝업으로 "해당 연월 데이터 초기화" 여부 확인
        #    예: "TGO202501" → "2025년 1월" 로 변환하여 안내
        # ────────────────
        user_friendly_list = []
        for prefix in monthly_prefixes:
            # 예: prefix = "TGO202501"
            if prefix.startswith("TGO") and len(prefix) == 9:
                year = prefix[3:7]  # "2025"
                month_str = prefix[7:9]  # "01"
                # int(month_str)로 01 -> 1 변환
                user_friendly_list.append(f"{year}년 {int(month_str)}월")
            else:
                # 혹시라도 형식이 다르면 그냥 prefix 그대로 표시
                user_friendly_list.append(prefix)

        user_friendly_text = ', '.join(user_friendly_list)
        confirm = messagebox.askyesno(
            "확인",
            f"업로드 파일 내에는 {user_friendly_text} 데이터가 포함되었습니다.\n"
            f"해당 월의 데이터를 초기화(삭제)하시겠습니까?"
        )
        if not confirm:
            messagebox.showinfo("취소", "업로드가 취소되었습니다.")
            return  # 함수 종료

        # DB 연결
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # 트랜잭션 시작
        conn.autocommit = False

        # ────────────────
        # ③ 삭제 쿼리: output_number의 앞 9자리가 monthly_prefixes 중 하나와 일치하면 삭제
        # ────────────────
        print("기존 출고번호 데이터 삭제 중...")
        monthly_prefixes_list = list(monthly_prefixes)
        delete_query = f"""
            DELETE FROM mds_account_substitution_output
            WHERE LEFT(output_number, 9) = ANY(%s)
        """
        cursor.execute(delete_query, (monthly_prefixes_list,))
        deleted_count = cursor.rowcount
        conn.commit()
        print(f"기존 데이터 {deleted_count}개 삭제 완료.")

        # ────────────────
        # ④ 임시 테이블 생성
        # ────────────────
        print("임시 테이블 생성 중...")
        temp_table_name = 'temp_account_substitution_output'
        cursor.execute(f"""
            DROP TABLE IF EXISTS {temp_table_name};
            CREATE TEMP TABLE {temp_table_name} AS
            SELECT * FROM mds_account_substitution_output LIMIT 0
        """)

        # 임시 테이블에 데이터 삽입 (1000개씩)
        print("임시 테이블에 데이터 삽입 중...")
        tuples = [tuple(x) for x in data.to_numpy()]
        cols = ','.join(data.columns)
        values_placeholder = ','.join(['%s'] * len(data.columns))
        total_rows = len(tuples)
        batch_size = 1000

        for i in range(0, total_rows, batch_size):
            batch_tuples = tuples[i:i + batch_size]
            cursor.executemany(
                f"INSERT INTO {temp_table_name} ({cols}) VALUES ({values_placeholder})",
                batch_tuples
            )
            conn.commit()
            print(f"임시 테이블: {i + len(batch_tuples)} / {total_rows} 행 삽입 완료")

        # 실제 테이블로 데이터 삽입
        print("실제 테이블로 데이터 삽입 중...")
        cursor.execute(f"""
            INSERT INTO mds_account_substitution_output (
                {cols}, last_updated
            )
            SELECT {cols}, NOW()
            FROM {temp_table_name}
        """)
        inserted_count = cursor.rowcount
        conn.commit()
        print(f"새로운 데이터 {inserted_count}개 삽입 완료.")

        # 제외 품목코드 삭제
        print("제외 품목코드 데이터 삭제 중...")
        cursor.execute("""
            DELETE FROM mds_account_substitution_output
            WHERE item_code IN (SELECT item_code FROM exclude_item_codes)
        """)
        excluded_deleted_count = cursor.rowcount
        conn.commit()
        print(f"제외 품목코드 {excluded_deleted_count}개 삭제 완료.")

        # 결과 알림
        messagebox.showinfo(
            "성공",
            f"데이터 업로드 완료\n"
            f"삭제된 기존 데이터: {deleted_count}\n"
            f"추가된 데이터: {total_rows - excluded_deleted_count}\n"
            f"제외된 데이터: {excluded_deleted_count}"
        )

    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"오류 발생: {e}")
        print(f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.autocommit = True
            conn.close()


def upload_account_substitution_input_data():
    """
    계정대체입고현황 데이터를 업로드하는 함수
    (TGIYYYYMM 기준으로 기존 데이터 삭제 후 새로 입력)
    """
    # 파일 경로 설정
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
    )
    if not file_path:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")
        return

    try:
        print("엑셀 파일을 읽는 중...")
        # 파일 읽기 (헤더 1라인, 2라인은 정보이므로 스킵)
        data = pd.read_excel(file_path, header=0, skiprows=[1], dtype=str)
        print("엑셀 파일 로드 완료.")

        # 필요한 컬럼만 선택
        required_columns = [
            '품목코드', '품목명', '규격', '입고량', '단가', '금액', '대체유형명',
            '창고', '입고번호', '의뢰번호', '담당부서', '담당자', '비고', '비고(라인)'
        ]
        data = data[required_columns]

        # 컬럼명 매핑
        data = data.rename(columns={
            '품목코드': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '입고량': 'quantity',
            '단가': 'unit_price',
            '금액': 'amount',
            '대체유형명': 'substitution_type',
            '창고': 'warehouse',
            '입고번호': 'input_number',
            '의뢰번호': 'request_number',
            '담당부서': 'department',
            '담당자': 'manager',
            '비고': 'header_note',
            '비고(라인)': 'line_note'
        })

        # 숫자형 컬럼 변환
        integer_columns = ['quantity', 'unit_price', 'amount']
        for col in integer_columns:
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0).astype(int)

        # 문자열 컬럼 변환
        string_columns = [col for col in data.columns if col not in integer_columns]
        for col in string_columns:
            data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})

        # ─────────────────────────────────────────
        # [1] 업로드된 엑셀 'input_number'에서 앞 9자리를 추출 (TGIYYYYMM)
        # ─────────────────────────────────────────
        monthly_prefixes = set()
        for val in data['input_number']:
            if val and len(val) >= 9:
                prefix = val[:9]  # 예: "TGI202501"
                monthly_prefixes.add(prefix)

        if not monthly_prefixes:
            raise ValueError("입고번호(최소 9자리 이상)가 없습니다. (예: TGI202501XXXX)")

        # ─────────────────────────────────────────
        # [2] 팝업으로 사용자에게 삭제 대상 안내 + 진행 여부 확인
        #     예) TGI202501 → 2025년 1월
        # ─────────────────────────────────────────
        user_friendly_list = []
        for prefix in monthly_prefixes:
            # "TGI202501" 형태인지 검사 후 변환
            if prefix.startswith("TGI") and len(prefix) == 9:
                year = prefix[3:7]   # "2025"
                month_str = prefix[7:9]  # "01"
                user_friendly_list.append(f"{year}년 {int(month_str)}월")
            else:
                user_friendly_list.append(prefix)

        user_friendly_text = ', '.join(user_friendly_list)
        confirm = messagebox.askyesno(
            "확인",
            f"업로드 파일에는 {user_friendly_text} 데이터가 포함되었습니다.\n"
            f"이 월(들)의 기존 데이터를 초기화(삭제)하시겠습니까?"
        )
        if not confirm:
            messagebox.showinfo("취소", "업로드가 취소되었습니다.")
            return

        # DB 연결
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # 트랜잭션 시작
        conn.autocommit = False

        # ─────────────────────────────────────────
        # [3] 기존 데이터 삭제 (앞 9자리가 monthly_prefixes 중 하나인 것)
        # ─────────────────────────────────────────
        print("기존 입고번호 데이터 삭제 중...")
        monthly_prefixes_list = list(monthly_prefixes)  # tuple, list 변환
        delete_query = """
            DELETE FROM mds_account_substitution_input
            WHERE LEFT(input_number, 9) = ANY(%s)
        """
        cursor.execute(delete_query, (monthly_prefixes_list,))
        deleted_count = cursor.rowcount
        conn.commit()
        print(f"기존 데이터 {deleted_count}개 삭제 완료.")

        # ─────────────────────────────────────────
        # [4] 임시 테이블 생성
        # ─────────────────────────────────────────
        temp_table_name = 'temp_account_substitution_input'
        print("임시 테이블 초기화 중...")
        cursor.execute(f"""
            DROP TABLE IF EXISTS {temp_table_name};
            CREATE TEMP TABLE {temp_table_name} AS
            SELECT * FROM mds_account_substitution_input LIMIT 0;
        """)
        print("임시 테이블 생성 완료.")

        # 임시 테이블에 데이터 삽입 (1000개씩)
        print("임시 테이블에 데이터 삽입 중...")
        tuples = [tuple(x) for x in data.to_numpy()]
        cols = ','.join(data.columns)
        values_placeholder = ','.join(['%s'] * len(data.columns))
        insert_query_temp = f"INSERT INTO {temp_table_name} ({cols}) VALUES ({values_placeholder})"

        total_rows = len(tuples)
        batch_size = 1000
        for i in range(0, total_rows, batch_size):
            batch_tuples = tuples[i:i + batch_size]
            cursor.executemany(insert_query_temp, batch_tuples)
            conn.commit()
            print(f"임시 테이블: {i + len(batch_tuples)} / {total_rows} 행 삽입 완료")
        print("임시 테이블에 데이터 삽입 완료.")

        # ─────────────────────────────────────────
        # [5] 임시 테이블 → 실제 테이블 삽입
        # ─────────────────────────────────────────
        print("새로운 데이터 삽입 중...")
        insert_query_main = f"""
            INSERT INTO mds_account_substitution_input (
                {cols}, last_updated
            )
            SELECT {cols}, NOW()
            FROM {temp_table_name}
        """
        cursor.execute(insert_query_main)
        conn.commit()
        inserted_count = cursor.rowcount
        print(f"새로운 데이터 {inserted_count}개 삽입 완료.")

        # 제외 품목코드 삭제
        print("제외 품목코드 데이터 삭제 중...")
        delete_excluded_query = """
            DELETE FROM mds_account_substitution_input
            WHERE item_code IN (SELECT item_code FROM exclude_item_codes)
        """
        cursor.execute(delete_excluded_query)
        excluded_deleted_count = cursor.rowcount
        conn.commit()
        print(f"제외 품목코드 {excluded_deleted_count}개 삭제 완료.")

        # 결과 메시지
        messagebox.showinfo(
            "성공",
            f"계정대체입고현황 데이터가 성공적으로 처리되었습니다.\n"
            f"삭제된 기존 데이터: {deleted_count}\n"
            f"신규 추가된 레코드 수: {total_rows - excluded_deleted_count}\n"
            f"제외된 데이터(품목) 수: {excluded_deleted_count}"
        )

    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"오류 발생: {e}")
        print(f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.autocommit = True
            conn.close()




def upload_shipment_status_data():
    """
    [출하현황 데이터를 업로드하는 함수 - COPY + 임시테이블 + 단일 commit]

    1) 기준월 입력
    2) 엑셀 파일 선택
       - 1행: 컬럼 헤더
       - 2행: 합계 (skiprows=[1]) 
       - 3행 이후: 실제 데이터
    3) 'mds_shipment_status' 테이블에 최종 저장
       - 임시테이블(temp_shipment_upload)에 COPY로 적재 → INSERT SELECT
    4) 기존 월 데이터가 있으면 삭제
    5) 제외 품목( exclude_item_codes ) 마지막에 삭제
    6) 단일 트랜잭션(중간 commit 없음), 마지막에 한 번만 commit
    """

    # (1) 기준월 입력
    current_year_month = datetime.now().strftime("%Y/%m")
    year_month = simpledialog.askstring(
        "기준월 선택", 
        "년/월을 입력하세요 (예: 2025/01):",
        initialvalue=current_year_month
    )
    if not year_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    # (2) Excel 파일 선택
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        print("파일 선택이 취소되었습니다.")
        return

    conn = None
    cursor = None
    csv_path = None

    try:
        print("엑셀 파일을 읽는 중...")
        # 1행=헤더, 2행=합계, skiprows=[1], dtype=str
        data = pd.read_excel(file_path, sheet_name=0, skiprows=[1], dtype=str)
        print(f"엑셀 파일 로드 완료. 총 데이터 행 수: {len(data)}")

        # 필요한 컬럼
        required_columns = [
            '품목코드', '품목명', '규격', '출하수량', '단가', '금액', '원화금액(출하)',
            '부가세(출하)', '원화금액(매출)', '부가세(매출)', '총금액(출하)', '총금액(매출)', '중량'
        ]
        # 필수 컬럼 체크
        data = data[required_columns]

        # 컬럼명 매핑
        data.rename(columns={
            '품목코드': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '출하수량': 'shipment_quantity',
            '단가': 'unit_price',
            '금액': 'amount',
            '원화금액(출하)': 'won_amount_shipment',
            '부가세(출하)': 'vat_shipment',
            '원화금액(매출)': 'won_amount_sales',
            '부가세(매출)': 'vat_sales',
            '총금액(출하)': 'total_amount_shipment',
            '총금액(매출)': 'total_amount_sales',
            '중량': 'weight'
        }, inplace=True)

        # 숫자 변환
        integer_columns = [
            'shipment_quantity', 'unit_price', 'amount', 'won_amount_shipment', 
            'vat_shipment', 'won_amount_sales', 'vat_sales', 
            'total_amount_shipment', 'total_amount_sales', 'weight'
        ]
        for column in integer_columns:
            data[column] = data[column].str.replace(',', '', regex=True).replace('', '0')
            data[column] = pd.to_numeric(data[column], errors='coerce').fillna(0).astype(int)

        # 문자열 컬럼
        string_columns = ['item_code', 'item_name', 'specification']
        for col in string_columns:
            data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})

        # 기준월 컬럼 추가
        data['reference_month'] = year_month

        # DB 연결 + 단일 트랜잭션
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # 기존 동일 월 데이터 삭제
        cursor.execute("""
            SELECT COUNT(*) FROM mds_shipment_status
            WHERE reference_month = %s
        """, (year_month,))
        existing_data_count = cursor.fetchone()[0]
        if existing_data_count > 0:
            ans = messagebox.askyesno(
                "기존 데이터 확인",
                f"기준월 {year_month}의 데이터가 이미 {existing_data_count}건 존재합니다.\n삭제 후 업로드 하시겠습니까?"
            )
            if ans:
                cursor.execute("""
                    DELETE FROM mds_shipment_status
                    WHERE reference_month = %s
                """, (year_month,))
                print(f"기존 데이터 {year_month} 초기화 완료 ({existing_data_count}건 삭제).")
            else:
                print("업로드 취소되었습니다.")
                return

        # 임시 CSV 파일 생성
        tmp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8')
        csv_path = tmp_file.name
        data.to_csv(tmp_file, index=False, header=False)  # COPY 시 헤더 불필요
        tmp_file.close()

        # 임시테이블 (TEMP TABLE + ON COMMIT DROP)
        temp_table = "temp_shipment_upload"  # 함수명과 유사한 임시테이블명
        cursor.execute(f"""
            CREATE TEMP TABLE {temp_table} (
                item_code VARCHAR(255),
                item_name VARCHAR(255),
                specification VARCHAR(100),
                shipment_quantity INT,
                unit_price INT,
                amount INT,
                won_amount_shipment INT,
                vat_shipment INT,
                won_amount_sales INT,
                vat_sales INT,
                total_amount_shipment INT,
                total_amount_sales INT,
                weight INT,
                reference_month VARCHAR(7)
            ) ON COMMIT DROP
        """)
        print("임시테이블 생성 완료")

        # COPY FROM CSV → 임시테이블
        with open(csv_path, 'r', encoding='utf-8') as f:
            cursor.copy_expert(f"""
                COPY {temp_table}
                FROM STDIN
                WITH (FORMAT CSV, DELIMITER ',', NULL '')
            """, f)
        print("임시테이블 COPY 완료")

        # CSV 임시파일 삭제
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)

        # 임시테이블 → 최종 테이블
        cursor.execute(f"""
            INSERT INTO mds_shipment_status (
                item_code, item_name, specification, shipment_quantity,
                unit_price, amount, won_amount_shipment, vat_shipment,
                won_amount_sales, vat_sales, total_amount_shipment, 
                total_amount_sales, weight, reference_month
            )
            SELECT
                item_code, item_name, specification, shipment_quantity,
                unit_price, amount, won_amount_shipment, vat_shipment,
                won_amount_sales, vat_sales, total_amount_shipment,
                total_amount_sales, weight, reference_month
            FROM {temp_table}
        """)
        inserted_count = cursor.rowcount
        print(f"최종 테이블에 {inserted_count}행 INSERT 예정.")

        # 제외 품목 삭제
        delete_excluded_query = """
            DELETE FROM mds_shipment_status
            WHERE item_code IN (SELECT item_code FROM exclude_item_codes)
              AND reference_month = %s
        """
        cursor.execute(delete_excluded_query, (year_month,))
        excluded_deleted_count = cursor.rowcount
        print(f"제외 품목코드 {excluded_deleted_count}행 삭제 예정.")

        # 마지막 한 번만 commit
        conn.commit()
        print("커밋 완료. 임시테이블 삭제됨.")

        messagebox.showinfo(
            "업로드 완료",
            f"{year_month} 데이터 업로드 완료.\n"
            f"업로드된 행 수: {inserted_count}\n"
            f"제외된 품목코드 행 수: {excluded_deleted_count}"
        )
        print(f"업로드 완료: {year_month} 데이터 {inserted_count}행 업로드, {excluded_deleted_count}행 제외.")

    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"데이터 업로드 중 오류가 발생했습니다: {e}")
        print(f"오류 발생: {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)




def upload_purchase_receipt_status_data():
    """
    [구매입고현황(새 xls) 업로드 함수 - 임시테이블 + copy + 단일 commit 최종본]
    
    1) 기준월 입력
    2) 엑셀 파일 선택 (새 구조)
       - 1행: 컬럼 헤더(한글)
       - 2행: 합계(무시, skiprows=[1])
       - 3행 이후: 실제 데이터
    3) DB: mds_purchase_receipt_status 테이블 최종 저장
       - 임시테이블(temp_purchase_upload)에 COPY로 적재 → INSERT SELECT
    4) exclude_item_codes 테이블에 있는 item_code는 최종 업로드에서 제외
    5) 모든 단계를 **하나의 트랜잭션**으로 진행 (중간 commit 없음)
    6) 마지막에 commit
    """
    # (1) 기준월 입력
    current_year_month = datetime.now().strftime("%Y/%m")
    year_month = simpledialog.askstring(
        "기준월 선택",
        "년/월을 입력하세요 (예: 2025/01):",
        initialvalue=current_year_month
    )
    if not year_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    # (2) Excel 파일 선택
    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        print("파일 선택이 취소되었습니다.")
        return

    conn = None
    cursor = None
    csv_path = None

    try:
        print("엑셀 파일 로드 중...")
        # 1행=헤더, 2행=합계(무시), 3행부터 실제 데이터
        data = pd.read_excel(
            file_path,
            sheet_name=0,
            skiprows=[1],  # 2행(합계) 제거
            header=0,
            dtype=str
        )
        print(f"엑셀 파일 로드 완료. 총 데이터 행 수: {len(data)}")

        # (A) 필수 한글 컬럼
        required_columns = [
            '품목코드', '품목명', '규격', '부가세', '총금액',
            '관리수량', '원화금액', '거래처명', '거래처코드'
        ]
        for col in required_columns:
            if col not in data.columns:
                messagebox.showerror("오류", f"필수 컬럼 '{col}'이 엑셀에 없습니다.")
                return

        data = data[required_columns]

        # (B) 한글→영문 컬럼명 매핑
        data.rename(columns={
            '품목코드': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '부가세': 'vat',
            '총금액': 'total_amount',
            '관리수량': 'management_quantity',
            '원화금액': 'won_amount',
            '거래처명': 'supplier_name',
            '거래처코드': 'supplier_code'
        }, inplace=True)

        # (C) 숫자 변환
        numeric_cols = ['vat','total_amount','management_quantity','won_amount']
        for col in numeric_cols:
            data[col] = data[col].str.replace(',', '', regex=True).fillna('0')
            data[col] = pd.to_numeric(data[col], errors='coerce').fillna(0).astype(int)

        # (D) 문자열 컬럼: None 처리
        str_cols = ['item_code','item_name','specification','supplier_name','supplier_code']
        for col in str_cols:
            data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})

        # (E) 기준월 컬럼
        data['reference_month'] = year_month

        # (F) DB 연결 (하나의 트랜잭션으로)
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # (G) 기존 데이터 확인 & 삭제
        cursor.execute("""
            SELECT COUNT(*)
            FROM mds_purchase_receipt_status
            WHERE reference_month = %s
        """, (year_month,))
        existing_count = cursor.fetchone()[0]
        if existing_count > 0:
            ans = messagebox.askyesno(
                "기존 데이터 확인",
                f"{year_month} 데이터가 이미 {existing_count}행 존재합니다.\n삭제 후 업로드 하시겠습니까?"
            )
            if ans:
                cursor.execute("""
                    DELETE FROM mds_purchase_receipt_status
                    WHERE reference_month = %s
                """, (year_month,))
                print(f"{year_month} 기존 데이터 {existing_count}행 삭제 완료.")
            else:
                print("업로드 취소됨.")
                return

        # (H) 임시 CSV파일 생성
        tmp_file = tempfile.NamedTemporaryFile(
            mode='w',
            suffix='.csv',
            delete=False,
            newline='',
            encoding='utf-8'
        )
        csv_path = tmp_file.name
        data.to_csv(tmp_file, index=False, header=False)  # COPY 시 헤더 불필요
        tmp_file.close()

        # (I) 임시테이블 생성 (TEMP TABLE + ON COMMIT DROP)
        temp_table = "temp_purchase_upload"  # ← 임시테이블명
        cursor.execute(f"""
            CREATE TEMP TABLE {temp_table} (
                item_code VARCHAR(255),
                item_name VARCHAR(255),
                specification VARCHAR(255),
                vat INT,
                total_amount INT,
                management_quantity INT,
                won_amount INT,
                supplier_name VARCHAR(255),
                supplier_code VARCHAR(255),
                reference_month VARCHAR(7)
            ) ON COMMIT DROP
        """)
        print("임시테이블 생성 완료")

        # (J) COPY FROM CSV → 임시테이블
        with open(csv_path, 'r', encoding='utf-8') as f:
            cursor.copy_expert(f"""
                COPY {temp_table}
                FROM STDIN
                WITH (FORMAT CSV, DELIMITER ',', NULL '')
            """, f)
        print("임시테이블 COPY 완료")

        # (K) 임시파일 삭제
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)

        # (L) 임시테이블 → 최종 테이블(mds_purchase_receipt_status)
        #     여기서 exclude_item_codes 에 있는 item_code는 제외
        cursor.execute(f"""
            INSERT INTO mds_purchase_receipt_status (
                item_code, item_name, specification, vat,
                total_amount, management_quantity, won_amount,
                supplier_name, supplier_code, reference_month
            )
            SELECT
                tmp.item_code,
                tmp.item_name,
                tmp.specification,
                tmp.vat,
                tmp.total_amount,
                tmp.management_quantity,
                tmp.won_amount,
                tmp.supplier_name,
                tmp.supplier_code,
                tmp.reference_month
            FROM {temp_table} AS tmp
            WHERE tmp.item_code NOT IN (
                SELECT item_code FROM exclude_item_codes
            )
        """)
        inserted_count = cursor.rowcount
        print(f"최종 테이블에 {inserted_count}건 INSERT 예정.")

        # (M) 마지막 한 번만 commit
        conn.commit()
        print("커밋 완료. 임시테이블 삭제됨.")

        messagebox.showinfo(
            "완료",
            f"{year_month} 업로드 완료.\n총 {inserted_count}행 업로드됨.\n"
            "exclude_item_codes에 포함된 item_code는 제외되었습니다."
        )

    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"업로드 중 오류 발생: {e}")
        print(f"오류 발생: {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        # 혹시 CSV파일 남아있으면 제거
        if csv_path and os.path.exists(csv_path):
            os.remove(csv_path)









def upload_inventory_evaluation_data():
    """
    재고평가 데이터를 업로드하는 함수
    """
    # 기준월 입력
    current_year_month = datetime.now().strftime("%Y/%m")  # 기본값은 현재 년/월
    year_month = simpledialog.askstring("기준월 선택", "년/월을 입력하세요 :", initialvalue=current_year_month)
    
    if not year_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    # Excel 파일 선택
    directory = 'C:/ERPUExport'  # 필요에 따라 경로 수정
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        print("파일 선택이 취소되었습니다.")
        return

    try:
        print("엑셀 파일을 읽는 중...")
        # Excel 파일 읽기 (첫 번째 시트의 모든 데이터를 가져오며 두 번째 라인을 건너뜀)
        data = pd.read_excel(file_path, sheet_name=0, skiprows=[1], dtype=str)
        print("엑셀 파일 로드 완료.")
        print(f"총 데이터 행 수 (헤더 제외): {len(data)}")

        # 마지막 라인 (총계 라인) 제거
        data = data.iloc[:-1]
        print(f"총계 라인 제거 후 데이터 행 수: {len(data)}")

        # 필요한 컬럼만 선택
        required_columns = [
            '품목', '품목명', '규격', '기초수량', '기초단가', '기초금액',
            '입고수량', '입고금액', '대체수량', '대체금액', '출고수량',
            '출고금액', '재고수량', '재고단가', '재고금액'
        ]
        data = data[required_columns]
        print("필요한 컬럼 선택 완료.")

        # 컬럼명 매핑
        data = data.rename(columns={
            '품목': 'item_code',
            '품목명': 'item_name',
            '규격': 'specification',
            '기초수량': 'beginning_quantity',
            '기초단가': 'beginning_unit_price',
            '기초금액': 'beginning_amount',
            '입고수량': 'receipt_quantity',
            '입고금액': 'receipt_amount',
            '대체수량': 'substitution_quantity',
            '대체금액': 'substitution_amount',
            '출고수량': 'shipment_quantity',
            '출고금액': 'shipment_amount',
            '재고수량': 'inventory_quantity',
            '재고단가': 'inventory_unit_price',
            '재고금액': 'inventory_amount'
        })
        print("컬럼명 매핑 완료.")

        # 쉼표 및 통화 형식 제거 후 데이터 타입 변환
        integer_columns = [
            'beginning_quantity', 'beginning_unit_price', 'beginning_amount',
            'receipt_quantity', 'receipt_amount', 'substitution_quantity',
            'substitution_amount', 'shipment_quantity', 'shipment_amount',
            'inventory_quantity', 'inventory_unit_price', 'inventory_amount'
        ]
        for column in integer_columns:
            # 쉼표 제거 및 숫자 변환, 소수점 유지
            data[column] = data[column].astype(str).str.replace(',', '').replace('', '0')
            data[column] = pd.to_numeric(data[column], errors='coerce').fillna(0).astype(float)  # int -> float
        print("숫자 컬럼 데이터 타입 변환 완료.")


        # 문자열 컬럼의 결측치 처리
        string_columns = ['item_code', 'item_name', 'specification']
        for col in string_columns:
            data[col] = data[col].astype(str).replace({'nan': None, 'NaN': None, 'None': None})
        print("문자열 컬럼 결측치 처리 완료.")

        # 기준월 컬럼 추가
        data['reference_month'] = year_month
        print("기준월 컬럼 추가 완료.")

        # DB 연결
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()
        print("데이터베이스 연결 성공.")

        # 같은 기준월의 데이터가 이미 존재하는지 확인
        cursor.execute('SELECT COUNT(*) FROM mds_inventory_evaluation WHERE reference_month = %s', (year_month,))
        existing_data_count = cursor.fetchone()[0]

        if existing_data_count > 0:
            # 이미 데이터가 존재할 때 덮어쓸지 여부 확인
            if messagebox.askyesno("기존 데이터 확인", f"기준월 {year_month}의 데이터가 이미 존재합니다. 초기화하고 업로드 하시겠습니까?"):
                cursor.execute('DELETE FROM mds_inventory_evaluation WHERE reference_month = %s', (year_month,))
                conn.commit()
                print(f"기존 데이터 {year_month}가 초기화되었습니다.")
            else:
                messagebox.showinfo("업로드 취소", "업로드가 취소되었습니다.")
                print("업로드가 취소되었습니다.")
                return

        # 데이터 삽입 시작
        print("데이터베이스에 데이터 삽입 중...")
        tuples = [tuple(x) for x in data.to_numpy()]
        cols = ','.join(data.columns)
        values_placeholder = ','.join(['%s'] * len(data.columns))
        insert_query = f"INSERT INTO mds_inventory_evaluation ({cols}) VALUES ({values_placeholder})"

        total_rows = len(tuples)
        batch_size = 1000  # 한 번에 처리할 행의 수
        for i in range(0, total_rows, batch_size):
            batch_tuples = tuples[i:i+batch_size]
            cursor.executemany(insert_query, batch_tuples)
            conn.commit()
            print(f"{i + len(batch_tuples)} / {total_rows} 행 처리 완료.")

        print("데이터베이스에 데이터 삽입 완료.")

        # 마지막 단계에서 제외할 품목 코드를 가진 행 삭제
        print("데이터베이스에서 제외할 품목 코드를 가진 행 삭제 중...")
        delete_excluded_query = """
            DELETE FROM mds_inventory_evaluation
            WHERE item_code IN (SELECT item_code FROM exclude_item_codes)
            AND reference_month = %s
        """
        cursor.execute(delete_excluded_query, (year_month,))
        excluded_deleted_count = cursor.rowcount
        print(f"{excluded_deleted_count}개의 제외할 품목코드 레코드 삭제됨.")

        conn.commit()
        cursor.close()
        conn.close()
        print("데이터베이스 연결 종료.")

        # 업로드된 라인 수 계산
        uploaded_lines = len(data)
        messagebox.showinfo(
            "업로드 완료",
            f"{year_month} 데이터 업로드 완료.\n"
            f"업로드된 라인 수: {uploaded_lines}\n"
            f"제외된 품목코드 레코드 수: {excluded_deleted_count}"
        )
        print(f"업로드 완료: {year_month} 데이터 업로드 완료되었습니다. 업로드된 라인 수: {uploaded_lines}, 제외된 품목코드 레코드 수: {excluded_deleted_count}")

    except Exception as e:
        messagebox.showerror("오류", f"데이터 업로드 중 오류가 발생했습니다: {e}")
        print(f"오류 발생: {e}")




def upload_basic_data():
    """엑셀 기초데이터를 PostgreSQL로 고속 업로드 (execute_values 사용)"""
    global root

    # ───────────── 1. 기준년 입력 ─────────────
    year = simpledialog.askstring("기준년 입력", "기준년을 입력하세요 (예: 2024):", parent=root)
    if not year:
        messagebox.showwarning("경고", "기준년을 입력해야 합니다.")
        return

    # ───────────── 2. 파일 선택 ─────────────
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        return

    # ───────────── 3. DB 연결 ─────────────
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # ───────────── 4. 기존 연도 데이터 삭제 여부 ─────────────
        cursor.execute("SELECT COUNT(*) FROM mds_basic_data WHERE reference_year = %s", (year,))
        if cursor.fetchone()[0] > 0:
            if not messagebox.askyesno("확인", f"{year}년도 데이터를 삭제하고 다시 업로드하시겠습니까?"):
                messagebox.showinfo("정보", "업로드를 취소합니다.")
                return
            cursor.execute("DELETE FROM mds_basic_data WHERE reference_year = %s", (year,))
            conn.commit()
            messagebox.showinfo("정보", f"{year}년도 기존 데이터를 삭제했습니다.")

        # ───────────── 5. 엑셀 로드 & 전처리 ─────────────
        df = pd.read_excel(file_path, header=[0, 1])
        df.columns = pd.MultiIndex.from_tuples([
            (str(c[0]).strip() if c[0] else "", str(c[1]).strip() if "Unnamed" not in str(c[1]) else "")
            for c in df.columns
        ])
        df.columns = ["_".join(filter(None, c)) for c in df.columns]

        rename_map = {
            "품목코드": "item_code",
            "품명": "item_name",
            "규격": "specification",
            "단위": "unit",
            "분류": "category",
            "이월재고_단가": "beginning_unit_price",
            "이월재고_수량": "beginning_quantity",
            "이월재고_금액": "beginning_amount",
        }
        df = df.rename(columns=rename_map)

        required = [
            "item_code", "item_name", "specification", "unit", "category",
            "beginning_unit_price", "beginning_quantity", "beginning_amount",
        ]
        miss = [c for c in required if c not in df.columns]
        if miss:
            messagebox.showerror("에러", f"다음 컬럼이 누락되었습니다: {', '.join(miss)}")
            return

        # "합계" 이후 행 제거
        df = df.reset_index(drop=True)
        sum_idx = df.index[df["item_code"].astype(str).str.strip() == "합계"].tolist()
        if sum_idx:
            df = df.loc[: sum_idx[0] - 1]

        # 숫자 컬럼 정제 ➜ Decimal(소수 5자리)
        for col in ["beginning_unit_price", "beginning_quantity", "beginning_amount"]:
            df[col] = (
                pd.to_numeric(df[col], errors="coerce")
                .fillna(0)
                .round(5)
                .apply(lambda x: Decimal(f"{x:.5f}"))
            )

        df["reference_year"] = year

        # ───────────── 6. execute_values 고속 업로드 ─────────────
        data_cols = [
            "item_code",
            "item_name",
            "specification",
            "unit",
            "category",
            "beginning_unit_price",
            "beginning_quantity",
            "beginning_amount",
            "reference_year",
        ]
        data = list(df[data_cols].itertuples(index=False, name=None))

        insert_query = (
            "INSERT INTO mds_basic_data (" + ", ".join(data_cols) + ") VALUES %s"
        )
        execute_values(cursor, insert_query, data)

        conn.commit()
        messagebox.showinfo("성공", f"업로드 완료: {len(df)}개 행이 업로드되었습니다.")

    except Exception as e:
        conn.rollback()
        messagebox.showerror("에러", f"업로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



def upload_account_ledger_data():
    """
    계정별원장(mds_account_ledger) 데이터를 업로드하는 함수.
    - '계정명'이 '판매장려수익'인 행만 남김
    - '적요'에 '재고조정' 포함된 행 제외
    - '회계일자'가 '전월이월','월계','누계'인 행 제외
    - 동일 회계월 데이터가 있다면 삭제 후 업로드
    - 새로 추가한 '거래처코드' 컬럼도 함께 INSERT
    """

    directory = 'C:/ERPUExport'
    file_path = filedialog.askopenfilename(
        initialdir=directory,
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        print("파일 선택이 취소되었습니다.")
        return

    try:
        print("엑셀 파일을 읽는 중...")
        data = pd.read_excel(file_path, sheet_name=0, dtype=str)
        print("엑셀 파일 로드 완료.")
        print(f"총 데이터 행 수: {len(data)}")

        # (1) 필수 컬럼 체크
        required_columns = ['회계일자', '적요', '거래처', '대변', 'C/C', '계정명', '거래처코드']
        for col in required_columns:
            if col not in data.columns:
                messagebox.showerror("오류", f"필수 컬럼 '{col}'이 누락되었습니다.")
                return

        # (2) 특정 값 제외 로직
        exclude_dates = ['전월이월', '월계', '누계']
        data = data[~data['회계일자'].isin(exclude_dates)]

        # '계정명'이 '판매장려수익'인 행만 남김
        data = data[data['계정명'] == '판매장려수익']

        # '적요'에 '재고조정' 포함 행 제외
        data = data[~data['적요'].astype(str).str.contains('재고조정', na=False)]

        # '대변' 숫자 변환
        data['대변'] = data['대변'].str.replace(',', '', regex=True)
        data['대변'] = pd.to_numeric(data['대변'], errors='coerce').fillna(0)

        # (3) 회계일자에서 년/월 추출 함수
        def get_year_month(date_str):
            if isinstance(date_str, str) and len(date_str) >= 7:
                return date_str[:7]  # 예) "2023-05"
            return None

        # 업로드 대상 (년-월) 목록
        year_months_in_file = data['회계일자'].dropna().apply(get_year_month).unique()
        if len(year_months_in_file) == 0:
            messagebox.showerror("오류", "유효한 회계일자 데이터가 없습니다.")
            return

        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # (4) 월별로 기존 데이터 삭제 + 업로드
        for ym in year_months_in_file:
            cursor.execute("""
                SELECT COUNT(*) 
                FROM mds_account_ledger 
                WHERE 회계일자 LIKE %s||'%%'
            """, (ym,))
            existing_count = cursor.fetchone()[0]
            if existing_count > 0:
                ans = messagebox.askyesno("데이터 존재 확인", f"{ym}월의 기존 데이터가 있습니다. 삭제하고 진행하시겠습니까?")
                if ans:
                    cursor.execute("""
                        DELETE FROM mds_account_ledger 
                        WHERE 회계일자 LIKE %s||'%%'
                    """, (ym,))
                    conn.commit()
                    print(f"{ym}월 데이터 삭제 완료.")
                else:
                    messagebox.showinfo("취소", f"{ym}월 데이터 저장이 취소되었습니다.")
                    return

        # (5) INSERT 시 '거래처코드' 컬럼 추가
        insert_query = """
            INSERT INTO mds_account_ledger (
                회계일자, 적요, 거래처, 대변, "C/C", 계정명, 거래처코드
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        # (6) data_to_insert 에 '거래처코드' 포함
        data_to_insert = data[
            ['회계일자', '적요', '거래처', '대변', 'C/C', '계정명', '거래처코드']
        ].values.tolist()

        batch_size = 1000
        total_rows = len(data_to_insert)
        for i in range(0, total_rows, batch_size):
            batch = data_to_insert[i:i+batch_size]
            cursor.executemany(insert_query, batch)
            conn.commit()
            print(f"{i + len(batch)}/{total_rows} 업로드 완료.")

        messagebox.showinfo("완료", f"데이터 업로드 완료. 총 {total_rows}행 업로드됨.")
        cursor.close()
        conn.close()

    except Exception as e:
        messagebox.showerror("오류", f"데이터 업로드 중 오류가 발생했습니다: {e}")
        print(f"오류 발생: {e}")










def open_exclude_item_codes_window():
    window = tk.Toplevel(root)
    window.title("비수불 계정관리 - 제외할 품목코드 리스트")
    window.geometry("400x400")

    # 데이터베이스에서 코드 목록 가져오기
    conn = get_postgres_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT item_code FROM exclude_item_codes ORDER BY item_code")
    codes = [row[0] for row in cursor.fetchall()]
    cursor.close()
    conn.close()

    # 리스트박스 생성
    listbox = tk.Listbox(window)
    for code in codes:
        listbox.insert(tk.END, code)
    listbox.pack(fill='both', expand=True)

    # 추가, 수정, 삭제 버튼 생성
    add_button = ttk.Button(window, text="추가", command=lambda: add_code(listbox))
    edit_button = ttk.Button(window, text="수정", command=lambda: edit_code(listbox))
    delete_button = ttk.Button(window, text="삭제", command=lambda: delete_code(listbox))
    add_button.pack(side='left', padx=5, pady=5)
    edit_button.pack(side='left', padx=5, pady=5)
    delete_button.pack(side='left', padx=5, pady=5)

    # 저장, 취소 버튼 생성
    save_button = ttk.Button(window, text="저장", command=lambda: save_codes(listbox))
    cancel_button = ttk.Button(window, text="취소", command=window.destroy)
    save_button.pack(side='left', padx=5, pady=5)
    cancel_button.pack(side='left', padx=5, pady=5)


def reset_treeview_state(treeview):
    """
    트리뷰의 상태와 스크롤을 초기화하는 함수
    """
    treeview.yview_moveto(0)
    treeview.xview_moveto(0)

# 트리뷰 표시 함수
def show_treeview(show_frame, hide_frames, treeview_name, date_frame=None):
    global current_treeview
    global date_frame_treeview1, date_frame_treeview2, date_frame_treeview3
    global date_frame_treeview4, date_frame_treeview5, date_frame_treeview6, date_frame_treeview7
    global total_frame_treeview2, total_frame_treeview3, total_frame_treeview4
    global total_frame_treeview5, total_frame_treeview6, total_frame_treeview7, total_frame_treeview8

    # ★ 라벨맵: 트리뷰 이름(식별자) -> 화면에 보여줄 라벨 텍스트
    label_map = {
        'treeview0': '[0] 월별수불현황',
        'treeview1': '[1] 창고별재고현황',
        'treeview2': '[2] 계정대체출고현황',
        'treeview3': '[3] 계정대체입고현황',
        'treeview4': '[4] 출하현황',
        'treeview5': '[5] 구매입고현황',
        'treeview6': '[6] 재고평가',
        'treeview7': '[7] 기초데이터',
        'treeview8': '[8] 장려금',
    }

    # 다른 트리뷰 프레임 숨기기
    for frame in hide_frames:
        frame.pack_forget()

        # 날짜 프레임과 총계 프레임도 숨기기
        if frame == treeview2_frame:
            date_frame_treeview2.pack_forget()
            total_frame_treeview2.pack_forget()
        elif frame == treeview3_frame:
            date_frame_treeview3.pack_forget()
            total_frame_treeview3.pack_forget()
        elif frame == treeview4_frame:
            date_frame_treeview4.pack_forget()
            total_frame_treeview4.pack_forget()
        elif frame == treeview5_frame:
            date_frame_treeview5.pack_forget()
            total_frame_treeview5.pack_forget()
        elif frame == treeview6_frame:
            date_frame_treeview6.pack_forget()
            total_frame_treeview6.pack_forget()
        elif frame == treeview7_frame:
            date_frame_treeview7.pack_forget()
            total_frame_treeview7.pack_forget()
        elif frame == treeview8_frame:
            # 트리뷰8은 날짜 프레임이 없으므로 총계 프레임만 숨김
            total_frame_treeview8.pack_forget()
        elif frame == treeview1_frame:
            date_frame_treeview1.pack_forget()
            # 트리뷰1 total_frame이 있다면 여기도 pack_forget() 해주면 됨

    # 선택한 트리뷰 프레임 표시
    show_frame.pack(fill='both', expand=True)

    # 선택된 트리뷰의 날짜 프레임 표시
    if date_frame:
        date_frame.pack(side="top", fill='x', padx=5, pady=5)

    # Treeview0과 Treeview8의 총계 프레임은 제외
    if treeview_name not in ['treeview0', 'treeview8']:
        total_frame_name = f"total_frame_{treeview_name}"
        total_frame = globals().get(total_frame_name)
        if total_frame:
            total_frame.pack(side="top", fill='x')

    # 현재 트리뷰를 갱신
    current_treeview = treeview_name

    # ★ 라벨맵에서 제목 가져오기
    view_title = label_map.get(treeview_name, treeview_name)  # 혹은 디폴트값 사용

    # ★ 라벨에 표시 (전역으로 만든 current_treeview_label이 있다고 가정)
    if current_treeview_label:
        current_treeview_label.config(text=f"현재 트리뷰: {view_title}")

    # 로그 출력
    logging.info(f"{treeview_name}으로 전환됨")




# ------------------------------------------------------------------------------
# **추가: 숫자 변환 함수** (콤마/공백 등 제거 후 float 변환 시도)
# ------------------------------------------------------------------------------
def parse_as_float(val):
    """
    문자열(또는 기타 값) val이 숫자로 변환 가능하면 float 값 반환.
    콤마(',', 공백 등) 무시. 변환 불가능하면 None 반환.
    """
    if val is None:
        return None
    s = str(val).strip()
    if s == '':
        return None
    s = s.replace(',', '')  # 콤마 제거
    try:
        return float(s)
    except ValueError:
        return None

# ------------------------------------------------------------------------------
# download_current_treeview 함수 (전체)
# ------------------------------------------------------------------------------

def download_current_treeview():
    if not current_treeview:
        messagebox.showwarning("Warning", "현재 표시된 트리뷰가 없습니다.")
        return

    downloads_folder = os.path.join(os.path.expanduser('~'), 'Downloads')
    now = datetime.now().strftime('%Y%m%d_%H%M%S')

    if current_treeview == 'treeview0':
        default_filename = f"월별수불현황_{now}.xlsx"
    elif current_treeview == 'treeview1':
        default_filename = f"창고별재고현황_{now}.xlsx"
    elif current_treeview == 'treeview2':
        default_filename = f"계정대체출고현황_{now}.xlsx"
    elif current_treeview == 'treeview3':
        default_filename = f"계정대체입고현황_{now}.xlsx"
    elif current_treeview == 'treeview4':
        default_filename = f"출하현황_{now}.xlsx"
    elif current_treeview == 'treeview5':
        default_filename = f"구매입고현황_{now}.xlsx"
    elif current_treeview == 'treeview6':
        default_filename = f"재고평가_{now}.xlsx"
    elif current_treeview == 'treeview7':
        default_filename = f"기초데이터_{now}.xlsx"
    elif current_treeview == 'treeview8':
        default_filename = f"장려금_{now}.xlsx"
    else:
        default_filename = f"data_{now}.xlsx"

    file_path = filedialog.asksaveasfilename(
        initialdir=downloads_folder,
        initialfile=default_filename,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not file_path:
        return

    treeview = globals()[current_treeview]

    if current_treeview == 'treeview0':
        # 여기서는 전역변수 columns( (col_id, col_text, col_width) 형태 )를 그대로 사용
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Font, Border, Side

            wb = Workbook()
            ws = wb.active

            global columns, multiheader_groups

            # 첫 번째 행: 고정 컬럼(품목코드, 품명, 규격, 단위, 분류) 처리
            col_idx = 1
            for idx, (col_id, col_text, _) in enumerate(columns, start=1):
                if idx <= 5:  # 품목코드, 품명, 규격, 단위, 분류
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = col_text
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    # 첫 번째 행부터 두 번째 행까지 병합
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    col_idx += 1
                else:  
                    # 고정 5개 컬럼 이후는 모두 멀티헤더 그룹 처리 부분에서 진행
                    break

            # 이후 멀티헤더 그룹 작성
            # multiheader_groups는 (group_name, group_columns) 형태
            for group_name, group_columns in multiheader_groups:
                num_columns = len(group_columns)
                cell = ws.cell(row=1, column=col_idx)
                cell.value = group_name
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)

                if num_columns > 1:
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + num_columns - 1)

                # 두 번째 행에 그룹에 속한 각 컬럼 헤더를 작성
                # (사용자 코드에 맞게 조정)
                for i, (col_id, col_text, _) in enumerate(columns[5:], start=6):
                    if col_id in group_columns:
                        cell2 = ws.cell(row=2, column=col_idx)
                        cell2.value = col_text
                        cell2.alignment = Alignment(horizontal='center', vertical='center')
                        cell2.font = Font(bold=True)
                        col_idx += 1

            # 데이터 작성 (3행부터)
            data = []
            for item in treeview.get_children():
                # 합계행인지 확인
                if 'totalrow' in treeview.item(item, 'tags'):
                    continue
                values = treeview.item(item)["values"]
                data.append(values)

            row_start = 3
            for r_idx, row_data in enumerate(data, start=row_start):
                for c_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)

                    # **숫자 변환 시도**
                    numeric_val = parse_as_float(value)
                    if numeric_val is not None:
                        cell.value = numeric_val
                        # 천 단위(소수점 3자리까지 등) → 필요에 따라 조정
                        cell.number_format = "#,##0.#####"
                    else:
                        cell.value = str(value) if value is not None else ''

            # 테두리 설정
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # 열 너비 자동 조정
            for i in range(1, ws.max_column + 1):
                max_length = 0
                column = get_column_letter(i)
                for cell in ws[column]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(file_path)
            os.startfile(file_path)
            messagebox.showinfo("완료", "데이터가 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("Error", f"엑셀 파일을 저장하거나 열 수 없습니다: {e}")

    else:
        # columns를 재정의해도 됨
        col_texts = [treeview.heading(col)["text"] for col in treeview["columns"]]
        data = []
        for item in treeview.get_children():
            if 'totalrow' in treeview.item(item, 'tags'):
                continue
            values = treeview.item(item)["values"]
            data.append(values)

        # pandas → excel 저장 후 openpyxl로 숫자 서식 지정할 수도 있지만,
        # 여기서는 간단히 직접 openpyxl에 쓰는 방식 시연
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Border, Side

            wb = Workbook()
            ws = wb.active

            # 첫 행(헤더)
            for col_idx, col_name in enumerate(col_texts, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)

            # 실제 데이터 (2행부터)
            start_row = 2
            for r_idx, row_values in enumerate(data, start=start_row):
                for c_idx, raw_val in enumerate(row_values, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    numeric_val = parse_as_float(raw_val)
                    if numeric_val is not None:
                        cell.value = numeric_val
                        cell.number_format = "#,##0.#####"
                    else:
                        cell.value = str(raw_val) if raw_val is not None else ''

            # 테두리 설정
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1,
                                    max_row=ws.max_row,
                                    min_col=1,
                                    max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

            # 열너비 자동조정
            for i in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(i)
                for cell in ws[col_letter]:
                    val = cell.value
                    val_len = len(str(val)) if val else 0
                    if val_len > max_length:
                        max_length = val_len
                ws.column_dimensions[col_letter].width = max_length + 2

            wb.save(file_path)
            os.startfile(file_path)
            messagebox.showinfo("완료", "데이터가 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("Error", f"엑셀 파일을 저장하거나 열 수 없습니다: {e}")


# 헤더 높이 정의
header_height = 50  # 헤더의 높이를 픽셀 단위로 지정합니다.

# 필터 설정 정의
FILTERS_CONFIG = {
    'treeview0': {
        '필터그룹1': {
            '이월재고': {
                'beginning_unit_price': '단가',
                'beginning_quantity': '수량',
                'beginning_amount': '금액',
            },
            '입고내역': {
                'incoming_unit_price': '입고단가',
                'incoming_quantity': '입고수량',
                'incoming_amount': '입고금액',
            },
            '잡이익': {
                'misc_profit_quantity': '수량',
                'misc_profit_amount': '금액',
            },
            '장려금': {
                'incentive_quantity': '수량',
                'incentive_amount': '금액',
            },
            '무상지원': {
                'transfer_in_free_quantity': '수량',
                'transfer_in_free_amount': '금액',
            },
            '코드변경': {
                'transfer_in_code_change_quantity': '수량',
                'transfer_in_code_change_amount': '금액',
            },
        },
        '필터그룹2': {
            '출고내역': {
                'outgoing_unit_price': '출고단가',
                'outgoing_quantity': '출고수량',
                'outgoing_amount': '금액',
            },
            '기부': {
                'transfer_out_donation_quantity': '수량',
                'transfer_out_donation_amount': '금액',
            },
            '무상지원': {
                'transfer_out_free_quantity': '수량',
                'transfer_out_free_amount': '금액',
            },
            '사내소비': {
                'transfer_out_internal_use_quantity': '수량',
                'transfer_out_internal_use_amount': '금액',
            },
            '샘플': {
                'transfer_out_sample_quantity': '수량',
                'transfer_out_sample_amount': '금액',
            },
            '직원선물대': {
                'transfer_out_employee_gift_quantity': '수량',
                'transfer_out_employee_gift_amount': '금액',
            },
            '코드변경': {
                'transfer_out_code_change_quantity': '수량',
                'transfer_out_code_change_amount': '금액',
            },
            'LOSS': {
                'transfer_out_loss_quantity': '수량',
                'transfer_out_loss_amount': '금액',
            },
            '계정대체유형(ERP-iU기본)': {
                'transfer_out_account_substitution_quantity': '수량',
                'transfer_out_account_substitution_amount': '금액',
            },
            '사고보상건(분실, 파손 등)': {
                'transfer_out_accident_compensation_quantity': '수량',
                'transfer_out_accident_compensation_amount': '금액',
            },
            '유통기한경과': {
                'transfer_out_expired_quantity': '수량',
                'transfer_out_expired_amount': '금액',
            },
            '재고조정': {
                'transfer_out_inventory_adjustment_quantity': '수량',
                'transfer_out_inventory_adjustment_amount': '금액',
            },
            '정기재고실사': {
                'transfer_out_regular_inventory_check_quantity': '수량',
                'transfer_out_regular_inventory_check_amount': '금액',
            },
            '클레임처리': {
                'transfer_out_claim_processing_quantity': '수량',
                'transfer_out_claim_processing_amount': '금액',
            },
        },
        '필터그룹3': {
            '현재고': {
                'current_unit_price': '단가',
                'current_quantity': '수량',
                'current_amount': '금액',
            },
        '검증': {
                'verification_inventory': '재고검증',
                'verification_negative_inventory': '음수재고검증',
                'verification_outgoing': '출고검증',
                'verification_return': '반품검증',
                'verification_negative_stock_check': '재고음수체크',
                'verification_incentive': '장려금검증',  # 새로 추가
                'verification_free_support': '무상지원 조정',  # 새로 추가
                'verification_quantity': '수량검증',
                'verification_amount': '금액검증',
                'verification_inventory_unit_price': '재고단가검증',
            },
        '재고실사': {
                # 차산점 관련
                'inventory_inspection_chasan_sum_quantity': '차산점 합계_수량',
                'inventory_inspection_chasan_sum_amount': '차산점 합계_금액',
                'inventory_inspection_chasan': '차산점_수량',
                'inventory_inspection_chasan_amount': '차산점_금액',
                'inventory_inspection_chasan_a_quantity': '차산점A_수량',
                'inventory_inspection_chasan_a_amount': '차산점A_금액',
                'inventory_inspection_import_warehouse_chasan_quantity': '수입창고(차산)_수량',
                'inventory_inspection_import_warehouse_chasan_amount': '수입창고(차산)_금액',
                'inventory_inspection_import_warehouse_storage_quantity': '수입창고(보관)_수량',
                'inventory_inspection_import_warehouse_storage_amount': '수입창고(보관)_금액',
                'inventory_inspection_chasan_return_quantity': '차산점반품_수량',
                'inventory_inspection_chasan_return_amount': '차산점반품_금액',
                
                # 청량리점 관련
                'inventory_inspection_cheongnyangni_sum_quantity': '청량리점 합계_수량',
                'inventory_inspection_cheongnyangni_sum_amount': '청량리점 합계_금액',
                'inventory_inspection_cheongnyangni': '청량리점_수량',
                'inventory_inspection_cheongnyangni_amount': '청량리점_금액',
                'inventory_inspection_cheongnyangni_return_quantity': '청량리점반품_수량',
                'inventory_inspection_cheongnyangni_return_amount': '청량리점반품_금액',
                
                # 이천점 관련
                'inventory_inspection_icheon_sum_quantity': '이천점 합계_수량',
                'inventory_inspection_icheon_sum_amount': '이천점 합계_금액',
                'inventory_inspection_icheon': '이천점_수량',
                'inventory_inspection_icheon_amount': '이천점_금액',
                'inventory_inspection_catering_quantity': '케이터링_수량',
                'inventory_inspection_catering_amount': '케이터링_금액',
                'inventory_inspection_ecommerce_quantity': '이커머스_수량',
                'inventory_inspection_ecommerce_amount': '이커머스_금액',
                'inventory_inspection_icheon_return_quantity': '이천점반품_수량',
                'inventory_inspection_icheon_return_amount': '이천점반품_금액',
                
                # 기타
                'inventory_inspection_hanam': '하남점_수량',
                'inventory_inspection_hanam_amount': '하남점_금액',
                'inventory_inspection_prepurchase_quantity': '선매입창고_수량',
                'inventory_inspection_prepurchase_amount': '선매입창고_금액',
                
                # 합계 관련
                'inventory_inspection_total_quantity': '합계수량',
                'inventory_inspection_total_amount': '합계금액',
                'inventory_inspection_difference_quantity': '차이수량',
                'inventory_inspection_difference_amount': '차이금액',
            },
        },
    },
}

# 필터 변수 딕셔너리
filter_vars = {}

# 필터 UI 생성 함수
def create_filters_frame(parent, treeview_name, load_function):
    global filter_vars, treeview0
    filter_vars[treeview_name] = {}

    treeview_filters = FILTERS_CONFIG.get(treeview_name, {})

    # 각 필터 그룹별로 프레임 생성
    for group_idx, (group_name, categories) in enumerate(treeview_filters.items()):
        group_frame = tk.Frame(parent)
        group_frame.pack(side='top', fill='x', padx=5, pady=5)

        # 그룹별 변수 딕셔너리 초기화
        vars_dict = {}
        filter_vars[treeview_name][group_name] = vars_dict

        # '전체' 체크박스와 그룹명을 담을 헤더 프레임 생성
        header_frame = tk.Frame(group_frame)
        header_frame.pack(side='top', fill='x')

        # 그룹 레이블
        group_label = tk.Label(header_frame, text=group_name)
        group_label.pack(side='left')

        # '전체' 체크박스 생성
        group_var = tk.BooleanVar()
        group_var.set(True)
        vars_dict['_group_var'] = group_var

        ttk.Checkbutton(
            header_frame,
            text="전체",
            variable=group_var,
            command=lambda gv=group_var, gn=group_name: toggle_group(gv, filter_vars[treeview_name][gn])
        ).pack(side='left')

        # 옵션들을 담을 프레임 생성
        options_frame = tk.Frame(group_frame)
        options_frame.pack(side='top', fill='x')

        # 옵션 배치 설정
        if group_name == '필터그룹1':
            max_cols = 7  # 한 줄로 배치
        elif group_name == '필터그룹2':
            max_cols = 20  # 한 줄로 배치되도록 큰 값 설정
        else:
            max_cols = 6  # 기본 값

        col_index = 0
        row_index = 0

        # 각 카테고리별로 옵션 생성
        for category_name, options in categories.items():
            # 필터명에서 '이체입고'와 '이체출고' 생략 및 괄호 안 내용 제거
            display_name = category_name.replace('이체입고 ', '').replace('이체출고 ', '')
            display_name = re.sub(r'\(.*?\)', '', display_name).strip()

            var = tk.BooleanVar()
            var.set(True)
            vars_dict[category_name] = var

            ttk.Checkbutton(
                options_frame,
                text=display_name,
                variable=var,
                command=lambda cn=category_name, gn=group_name: toggle_category_var(gn, treeview_name)
            ).grid(row=row_index, column=col_index, sticky='w', padx=2, pady=2)

            col_index += 1
            if col_index >= max_cols:
                col_index = 0
                row_index += 1

    return filter_vars[treeview_name]


def toggle_category_var(group_name, treeview_name):
    treeview_filter_vars = filter_vars[treeview_name]
    group_vars = treeview_filter_vars[group_name]
    all_checked = all(var.get() for key, var in group_vars.items() if key != '_group_var')
    group_vars['_group_var'].set(all_checked)
    apply_column_filters(treeview0)


def toggle_group(group_var, vars_dict):
    global treeview0
    new_state = group_var.get()
    for key, var in vars_dict.items():
        if key != '_group_var':
            var.set(new_state)
    apply_column_filters(treeview0)



# 컬럼 필터 적용 함수
def apply_column_filters(treeview):
    """
    체크박스 상태에 따라 트리뷰의 컬럼을 업데이트하는 함수
    """
    toggle_treeview_columns(treeview, filter_vars['treeview0'])
    update_multi_header_canvas()  # 헤더 캔버스 업데이트



# 항상 표시되어야 할 컬럼 목록
always_display_columns = [
    'item_code', 'item_name', 'specification', 'unit', 'category'
]

def toggle_treeview_columns(treeview, treeview_filter_vars):
    display_columns = always_display_columns.copy()

    treeview_filters = FILTERS_CONFIG.get('treeview0', {})

    for group_name, categories in treeview_filters.items():
        for category_name, options in categories.items():
            category_var = treeview_filter_vars[group_name][category_name]
            if category_var.get():
                for col_key in options.keys():
                    display_columns.append(col_key)

    # 트리뷰의 displaycolumns 설정
    treeview.configure(displaycolumns=display_columns)

    # 멀티헤더 캔버스 업데이트
    update_multi_header_canvas()




# 멀티헤더 캔버스 업데이트 함수
def update_multi_header_canvas():
    """
    멀티헤더 캔버스를 업데이트하는 함수
    """
    draw_multiheader()


# 필터 설정 저장 함수
def save_filter_settings(filter_vars_dict):
    config = configparser.ConfigParser()
    for treeview_name, filters in filter_vars_dict.items():
        section = f'{treeview_name}_Filters'
        config[section] = {}
        for category, vars_dict in filters.items():
            for key, var in vars_dict.items():
                config_key = f"{category}_{key}"
                config[section][config_key] = str(var.get())

    with open('config_msd.ini', 'w', encoding='utf-8') as configfile:
        config.write(configfile)

# 필터 설정 로드 함수
def load_filter_settings(filter_vars_dict):
    config = configparser.ConfigParser()
    config.read('config_msd.ini', encoding='utf-8')

    for treeview_name, filters in filter_vars_dict.items():
        section = f'{treeview_name}_Filters'
        if section in config:
            for category, vars_dict in filters.items():
                for key, var in vars_dict.items():
                    config_key = f"{category}_{key}"
                    if config_key in config[section]:
                        var.set(config[section].getboolean(config_key))

# 프로그램 시작 시 필터 설정 로드
load_filter_settings(filter_vars)

# 전역 변수 선언
treeview0 = None
header_canvas0 = None
header_height = 60  # 헤더 높이 설정

# 멀티헤더 그리기 함수 (전역 함수로 이동)
def draw_multiheader():
    global treeview0, header_canvas0
    header_canvas0.delete('all')  # 기존 헤더 삭제

    # 현재 표시된 컬럼 가져오기
    displayed_columns = treeview0['displaycolumns']
    if displayed_columns == '#all' or displayed_columns == ('#all',):
        displayed_columns = treeview0['columns']
    else:
        displayed_columns = list(displayed_columns)

    # 컬럼 위치와 너비 계산
    positions = []
    total_width = 0
    for col_id in displayed_columns:
        col_width = treeview0.column(col_id, width=None)
        col_text = canvas_column_names.get(col_id, col_id)
        positions.append((total_width, total_width + col_width, col_id, col_text))
        total_width += col_width

    # 컬럼 ID를 키로 하는 딕셔너리로 위치 정보 저장
    col_positions = {col_id: (x1, x2, col_text) for x1, x2, col_id, col_text in positions}

    # 고정 컬럼 그리기
    for col_id in always_display_columns:
        if col_id in col_positions:
            x1, x2, col_text = col_positions[col_id]
            rect_id = header_canvas0.create_rectangle(x1, 0, x2, 40, fill='lightgray', outline='black')
            text_id = header_canvas0.create_text((x1 + x2) / 2, 20, text=col_text)
            # 클릭 이벤트 바인딩
            header_canvas0.tag_bind(rect_id, '<Button-1>',
                                    lambda event, col_id=col_id: handle_header_click(treeview0, col_id, numeric_columns_treeview0))
            header_canvas0.tag_bind(text_id, '<Button-1>',
                                    lambda event, col_id=col_id: handle_header_click(treeview0, col_id, numeric_columns_treeview0))

    # 그룹 헤더 그리기
    for group_name, col_ids in multiheader_groups:
        # 그룹에 속한 표시된 컬럼들만 선택
        group_displayed_cols = [col_id for col_id in col_ids if col_id in displayed_columns]
        if not group_displayed_cols:
            continue  # 그룹에 표시되는 컬럼이 없으면 건너뜀

        # 그룹의 시작과 끝 위치 계산
        x1 = col_positions[group_displayed_cols[0]][0]
        x2 = col_positions[group_displayed_cols[-1]][1]

        # 그룹 헤더 그리기
        header_canvas0.create_rectangle(x1, 0, x2, 20, fill='lightblue', outline='black')
        header_canvas0.create_text((x1 + x2) / 2, 10, text=group_name)

        # 하위 컬럼 헤더 그리기
        for col_id in group_displayed_cols:
            x1, x2, col_text = col_positions[col_id]
            rect_id = header_canvas0.create_rectangle(x1, 20, x2, 40, fill='white', outline='black')
            text_id = header_canvas0.create_text((x1 + x2) / 2, 30, text=col_text)

            # 클릭 이벤트 바인딩 (numeric_columns_treeview0 추가)
            header_canvas0.tag_bind(rect_id, '<Button-1>',
                                    lambda event, col_id=col_id: handle_header_click(treeview0, col_id, numeric_columns_treeview0))
            header_canvas0.tag_bind(text_id, '<Button-1>',
                                    lambda event, col_id=col_id: handle_header_click(treeview0, col_id, numeric_columns_treeview0))


    # 캔버스 크기 조정
    header_canvas0.config(scrollregion=(0, 0, total_width, header_height), width=total_width)


# 트리뷰 설정 및 조회 함수
def hide_treeview_header(treeview):
    style = ttk.Style()
    style.layout("NoHeader.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])
    treeview.configure(style="NoHeader.Treeview")

# 전역 변수로 columns와 multiheader_groups 정의
columns = [
    ('item_code', '품목코드', 100),
    ('item_name', '품명', 300),
    ('specification', '규격', 100),
    ('unit', '단위', 50),
    ('category', '분류', 100),
    # 이월재고
    ('beginning_unit_price', '단가', 80),
    ('beginning_quantity', '수량', 80),
    ('beginning_amount', '금액', 100),
    # 입고내역
    ('incoming_unit_price', '단가', 80),
    ('incoming_quantity', '수량', 80),
    ('incoming_amount', '금액', 100),
    # 잡이익
    ('misc_profit_quantity', '수량', 80),
    ('misc_profit_amount', '금액', 100),
    # 장려금
    ('incentive_quantity', '수량', 80),
    ('incentive_amount', '금액', 100),
    # 이체입고 무상지원
    ('transfer_in_free_quantity', '수량', 80),
    ('transfer_in_free_amount', '금액', 100),
    # 이체입고 코드변경
    ('transfer_in_code_change_quantity', '수량', 80),
    ('transfer_in_code_change_amount', '금액', 100),
    # 출고내역
    ('outgoing_unit_price', '단가', 80),
    ('outgoing_quantity', '수량', 80),
    ('outgoing_amount', '금액', 100),
    # 이체출고 기부
    ('transfer_out_donation_quantity', '수량', 80),
    ('transfer_out_donation_amount', '금액', 100),
    # 이체출고 무상지원
    ('transfer_out_free_quantity', '수량', 80),
    ('transfer_out_free_amount', '금액', 100),
    # 이체출고 사내소비
    ('transfer_out_internal_use_quantity', '수량', 80),
    ('transfer_out_internal_use_amount', '금액', 100),
    # 이체출고 샘플
    ('transfer_out_sample_quantity', '수량', 80),
    ('transfer_out_sample_amount', '금액', 100),
    # 이체출고 직원선물대
    ('transfer_out_employee_gift_quantity', '수량', 80),
    ('transfer_out_employee_gift_amount', '금액', 100),
    # 이체출고 코드변경
    ('transfer_out_code_change_quantity', '수량', 80),
    ('transfer_out_code_change_amount', '금액', 100),
    # 이체출고 LOSS
    ('transfer_out_loss_quantity', '수량', 80),
    ('transfer_out_loss_amount', '금액', 100),
    # 이체출고 계정대체유형(ERP-iU기본)
    ('transfer_out_account_substitution_quantity', '수량', 80),
    ('transfer_out_account_substitution_amount', '금액', 120),
    # 이체출고 사고보상건(분실, 파손 등)
    ('transfer_out_accident_compensation_quantity', '수량', 80),
    ('transfer_out_accident_compensation_amount', '금액', 120),
    # 이체출고 유통기한경과
    ('transfer_out_expired_quantity', '수량', 80),
    ('transfer_out_expired_amount', '금액', 100),
    # 이체출고 재고조정
    ('transfer_out_inventory_adjustment_quantity', '수량', 80),
    ('transfer_out_inventory_adjustment_amount', '금액', 100),
    # 이체출고 정기재고실사
    ('transfer_out_regular_inventory_check_quantity', '수량', 80),
    ('transfer_out_regular_inventory_check_amount', '금액', 100),
    # 이체출고 클레임처리
    ('transfer_out_claim_processing_quantity', '수량', 80),
    ('transfer_out_claim_processing_amount', '금액', 100),
    # 현재고
    ('current_unit_price', '단가', 80),
    ('current_quantity', '수량', 80),
    ('current_amount', '금액', 100),
    # 검증
    ('verification_inventory', '재고검증', 100),
    ('verification_negative_inventory', '음수재고검증', 120),
    ('verification_outgoing', '출고검증', 100),
    ('verification_return', '반품검증', 100),
    ('verification_negative_stock_check', '재고음수체크', 120),
    ('verification_incentive', '장려금검증', 100),  # 새로 추가
    ('verification_free_support', '무상지원 조정', 120),  # 새로 추가
    ('verification_quantity', '수량검증', 100),
    ('verification_amount', '금액검증', 100),
    ('verification_inventory_unit_price', '재고단가검증', 120),


    # 재고실사 (12개: 각 창고 수량/금액 + 합계 수량/금액 + 차이 수량/금액)# === 새로운 재고실사 컬럼들 (검증 컬럼들 다음에 배치) ===
    
    # 차산점 관련
    ('inventory_inspection_chasan_sum_quantity', '차산점 합계_수량', 120),
    ('inventory_inspection_chasan_sum_amount', '차산점 합계_금액', 120),
    ('inventory_inspection_chasan', '차산점_수량', 100),
    ('inventory_inspection_chasan_amount', '차산점_금액', 100),
    ('inventory_inspection_chasan_a_quantity', '차산점A_수량', 100),
    ('inventory_inspection_chasan_a_amount', '차산점A_금액', 100),
    ('inventory_inspection_import_warehouse_chasan_quantity', '수입창고(차산)_수량', 130),
    ('inventory_inspection_import_warehouse_chasan_amount', '수입창고(차산)_금액', 130),
    ('inventory_inspection_import_warehouse_storage_quantity', '수입창고(보관)_수량', 130),
    ('inventory_inspection_import_warehouse_storage_amount', '수입창고(보관)_금액', 130),
    ('inventory_inspection_chasan_return_quantity', '차산점반품_수량', 120),
    ('inventory_inspection_chasan_return_amount', '차산점반품_금액', 120),
    
    # 청량리점 관련
    ('inventory_inspection_cheongnyangni_sum_quantity', '청량리점 합계_수량', 130),
    ('inventory_inspection_cheongnyangni_sum_amount', '청량리점 합계_금액', 130),
    ('inventory_inspection_cheongnyangni', '청량리점_수량', 100),
    ('inventory_inspection_cheongnyangni_amount', '청량리점_금액', 100),
    ('inventory_inspection_cheongnyangni_return_quantity', '청량리점반품_수량', 130),
    ('inventory_inspection_cheongnyangni_return_amount', '청량리점반품_금액', 130),
    
    # 이천점 관련
    ('inventory_inspection_icheon_sum_quantity', '이천점 합계_수량', 120),
    ('inventory_inspection_icheon_sum_amount', '이천점 합계_금액', 120),
    ('inventory_inspection_icheon', '이천점_수량', 100),
    ('inventory_inspection_icheon_amount', '이천점_금액', 100),
    ('inventory_inspection_catering_quantity', '케이터링_수량', 100),
    ('inventory_inspection_catering_amount', '케이터링_금액', 100),
    ('inventory_inspection_ecommerce_quantity', '이커머스_수량', 100),
    ('inventory_inspection_ecommerce_amount', '이커머스_금액', 100),
    ('inventory_inspection_icheon_return_quantity', '이천점반품_수량', 120),
    ('inventory_inspection_icheon_return_amount', '이천점반품_금액', 120),
    
    # 기타 창고
    ('inventory_inspection_hanam', '하남점_수량', 100),
    ('inventory_inspection_hanam_amount', '하남점_금액', 100),
    ('inventory_inspection_prepurchase_quantity', '선매입창고_수량', 120),
    ('inventory_inspection_prepurchase_amount', '선매입창고_금액', 120),
    
    # 기존 총계 관련 (유지)
    ('inventory_inspection_total_quantity', '합계수량', 100),
    ('inventory_inspection_total_amount', '합계금액', 120),
    ('inventory_inspection_difference_quantity', '차이수량', 100),
    ('inventory_inspection_difference_amount', '차이금액', 120),
]




multiheader_groups_info = [
    ('이월재고', 3),
    ('입고내역', 3),
    ('잡이익', 2),
    ('장려금', 2),
    ('이체입고 무상지원', 2),
    ('이체입고 코드변경', 2),
    ('출고내역', 3),
    ('이체출고 기부', 2),
    ('이체출고 무상지원', 2),
    ('이체출고 사내소비', 2),
    ('이체출고 샘플', 2),
    ('이체출고 직원선물대', 2),
    ('이체출고 코드변경', 2),
    ('이체출고 LOSS', 2),
    ('이체출고 계정대체유형(ERP-iU기본)', 2),
    ('이체출고 사고보상건(분실, 파손 등)', 2),
    ('이체출고 유통기한경과', 2),
    ('이체출고 재고조정', 2),
    ('이체출고 정기재고실사', 2),
    ('이체출고 클레임처리', 2),
    ('현재고', 3),
    ('검증', 10),
     # === 새로운 재고실사 그룹들 ===
    ('차산점 합계', 2),         # 차산점 합계_수량, 차산점 합계_금액
    ('차산점', 2),             # 차산점_수량, 차산점_금액
    ('차산점A', 2),            # 차산점A_수량, 차산점A_금액
    ('수입창고(차산)', 2),      # 수입창고(차산)_수량, 수입창고(차산)_금액
    ('수입창고(보관)', 2),      # 수입창고(보관)_수량, 수입창고(보관)_금액
    ('차산점반품', 2),          # 차산점반품_수량, 차산점반품_금액
    ('청량리점 합계', 2),       # 청량리점 합계_수량, 청량리점 합계_금액
    ('청량리점', 2),           # 청량리점_수량, 청량리점_금액
    ('청량리점반품', 2),        # 청량리점반품_수량, 청량리점반품_금액
    ('이천점 합계', 2),         # 이천점 합계_수량, 이천점 합계_금액
    ('이천점', 2),             # 이천점_수량, 이천점_금액
    ('케이터링', 2),           # 케이터링_수량, 케이터링_금액
    ('이커머스', 2),           # 이커머스_수량, 이커머스_금액
    ('이천점반품', 2),          # 이천점반품_수량, 이천점반품_금액
    ('하남점', 2),             # 하남점_수량, 하남점_금액
    ('선매입창고', 2),          # 선매입창고_수량, 선매입창고_금액
    ('합계', 2),               # 합계수량, 합계금액
    ('차이', 2),               # 차이수량, 차이금액
]

# 컬럼 ID 리스트 생성
columns_ids_in_order = [col[0] for col in columns]

# 고정 컬럼 제외
always_display_columns = ['item_code', 'item_name', 'specification', 'unit', 'category']
start_index = len(always_display_columns)  # 5

# 멀티헤더 그룹 생성
current_index = start_index
multiheader_groups = []

for group_name, num_columns in multiheader_groups_info:
    group_columns = columns_ids_in_order[current_index:current_index + num_columns]
    multiheader_groups.append((group_name, group_columns))
    current_index += num_columns

# 캔버스에 표시할 전체 컬럼명을 정의하는 딕셔너리
canvas_column_names = {
    col_id: col_text for col_id, col_text, col_width in columns
}


# 헤더 클릭 처리 함수
def handle_header_click(treeview, col_id, numeric_columns):
    """
    멀티헤더에서 열 클릭 시 정렬 방향을 토글하고 데이터를 정렬합니다.
    """
    # 정렬 상태 초기화
    if not hasattr(treeview, '_sort_states'):
        treeview._sort_states = {}

    # 정렬 방향 토글
    reverse = not treeview._sort_states.get(col_id, False)
    treeview._sort_states[col_id] = reverse

    # 정렬 함수 호출
    sort_treeview_column(treeview, col_id, numeric_columns, reverse)

def reset_treeview0():
    show_treeview(
        treeview0_frame,
        [treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame],
        'treeview0'
    )
    load_treeview0_data(treeview0)

def create_treeview0(parent_frame):
    """
    parent_frame: 트리뷰0가 들어갈 프레임(이미 main 또는 다른 곳에서 생성됨)
    이 함수에서 ttk.Treeview, 스크롤바, 우클릭 메뉴 등을 만들어 하나의 Treeview 위젯을 구성한다.
    한 번 생성되면, 이후 단순히 pack_forget()/pack()으로 보이기/숨기기가 가능하므로 데이터가 유지된다.
    """
    global treeview0, header_canvas0, header_height

    # 임의로 헤더 높이 설정
    header_height = 60

    # 1) 전체 구성 프레임 생성 (부모 프레임 내에 생성)
    frame = tk.Frame(parent_frame)
    frame.pack(fill='both', expand=True)

    # 2) 상단 버튼 영역 (필요한 버튼들을 추가)
    button_frame_top = tk.Frame(frame)
    button_frame_top.pack(side='top', fill='x')

    load_carryover_inventory_button = ttk.Button(
        button_frame_top,
        text="이월재고 불러오기",
        command=lambda: load_carryover_inventory(treeview0)
    )
    load_carryover_inventory_button.pack(side='left', padx=5, pady=5)

    one_stop_button = ttk.Button(
        button_frame_top,
        text="원스탑 작업진행",
        command=lambda: run_one_stop_process(treeview0, status_label)
    )
    one_stop_button.pack(side='left', padx=5, pady=5)

    status_label = tk.Label(button_frame_top, text="", width=40, anchor='w')
    status_label.pack(side='left', padx=5, pady=5)

    resume_codechange_button = ttk.Button(
        button_frame_top,
        text="코드변경 수정값 출고내역에 반영하고 현재고 계산하기 진행",
        command=lambda: resume_one_stop_after_code_change(treeview0, status_label)
    )
    resume_codechange_button.pack(side='left', padx=5, pady=5)

    load_results_button = ttk.Button(
        button_frame_top,
        text="결과 불러오기",
        command=lambda: load_results(treeview0, loaded_month_label)
    )
    load_results_button.pack(side='left', padx=5, pady=5)

    loaded_month_label = tk.Label(button_frame_top, text="", width=30, anchor='w')
    loaded_month_label.pack(side='left', padx=5, pady=5)

    reset_button = ttk.Button(
        button_frame_top,
        text="초기화",
        command=reset_treeview0
    )
    reset_button.pack(side='left', padx=5, pady=5)

    load_ending_inventory_button = ttk.Button(
        button_frame_top,
        text="기말재고 불러오기",
        command=lambda: load_ending_inventory_for_treeview0(treeview0, work_month=WORK_MONTH)
    )
    load_ending_inventory_button.pack(side='left', padx=5, pady=5)

    load_inventory_evaluation_button = ttk.Button(
        button_frame_top,
        text="재고평가 불러오기",
        command=lambda: load_inventory_evaluation_for_treeview0(treeview0)
    )
    load_inventory_evaluation_button.pack(side='left', padx=5, pady=5)

    # 3) 하단 버튼 영역 (추가 기능 버튼)
    button_frame_bottom = tk.Frame(frame)
    button_frame_bottom.pack(side='top', fill='x')

    load_basic_inventory_button = ttk.Button(
        button_frame_bottom,
        text="기초재고 불러오기",
        command=lambda: load_basic_inventory(treeview0)
    )
    load_basic_inventory_button.pack(side='left', padx=5, pady=5)

    load_incoming_data_button = ttk.Button(
        button_frame_bottom,
        text="입고내역 불러오기",
        command=lambda: load_incoming_data(treeview0)
    )
    load_incoming_data_button.pack(side='left', padx=5, pady=5)

    calculate_misc_profit_button = ttk.Button(
        button_frame_bottom,
        text="잡이익 계산하기",
        command=lambda: calculate_misc_profit(treeview0)
    )
    calculate_misc_profit_button.pack(side='left', padx=5, pady=5)

    load_incentive_button = ttk.Button(
        button_frame_bottom,
        text="장려금 불러오기",
        command=lambda: load_incentive_for_treeview0_choice(treeview0)
    )
    load_incentive_button.pack(side='left', padx=5, pady=5)

    load_transfer_in_data_button = ttk.Button(
        button_frame_bottom,
        text="이체입고 불러오기",
        command=lambda: load_transfer_in_data(treeview0)
    )
    load_transfer_in_data_button.pack(side='left', padx=5, pady=5)

    load_shipment_data_button = ttk.Button(
        button_frame_bottom,
        text="출고내역 불러오기",
        command=lambda: load_shipment_data(treeview0)
    )
    load_shipment_data_button.pack(side='left', padx=5, pady=5)

    load_transfer_out_data_button = ttk.Button(
        button_frame_bottom,
        text="이체출고 불러오기",
        command=lambda: load_transfer_out_data(treeview0)
    )
    load_transfer_out_data_button.pack(side='left', padx=5, pady=5)

    calculate_current_inventory_button = ttk.Button(
        button_frame_bottom,
        text="현재고 계산하기",
        command=lambda: calculate_current_inventory(treeview0)
    )
    calculate_current_inventory_button.pack(side='left', padx=5, pady=5)

    perform_verification_button = ttk.Button(
        button_frame_bottom,
        text="검증",
        command=lambda: perform_verification(treeview0)
    )
    perform_verification_button.pack(side='left', padx=5, pady=5)

    save_results_button = ttk.Button(
        button_frame_bottom,
        text="결과 저장하기",
        command=lambda: save_results(treeview0)
    )
    save_results_button.pack(side='left', padx=5, pady=5)

    # 4) 필터 프레임 (초기에는 숨김)
    filters_frame = tk.Frame(frame)
    create_filters_frame(filters_frame, 'treeview0', load_treeview0_data)
    filters_frame.pack_forget()



    def toggle_filters():
        nonlocal filters_visible
        if filters_visible:
            filters_frame.pack_forget()
            filters_visible = False
            toggle_filter_button.config(text="필터 보이기")
        else:
            filters_frame.pack(side='top', fill='x', padx=5, pady=5)
            filters_visible = True
            toggle_filter_button.config(text="필터 감추기")
    filters_visible = False
    toggle_filter_button = tk.Button(frame, text="필터 보이기", command=toggle_filters)
    toggle_filter_button.pack(side='top', anchor='w', padx=5, pady=5)

    # 5) 헤더 캔버스 생성
    header_canvas0 = tk.Canvas(frame, height=header_height)
    header_canvas0.pack(side='top', fill='x')

    # 여기서 **status_label** 생성
    status_label_local = tk.Label(button_frame_top, text="", width=40, anchor='w')
    status_label_local.pack(side='left', padx=5, pady=5)

    # 6) 트리뷰와 스크롤바를 담을 프레임 생성
    treeview_container = tk.Frame(frame)
    treeview_container.pack(fill='both', expand=True)

    # 수직 스크롤바 (부모는 treeview_container)
    y_scrollbar = ttk.Scrollbar(treeview_container, orient="vertical")
    y_scrollbar.pack(side='right', fill='y')

    tree_columns = [col[0] for col in columns]  # 전역 columns 사용
    treeview0 = ttk.Treeview(treeview_container, columns=tree_columns, show='', yscrollcommand=y_scrollbar.set)
    treeview0.pack(side='left', fill='both', expand=True)
    y_scrollbar.config(command=treeview0.yview)

    # 수평 스크롤바 생성 (부모는 frame)
    x_scrollbar = ttk.Scrollbar(frame, orient="horizontal")
    x_scrollbar.pack(side='bottom', fill='x')

    treeview0.configure(xscrollcommand=x_scrollbar.set)
    header_canvas0.configure(xscrollcommand=x_scrollbar.set)
    x_scrollbar.configure(command=lambda *args: [treeview0.xview(*args), header_canvas0.xview(*args)])

    # 7) 컬럼 설정 (여기서 'item_name'만 왼쪽 정렬로 설정)
    for col_id, col_text, col_width in columns:
        # 'item_name' 컬럼만 anchor='w' (왼쪽정렬), 나머지는 anchor='center'
        if col_id == 'item_name':
            treeview0.column(col_id, width=col_width, anchor='w', stretch=False)
        else:
            treeview0.column(col_id, width=col_width, anchor='center', stretch=False)

        treeview0.heading(
            col_id,
            text=col_text,
            command=lambda c=col_id: sort_treeview_column(treeview0, c, numeric_columns_treeview0)
        )

    # 8) 기본 헤더 감추기
    hide_treeview_header(treeview0)

    # 9) 컬럼 크기 조정 비활성화
    def disable_column_resize(event):
        if treeview0.identify_region(event.x, event.y) == "separator":
            return "break"
    treeview0.bind('<Button-1>', disable_column_resize)

    # 10) 멀티헤더 재그리기
    def on_treeview_configure(event):
        draw_multiheader()
    treeview0.bind('<Configure>', on_treeview_configure)

    # 11) 행의 음영 재적용 함수
    def reapply_row_tags(event=None):
        for index, item in enumerate(treeview0.get_children('')):
            tags = list(treeview0.item(item, 'tags'))

            # 합계(totalrow)는 건너뜀
            if 'totalrow' in tags:
                continue

            # pinned 행은 even/odd를 덮어씌우지 않음
            if 'pinned' in tags:
                continue

            # 그 외 행: 기존 evenrow/oddrow 제거 후 새로 지정
            tags = [tag for tag in tags if tag not in ('evenrow', 'oddrow')]
            row_tag = 'evenrow' if index % 2 == 0 else 'oddrow'
            tags.append(row_tag)
            treeview0.item(item, tags=tags)


    # 태그 스타일 설정
    treeview0.tag_configure('evenrow', background='white')
    treeview0.tag_configure('oddrow', background='lightgray')
    treeview0.tag_configure('pinned', background='gold')
    
    # 재계산된 값에 적용할 태그 (글씨 파란색)    
    treeview0.tag_configure('editedrow', foreground='red')


    # 13) 더블 클릭 이벤트 바인딩
    treeview0.bind("<Double-1>", on_treeview_double_click)

    # 14) 우클릭 메뉴 실제 생성 후 추가
    context_menu = tk.Menu(treeview0, tearoff=0)
    

    def show_edit_popup():
        info = getattr(treeview0, 'editing_info', None)
        if not info:
            return

        row_id = info['row_id']
        col_index_all = info['col_index_all']
        edit_col = info['col_id']

        values = list(treeview0.item(row_id, 'values'))
        columns = treeview0['columns']
        col_idx_map = {col_name: i for i, col_name in enumerate(columns)}

        edit_win = tk.Toplevel()
        edit_win.title("값 수정")
        edit_win.grab_set()

        tk.Label(edit_win, text="새로운 값:").pack(pady=5)
        entry = tk.Entry(edit_win)
        entry.insert(0, values[col_index_all])
        entry.pack(pady=5)
        
        # ★ 추가: popup 뜨자마자 입력란에 포커스
        entry.focus_set()
        entry.selection_range(0, tk.END)

        def apply_edit():
            new_val = entry.get().strip()
            try:
                # 입력값(쉼표 제거 후 float 변환)
                numeric_val = float(new_val.replace(',', ''))

                # 현재 행 값 복사
                updated_values = list(values)
                # 수정된 값 반영
                updated_values[col_index_all] = str(numeric_val)

                # ─────────────────────────────────────────
                # (1) 이체입고 코드변경금액 수정 시
                # ─────────────────────────────────────────
                if edit_col == 'transfer_in_code_change_amount':
                    sum_1_1_cols = [
                        'beginning_amount', 'incoming_amount', 'misc_profit_amount',
                        'incentive_amount', 'transfer_in_free_amount', 'transfer_in_code_change_amount'
                    ]
                    sum_1_2_cols = [
                        'beginning_quantity', 'incoming_quantity',
                        'transfer_in_free_quantity', 'transfer_in_code_change_quantity'
                    ]

                    total_1_1 = 0.0
                    for c in sum_1_1_cols:
                        idx_c = col_idx_map[c]
                        val_str = updated_values[idx_c]
                        val_f = float(val_str.replace(',', '')) if val_str else 0.0
                        total_1_1 += val_f

                    total_1_2 = 0.0
                    for c in sum_1_2_cols:
                        idx_c = col_idx_map[c]
                        val_str = updated_values[idx_c]
                        val_f = float(val_str.replace(',', '')) if val_str else 0.0
                        total_1_2 += val_f

                    # 출고단가 재계산
                    out_unit_idx = col_idx_map['outgoing_unit_price']
                    if total_1_2 != 0:
                        ratio = total_1_1 / total_1_2
                    else:
                        ratio = 0.0
                    updated_values[out_unit_idx] = f"{ratio:.5f}"

                    # 출고금액 = 출고단가 × 출고수량
                    out_qty_idx = col_idx_map['outgoing_quantity']
                    out_amt_idx = col_idx_map['outgoing_amount']
                    try:
                        out_qty_str = updated_values[out_qty_idx]
                        out_qty_val = float(out_qty_str.replace(',', '')) if out_qty_str else 0.0
                        new_out_amt = ratio * out_qty_val
                        updated_values[out_amt_idx] = f"{new_out_amt:.2f}"
                    except:
                        pass

                # ─────────────────────────────────────────
                # (2) 이월재고단가(beginning_unit_price)를 수정 시
                # ─────────────────────────────────────────
                elif edit_col == 'beginning_unit_price':
                    """
                    사용자가 '이월재고 단가'를 직접 입력한 경우,
                    (1) 이월재고 금액 = 이월재고 단가 × 이월재고 수량  
                    (2) 출고단가 = (이월재고금액 + 입고금액 + 잡이익 - 장려금 + 이체입고코드변경금액) / (이월재고수량 + 입고수량 + 무상수량 + 이체입고코드변경수량)  
                        → 단, 분모가 0이면, 먼저 이월재고 수량이 0인지 확인하여,
                            - 만약 이월재고 수량 ≠ 0: 출고단가 = 이월재고금액 ÷ 이월재고 수량  
                            - 만약 이월재고 수량도 0이면: 출고단가 = 사용자가 입력한 이월재고 단가  
                    (3) 출고금액 = 출고단가 × 출고수량
                    """
                    # (A) 이월재고 금액 재계산: beginning_amount = beginning_unit_price × beginning_quantity
                    beg_qty_idx = col_idx_map.get('beginning_quantity')
                    beg_amt_idx = col_idx_map.get('beginning_amount')
                    if beg_qty_idx is not None and beg_amt_idx is not None:
                        beg_qty_str = updated_values[beg_qty_idx]
                        beg_qty_val = float(beg_qty_str.replace(',', '')) if beg_qty_str else 0.0
                        new_beg_amt = numeric_val * beg_qty_val
                        updated_values[beg_amt_idx] = f"{new_beg_amt:.2f}"

                    # (B) 출고단가 재계산
                    
                    sum_1_1_cols = [
                        'beginning_amount', 'incoming_amount', 'misc_profit_amount',
                        'incentive_amount',  # 장려금은 음수로 처리
                        'transfer_in_code_change_amount'
                    ]
                    sum_1_2_cols = [
                        'beginning_quantity', 'incoming_quantity',
                        'transfer_in_free_quantity', 'transfer_in_code_change_quantity'
                    ]

                    total_1_1 = 0.0
                    for c in sum_1_1_cols:
                        idx_c = col_idx_map.get(c)
                        if idx_c is not None:
                            val_str = updated_values[idx_c]
                            val_f = float(val_str.replace(',', '')) if val_str else 0.0
                            # 장려금는 빼줍니다.
                            if c == 'incentive_amount':
                                total_1_1 -= val_f
                            else:
                                total_1_1 += val_f

                    total_1_2 = 0.0
                    for c in sum_1_2_cols:
                        idx_c = col_idx_map.get(c)
                        if idx_c is not None:
                            val_str = updated_values[idx_c]
                            val_f = float(val_str.replace(',', '')) if val_str else 0.0
                            total_1_2 += val_f

                    out_unit_idx = col_idx_map.get('outgoing_unit_price')
                    out_amt_idx    = col_idx_map.get('outgoing_amount')
                    out_qty_idx    = col_idx_map.get('outgoing_quantity')

                    if out_unit_idx is not None:
                        if total_1_2 != 0:
                            new_out_uprice = total_1_1 / total_1_2
                        else:
                            # 분모가 0인 경우, 추가로 이월재고 수량이 있는지 확인
                            if beg_qty_idx is not None:
                                beg_qty_str = updated_values[beg_qty_idx]
                                beg_qty_val = float(beg_qty_str.replace(',', '')) if beg_qty_str else 0.0
                                if beg_qty_val != 0:
                                    # 이월재고 수량이 있으면, 이월재고금액 / 이월재고수량
                                    new_out_uprice = float(updated_values[beg_amt_idx].replace(',', '')) / beg_qty_val
                                else:
                                    # 이월재고 수량도 0이면, fallback: 사용자 입력값 사용
                                    new_out_uprice = numeric_val
                            else:
                                new_out_uprice = numeric_val
                        updated_values[out_unit_idx] = f"{new_out_uprice:.5f}"

                    # (C) 출고금액 = 출고단가 × 출고수량
                    if out_qty_idx is not None and out_amt_idx is not None:
                        out_qty_str = updated_values[out_qty_idx] or '0'
                        out_qty_val = float(out_qty_str.replace(',', ''))
                        new_out_amt = new_out_uprice * out_qty_val
                        updated_values[out_amt_idx] = f"{new_out_amt:.2f}"



                # ─────────────────────────────────────────
                # (3) 출고내역단가(outgoing_unit_price)를 수정 시
                # ─────────────────────────────────────────
                elif edit_col == 'outgoing_unit_price':
                    # 사용자가 직접 입력한 단가 × 출고수량 = 출고금액
                    out_qty_idx = col_idx_map['outgoing_quantity']
                    out_amt_idx = col_idx_map['outgoing_amount']
                    try:
                        out_qty_str = updated_values[out_qty_idx] or '0'
                        out_qty_val = float(out_qty_str.replace(',', ''))
                        new_out_amt = numeric_val * out_qty_val
                        updated_values[out_amt_idx] = f"{new_out_amt:.2f}"
                    except:
                        pass

                treeview0.item(row_id, values=updated_values)

                # ───────── 여기부터 추가 ─────────
                # 수정된 행에 'editedrow' 태그 붙이기
                current_tags = list(treeview0.item(row_id, 'tags'))
                if 'editedrow' not in current_tags:
                    current_tags.append('editedrow')
                treeview0.item(row_id, tags=current_tags)
                # ────────────────────────────────

                # 숫자 포맷 적용 & 합계행 업데이트
                format_numeric_columns(treeview0, numeric_columns_treeview0)
                update_total_row(treeview0)  # ← 합계행 갱신

                edit_win.destroy()

            except ValueError:
                messagebox.showerror("입력오류", "숫자 형식이 아닙니다.")

        def cancel_edit():
            edit_win.destroy()
        
        # ESC 키 입력 시 팝업창 종료
        edit_win.bind("<Escape>", lambda event: edit_win.destroy())

        entry.bind("<Return>", lambda e: apply_edit())

        button_frame = tk.Frame(edit_win)
        button_frame.pack(pady=5)
        tk.Button(button_frame, text="반영", command=apply_edit).pack(side='left', padx=5)
        tk.Button(button_frame, text="취소", command=cancel_edit).pack(side='left', padx=5)

        # 우클릭 메뉴에서 이 함수를 불러옴
        context_menu.add_command(label="수정", command=show_edit_popup)


    def pin_selected_rows():
        selected_items = treeview0.selection()
        all_children = list(treeview0.get_children(''))
        totalrow_indexes = []
        pinned_indexes = []

        for idx, child in enumerate(all_children):
            row_tags = treeview0.item(child, 'tags')
            if 'totalrow' in row_tags:
                totalrow_indexes.append(idx)
            elif 'pinned' in row_tags:
                pinned_indexes.append(idx)

        selected_in_order = [child for child in all_children if child in selected_items]

        insertion_index = 0
        if pinned_indexes:
            insertion_index = pinned_indexes[-1] + 1
        elif totalrow_indexes:
            insertion_index = totalrow_indexes[-1] + 1

        # pinned_order 리스트가 없으면 초기화
        if not hasattr(treeview0, 'pinned_order'):
            treeview0.pinned_order = []

        for item in selected_in_order:
            current_tags = [tag for tag in treeview0.item(item, 'tags') if tag not in ('evenrow', 'oddrow')]
            if 'pinned' not in current_tags:
                current_tags.append('pinned')   # pinned 태그 추가
            treeview0.item(item, tags=current_tags)

            # pinned 행은 totalrow/pinned 바로 아래쪽에 배치하기
            treeview0.move(item, '', insertion_index)
            treeview0.pinned_order.append(item)  # 고정된 행 순서를 저장
            insertion_index += 1

        # 태그 재적용
        reapply_row_tags(treeview0)

    context_menu.add_command(label="정렬고정", command=pin_selected_rows)

    def unpin_selected_rows():
        # 1) 선택된 행에서 'pinned' 태그 제거
        selected_items = treeview0.selection()
        for item in selected_items:
            current_tags = list(treeview0.item(item, 'tags'))
            if 'pinned' in current_tags:
                current_tags.remove('pinned')
                treeview0.item(item, tags=current_tags)
        
        # 2) 전체 행 중에서 'totalrow'는 건너뛰고, 현재 pinned 태그가 없는 행들만 추출
        all_children = list(treeview0.get_children(''))
        non_pinned = [child for child in all_children if ('pinned' not in treeview0.item(child, 'tags')) 
                    and ('totalrow' not in treeview0.item(child, 'tags'))]
        
        # 만약 정렬된(=unpinned) 행들이 있다면,
        # 그 마지막 행 바로 밑 위치(인덱스)를 구함.
        if non_pinned:
            # all_children 내에서 non_pinned[-1]의 위치를 찾고 그 다음 위치를 insertion_index로 사용
            insertion_index = all_children.index(non_pinned[-1]) + 1
        else:
            # 없으면 그냥 맨 위에 삽입
            insertion_index = 0
        
        # 3) 선택된(지금 unpin된) 행들을 위에서 계산한 insertion_index부터 순서대로 이동
        for item in selected_items:
            treeview0.move(item, '', insertion_index)
            insertion_index += 1

        reapply_row_tags(treeview0)

    context_menu.add_command(label="정렬해제", command=unpin_selected_rows)

    def on_right_click(event):
        row_id = treeview0.identify_row(event.y)
        col_id = treeview0.identify_column(event.x)
        if not row_id or not col_id:
            return
        display_cols = treeview0['displaycolumns']
        col_index_in_display = int(col_id.replace('#', '')) - 1
        if col_index_in_display < 0 or col_index_in_display >= len(display_cols):
            return
        edit_col = display_cols[col_index_in_display]
        editable_columns = [
            # 단가·금액
            'beginning_unit_price',
            'transfer_in_code_change_amount',
            'outgoing_unit_price',
            # ───── 추가할 “수량” 컬럼 ─────
            'beginning_quantity',          # 이월재고 수량
            'incoming_quantity',           # 입고내역 수량
            'misc_profit_quantity',        # 잡이익 수량
            'incentive_quantity',          # 장려금 수량
            'transfer_in_free_quantity',   # 무상지원 수량
            'transfer_in_code_change_quantity',
            'outgoing_quantity',           # 출고내역 수량
            # 이체출고 전 유형
            'transfer_out_donation_quantity',
            'transfer_out_free_quantity',
            'transfer_out_internal_use_quantity',
            'transfer_out_sample_quantity',
            'transfer_out_employee_gift_quantity',
            'transfer_out_code_change_quantity',
            'transfer_out_loss_quantity',
            'transfer_out_account_substitution_quantity',
            'transfer_out_accident_compensation_quantity',
            'transfer_out_expired_quantity',
            'transfer_out_inventory_adjustment_quantity',
            'transfer_out_regular_inventory_check_quantity',
            'transfer_out_claim_processing_quantity',
            # 현재고·검증·재고실사 등 필요한 경우 추가
        ]

        context_menu.delete(0, 'end')
        if edit_col in editable_columns:
            all_columns = treeview0['columns']
            col_index_in_all = all_columns.index(edit_col)
            treeview0.editing_info = {
                'row_id': row_id,
                'col_index_display': col_index_in_display,
                'col_index_all': col_index_in_all,
                'col_id': edit_col
            }
            context_menu.add_command(label="수정", command=show_edit_popup)
        else:
            treeview0.editing_info = None
        context_menu.add_command(label="정렬고정", command=pin_selected_rows)
        context_menu.add_command(label="정렬해제", command=unpin_selected_rows)
        try:
            context_menu.post(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    treeview0.bind("<Button-3>", on_right_click)
    draw_multiheader()
    return (treeview0, status_label_local)



def load_basic_inventory(treeview):
    """
    기초재고 데이터를 불러와 트리뷰에 업데이트하는 함수
    (매칭되지 않은 행 전체를 xlsx로 저장, NaN은 빈 칸, item_name 등 추가 컬럼도 함께 저장)
    """

    # NaN/None -> '' 치환 함수
    def safe_str(value):
        if value is None:
            return ''
        if isinstance(value, float) and math.isnan(value):
            return ''
        return str(value)

    current_year = datetime.now().year
    selected_year = simpledialog.askstring(
        "기초재고 불러오기", 
        "연도를 입력하세요 (예: 2025):", 
        initialvalue=str(current_year)
    )
    if not selected_year:
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # exclude_item_codes 테이블에서 제외할 품목코드 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        excluded_codes = set(row[0] for row in cursor.fetchall())

        # 필요한 컬럼을 모두 SELECT
        # (item_code, item_name, specification, unit, category, beginning_unit_price, beginning_quantity, beginning_amount, reference_year)
        cursor.execute("""
            SELECT 
                item_code, 
                item_name, 
                specification, 
                unit, 
                category, 
                beginning_unit_price, 
                beginning_quantity, 
                beginning_amount,
                reference_year
            FROM mds_basic_data
            WHERE reference_year = %s
        """, (selected_year,))
        basic_data_rows = cursor.fetchall()

        if not basic_data_rows:
            messagebox.showinfo("정보", f"{selected_year}년도의 기초재고 데이터가 존재하지 않습니다.")
            return

        # 트리뷰의 품목코드와 매핑하여 데이터 업데이트
        item_code_to_item_id = {}
        for item_id in treeview.get_children():
            # 만약 '합계' 행이나 기타 제외해야 할 행이 있다면, 여기서 필터링 가능
            # 예: if '합계' in treeview.item(item_id, 'values'): continue

            item_values = treeview.item(item_id, 'values')
            item_code = item_values[0]  # '품목코드'는 첫 번째 컬럼
            if item_code not in excluded_codes:
                item_code_to_item_id[item_code] = item_id

        # 매칭되지 않은 전체 원본 행을 저장할 리스트
        unmatched_data = []

        for row in basic_data_rows:
            # row = (item_code, item_name, specification, unit, category, 
            #        beginning_unit_price, beginning_quantity, beginning_amount, reference_year)
            item_code = row[0]

            # exclude_item_codes 에 없고, 트리뷰에서도 찾지 못하면 unmatched
            if item_code in item_code_to_item_id:
                # 매칭 성공 -> 트리뷰 업데이트
                item_id = item_code_to_item_id[item_code]
                values = list(treeview.item(item_id, 'values'))

                columns = treeview['columns']
                col_indices = {col: idx for idx, col in enumerate(columns)}

                # 이 부분은 기존과 동일 (기초단가, 기초수량, 기초금액만 트리뷰에 업데이트)
                # 필요하다면 item_name 등도 트리뷰에 표시할 수 있음
                unit_price = row[5]
                quantity = row[6]
                amount = row[7]

                values[col_indices['beginning_unit_price']] = format_numeric_value(unit_price)
                values[col_indices['beginning_quantity']] = format_numeric_value(quantity)
                values[col_indices['beginning_amount']] = format_numeric_value(amount)
                treeview.item(item_id, values=values)

            else:
                # 매칭 실패
                if item_code not in excluded_codes:
                    unmatched_data.append(row)

        if unmatched_data:
            # 팝업 메시지
            messagebox.showinfo("완료", f"매칭되지 않은 품목코드가 {len(unmatched_data)}개 있습니다.")

            # --- (1) xlsx 파일 자동 생성 로직 ---
            global base_path
            timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
            xlsx_file_name = f"unmatched_codes_{timestamp_str}.xlsx"
            xlsx_file_path = os.path.join(base_path, xlsx_file_name)

            wb = Workbook()
            ws = wb.active
            ws.title = "UnmatchedData"

            # 헤더 작성 (원하는 컬럼명으로)
            ws.append(["품목코드", "품목명", "규격", "단위", "카테고리", 
                       "기초단가", "기초수량", "기초금액", "기준연도"])

            # 내용 작성 (NaN이나 None은 safe_str로 치환)
            for data_row in unmatched_data:
                # data_row는 (item_code, item_name, specification, unit, category, beginning_unit_price, beginning_quantity, beginning_amount, reference_year)
                ws.append([safe_str(x) for x in data_row])

            wb.save(xlsx_file_path)

            # --- (2) 저장된 xlsx 파일 실행(Windows 환경) ---
            if platform.system() == "Windows":
                os.startfile(xlsx_file_path)
            else:
                # macOS / Linux 등
                # subprocess.call(["open", xlsx_file_path])  # 필요 시
                pass

        else:
            messagebox.showinfo("완료", "기초재고 데이터 로딩이 완료되었습니다.")

        # 합계 행 업데이트
        update_total_row(treeview)
        # 행 음영 재적용
        reapply_row_tags(treeview)
        # 숫자 포맷팅
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()




def load_carryover_inventory(treeview):
    """
    이월재고 데이터를 불러와 트리뷰에 업데이트하는 함수 (가공 없이 그대로 불러오기)
    """
    from datetime import datetime, timedelta

    # 이전 달의 년/월을 가져오는 함수
    def get_previous_month():
        today = datetime.today()
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return last_month.strftime("%Y/%m")

    # 기준월 입력
    default_month = get_previous_month()
    reference_month = simpledialog.askstring("이월재고 불러오기", "기준월을 입력하세요 :", initialvalue=default_month)
    if not reference_month:
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # mds_monthly_inventory_transactions 테이블에서 선택된 기준월의 데이터 가져오기
        cursor.execute("""
            SELECT item_code, current_unit_price, current_quantity, current_amount
            FROM mds_monthly_inventory_transactions
            WHERE reference_month = %s
        """, (reference_month,))
        carryover_data_rows = cursor.fetchall()

        if not carryover_data_rows:
            messagebox.showinfo("정보", f"{reference_month}의 이월재고 데이터가 존재하지 않습니다.")
            return

        # 트리뷰의 품목코드와 매칭하여 데이터 업데이트
        item_code_to_item_id = {}
        for item_id in treeview.get_children():
            item_values = treeview.item(item_id, 'values')
            item_code = item_values[0]  # '품목코드'는 첫 번째 컬럼
            item_code_to_item_id[item_code] = item_id

        unmatched_codes = []
        for row in carryover_data_rows:
            item_code, unit_price, quantity, amount = row

            if item_code in item_code_to_item_id:
                item_id = item_code_to_item_id[item_code]
                # 현재 아이템의 값을 가져옴
                values = list(treeview.item(item_id, 'values'))

                # 컬럼 인덱스 가져오기
                columns = treeview['columns']
                col_indices = {col: idx for idx, col in enumerate(columns)}

                # 데이터 업데이트 (가공 없이 그대로 업데이트)
                values[col_indices['beginning_unit_price']] = unit_price if unit_price is not None else ''
                values[col_indices['beginning_quantity']] = quantity if quantity is not None else ''
                values[col_indices['beginning_amount']] = amount if amount is not None else ''

                treeview.item(item_id, values=values)
            else:
                unmatched_codes.append(item_code)

        # 매칭되지 않은 코드의 개수 출력
        if unmatched_codes:
            messagebox.showinfo("완료", f"매칭되지 않은 품목코드가 {len(unmatched_codes)}개 있습니다.")
        else:
            messagebox.showinfo("완료", "이월재고 데이터 로딩이 완료되었습니다.")

        print(f"이월재고 불러오기 진행 완료. 매칭되지 않은 품목코드: {len(unmatched_codes)}개")

        # 데이터 로드 완료 후 합계 행 업데이트
        update_total_row(treeview)

        # 음영 재적용
        reapply_row_tags(treeview)

        # 숫자 포맷팅 적용
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# =====================================================================
# (1) 원스탑 작업 함수
# =====================================================================
def run_one_stop_process(treeview, status_label):
    """
    원스탑 작업을 진행하는 함수
    """
    global WORK_MONTH
    def get_previous_month():
        today = datetime.today()
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return last_month.strftime("%Y/%m")

    # (A) 작업월 입력
    default_month = get_previous_month()
    work_month = simpledialog.askstring("작업월 선택",
                                        "작업월을 입력하세요 (예: 2025/01):",
                                        initialvalue=default_month)
    if not work_month:
        messagebox.showerror("오류", "작업월이 입력되지 않았습니다.")
        return
    WORK_MONTH = work_month 
    
    # (B) '이월재고' 체크
    columns = treeview['columns']
    col_indices = {col: idx for idx, col in enumerate(columns)}
    beginning_quantity_idx = col_indices.get('beginning_quantity')
    data_exists = False
    for item_id in treeview.get_children():
        values = treeview.item(item_id, 'values')
        if values[beginning_quantity_idx] not in [None, '', '0', '0.0']:
            data_exists = True
            break

    if not data_exists:
        messagebox.showinfo("알림", "이월재고를 먼저 불러와주세요.")
        return

    try:
        status_label.config(text=f"{work_month} 데이터가 원스탑으로 작업이 진행중입니다.")
        status_label.update_idletasks()

        # (1) 입고내역 불러오기
        load_incoming_data(treeview, work_month)

        # (2) 잡이익 계산하기
        calculate_misc_profit(treeview)

        # (3) 장려금 불러오기(팝업으로 선택)
        load_incentive_for_treeview0_choice(treeview, work_month)

        # (4) 이체입고 불러오기
        load_transfer_in_data(treeview, work_month)

        # (5) 출고내역 불러오기
        load_shipment_data(treeview, work_month)

        # (6) 이체출고 불러오기
        load_transfer_out_data(treeview, work_month)

        # === 여기를 기준으로 잠시 멈춤 ===
        response = messagebox.askyesno(
            "이체입고 코드변경 수정",
            "이체입고 코드변경 금액값을 수정하고, 출고내역에 반영한 뒤,\n"
            "현재고(7) 계산 및 검증(8)을 진행하시겠습니까?\n\n"
            "예(Y)를 누르면 현재 작업을 잠시 멈추고,\n"
            "수정 후에 (7)부터 진행할 수 있습니다.\n"
            "아니오(N)를 누르면 (7) 현재고 계산 → (8) 검증까지 바로 진행합니다."
        )
        if response:
            # 예 -> 여기서 멈추기
            messagebox.showinfo(
                "수정 대기",
                "이체입고 코드변경 금액을 수정한 뒤,\n"
                "‘코드변경 수정값 출고내역에 반영하고 현재고 계산하기’ 버튼을 눌러\n"
                "7번부터 진행하세요."
            )
            status_label.config(text="코드변경 수정 대기중...")
            return
        else:
            # 아니오 -> (7)현재고 계산, (8)검증
            calculate_current_inventory(treeview)
            perform_verification(treeview)

            status_label.config(text=f"{work_month} 데이터의 작업이 완료되었습니다.")
            status_label.update_idletasks()
            messagebox.showinfo("완료", f"{work_month} 데이터의 작업을 마쳤습니다.")

            # 결과 저장 여부
            save_response = messagebox.askyesno("결과 저장", f"{work_month} 월의 작업 결과를 저장하시겠습니까?")
            if save_response:
                save_results(treeview)
            else:
                messagebox.showinfo("알림", "결과 저장이 취소되었습니다.")

    except Exception as e:
        messagebox.showerror("오류", f"작업 중 오류가 발생했습니다: {e}")
        status_label.config(text=f"{work_month} 데이터의 작업 중 오류가 발생했습니다.")


def resume_one_stop_after_code_change(treeview, status_label):
    """
    (6) 이체출고 불러오기까지 완료된 상태에서,
    '이체입고 코드변경' 금액을 수정 후 (7)과 (8)을 이어서 진행하는 함수.
    수정된 금액/수량 값을 이용하여 출고내역 단가/금액을 업데이트하고 이어서 작업합니다.
    
    [A-1 방식]
    - 이미 사용자에 의해 수정된 outgoing_unit_price를 덮어쓰지 않고,
    - 여기서는 "outgoing_amount"만 (수정된 단가 × 출고수량)으로 다시 계산.
    """

    # (A) 작업월 확인
    work_month = simpledialog.askstring(
        "작업월 확인",
        "다시 작업월을 입력하세요 (예: 2025/01)\n"
        "※ (6) 이체출고 불러오기까지 이미 로드된 월과 동일해야 합니다."
    )
    if not work_month:
        messagebox.showerror("오류", "작업월이 입력되지 않았습니다.")
        return

    try:
        # (A) 먼저 "출고내역 단가" 변경된 값을 이용해
        #     모든 이체출고 금액 transfer_out_xxx_amount를 싹 재계산
        columns = list(treeview['columns'])
        col_idx_map = {col: i for i, col in enumerate(columns)}

         # 출고단가 인덱스
        out_uprice_idx = col_idx_map.get('outgoing_unit_price')

        # 이체출고에 해당하는 모든 수량/금액 컬럼 쌍
                
        transfer_out_pairs = [
            ('transfer_out_donation_quantity', 'transfer_out_donation_amount'),
            ('transfer_out_free_quantity', 'transfer_out_free_amount'),
            ('transfer_out_internal_use_quantity', 'transfer_out_internal_use_amount'),
            ('transfer_out_sample_quantity', 'transfer_out_sample_amount'),
            ('transfer_out_employee_gift_quantity', 'transfer_out_employee_gift_amount'),
            ('transfer_out_code_change_quantity', 'transfer_out_code_change_amount'),
            ('transfer_out_loss_quantity', 'transfer_out_loss_amount'),
            ('transfer_out_account_substitution_quantity', 'transfer_out_account_substitution_amount'),
            ('transfer_out_accident_compensation_quantity', 'transfer_out_accident_compensation_amount'),
            ('transfer_out_expired_quantity', 'transfer_out_expired_amount'),
            ('transfer_out_inventory_adjustment_quantity', 'transfer_out_inventory_adjustment_amount'),
            ('transfer_out_regular_inventory_check_quantity', 'transfer_out_regular_inventory_check_amount'),
            ('transfer_out_claim_processing_quantity', 'transfer_out_claim_processing_amount'),
        ]
        
        # # (사용 중인 컬럼만 추가)
        # transfer_out_pairs = [
        #     ('transfer_out_donation_amount'),
        #     ('transfer_out_free_amount'),
        #     ('transfer_out_internal_use_amount'),
        #     ('transfer_out_sample_amount'),
        #     ('transfer_out_employee_gift_amount'),
        #     ('transfer_out_code_change_amount'),
        #     ('transfer_out_loss_amount'),
        #     ('transfer_out_account_substitution_amount'),
        #     ('transfer_out_accident_compensation_amount'),
        #     ('transfer_out_expired_amount'),
        #     ('transfer_out_inventory_adjustment_amount'),
        #     ('transfer_out_regular_inventory_check_amount'),
        #     ('transfer_out_claim_processing_amount'),
        # ]

        for item_id in treeview.get_children():
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue

            values = list(treeview.item(item_id, 'values'))
            try:
                # 출고단가
                out_uprice_str = values[out_uprice_idx]
                out_uprice_val = float(out_uprice_str.replace(',', '')) if out_uprice_str else 0.0

                for (qty_col, amt_col) in transfer_out_pairs:
                    qty_idx = col_idx_map[qty_col]
                    amt_idx = col_idx_map[amt_col]
                    qty_str = values[qty_idx]
                    qty_val = float(qty_str.replace(',', '')) if qty_str else 0.0
                    # 새로운 금액
                    new_amount = out_uprice_val * qty_val
                    values[amt_idx] = format_numeric_value(new_amount)

                # 업데이트
                treeview.item(item_id, values=values)

            except Exception as ex:
                print(f"[이체출고 재계산 오류] item_id={item_id}, {ex}")

        # (C) (7) 현재고 계산
        calculate_current_inventory(treeview)

        # (8) 검증
        perform_verification(treeview)

        status_label.config(text=f"{work_month} 데이터: (7)~(8) 작업이 완료되었습니다.")
        status_label.update_idletasks()
        messagebox.showinfo("완료", f"{work_month} 데이터의 남은 작업(7~8)을 마쳤습니다.")

        # (D) 결과 저장 여부
        save_response = messagebox.askyesno("결과 저장", f"{work_month} 월의 수정 결과를 저장하시겠습니까?")
        if save_response:
            save_results(treeview)
        else:
            messagebox.showinfo("알림", "결과 저장이 취소되었습니다.")

    except Exception as e:
        messagebox.showerror("오류", f"작업 중 오류가 발생했습니다: {e}")
        status_label.config(text=f"{work_month} 데이터의 (7)~(8) 작업 중 오류가 발생했습니다.")





def load_incoming_data(treeview, work_month=None):
    """
    '입고내역 불러오기' 버튼 클릭 시 실행되는 함수입니다.
    - mds_purchase_receipt_status 테이블에서 reference_month가 일치하는
      item_code별 management_quantity, won_amount 합계를 GROUP BY로 가져옴.
    - 트리뷰0에서 해당 item_code를 찾아, 'incoming_quantity', 'incoming_amount',
      'incoming_unit_price' 컬럼에 값을 반영.
    """
    today = datetime.now()
    first_day_of_this_month = today.replace(day=1)
    previous_month_date = first_day_of_this_month - timedelta(days=1)
    default_month = previous_month_date.strftime('%Y/%m')


    if work_month is None:
        selected_month = simpledialog.askstring("입고내역 불러오기", 
                                                "조회할 년월을 입력하세요 (예: 2025/01):", 
                                                initialvalue=default_month)
        if not selected_month:
            return
        reference_month = selected_month
    else:
        reference_month = work_month

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # GROUP BY item_code로 합산된 결과 가져오기
        # management_quantity, won_amount를 SUM()으로 합산
        # 필요시 item_name, specification 등을 MAX(...)로 가져올 수도 있음
        cursor.execute("""
            SELECT 
                item_code,
                SUM(management_quantity) AS sum_qty,
                SUM(won_amount) AS sum_amt
            FROM mds_purchase_receipt_status
            WHERE reference_month = %s
            GROUP BY item_code
        """, (reference_month,))
        incoming_data_rows = cursor.fetchall()
        # incoming_data_rows = [(item_code, sum_qty, sum_amt), ...]

        if not incoming_data_rows:
            messagebox.showinfo("정보", f"{reference_month}의 입고내역 데이터가 존재하지 않습니다.")
            return

        # 트리뷰에서 item_code -> item_id 매핑
        treeview_items = treeview.get_children()
        item_code_to_item_id = {}
        for item_id in treeview_items:
            item_values = treeview.item(item_id, 'values')
            # '품목코드'가 첫 번째 컬럼이라고 가정
            item_code = item_values[0]  
            item_code_to_item_id[item_code] = item_id

        unmatched_codes = []
        for (item_code, sum_qty, sum_amt) in incoming_data_rows:
            try:
                quantity = float(sum_qty) if sum_qty else 0.0
                amount = float(sum_amt) if sum_amt else 0.0
                unit_price = amount / quantity if quantity != 0 else 0.0
            except ValueError:
                quantity, amount, unit_price = 0.0, 0.0, 0.0

            if item_code in item_code_to_item_id:
                item_id = item_code_to_item_id[item_code]
                values = list(treeview.item(item_id, 'values'))

                columns = treeview['columns']
                col_indices = {col: idx for idx, col in enumerate(columns)}

                # 'incoming_quantity', 'incoming_amount', 'incoming_unit_price' 컬럼에 값 반영
                if 'incoming_quantity' in col_indices:
                    values[col_indices['incoming_quantity']] = format_numeric_value(quantity)
                if 'incoming_amount' in col_indices:
                    values[col_indices['incoming_amount']] = format_numeric_value(amount)
                if 'incoming_unit_price' in col_indices:
                    values[col_indices['incoming_unit_price']] = format_numeric_value(unit_price)

                treeview.item(item_id, values=values)
            else:
                unmatched_codes.append(item_code)

        # 합계 행, 음영, 숫자 포맷팅
        update_total_row(treeview)
        reapply_row_tags(treeview)
        format_numeric_columns(treeview, numeric_columns_treeview0)

        # 필요시 매칭되지 않은 item_code 처리
        if unmatched_codes:
            print("매칭되지 않은 item_code:", unmatched_codes)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()




def calculate_misc_profit(treeview):
    """
    '잡이익 계산하기' 버튼 클릭 시 실행되는 함수
    변경사항:
    - None 비교 에러 방지 (incoming_quantity가 None일 경우 체크)
    - format_numeric_value 호출에서 decimals 인자 제거
    """
    item_ids = treeview.get_children()
    unmatched_items = 0

    columns = treeview['columns']
    col_indices = {col: idx for idx, col in enumerate(columns)}

    for item_id in item_ids:
        if 'totalrow' in treeview.item(item_id, 'tags'):
            continue

        values = list(treeview.item(item_id, 'values'))
        try:
            idx_incoming_quantity = col_indices.get('incoming_quantity')
            idx_beginning_unit_price = col_indices.get('beginning_unit_price')
            idx_incoming_amount = col_indices.get('incoming_amount')
            idx_misc_profit_amount = col_indices.get('misc_profit_amount')

            if None in [idx_incoming_quantity, idx_beginning_unit_price, idx_incoming_amount, idx_misc_profit_amount]:
                unmatched_items += 1
                continue

            def safe_float(val):
                if val is None or val == '':
                    return None
                try:
                    return float(str(val).replace(',', ''))
                except:
                    return None

            incoming_quantity = safe_float(values[idx_incoming_quantity])
            beginning_unit_price = safe_float(values[idx_beginning_unit_price])
            incoming_amount = safe_float(values[idx_incoming_amount])

            if incoming_quantity is not None and incoming_quantity < 0:
                # 계산식: ROUND(('입고내역_수량' * '이월재고_단가') - '입고내역_금액', 0)
                misc_profit_amount = round((incoming_quantity * (beginning_unit_price if beginning_unit_price else 0)) - (incoming_amount if incoming_amount else 0), 0)
            else:
                misc_profit_amount = 0

            values[idx_misc_profit_amount] = format_numeric_value(misc_profit_amount)
            treeview.item(item_id, values=values)

        except Exception as e:
            unmatched_items += 1
            print(f"아이템 처리 중 오류 발생 (ID: {item_id}): {e}")
            continue

    update_total_row(treeview)
    reapply_row_tags(treeview)
    format_numeric_columns(treeview, numeric_columns_treeview0)
    # messagebox.showinfo("완료", "잡이익 계산이 완료되었습니다.")
    # print(f"잡이익 계산 완료. 처리되지 않은 항목 수: {unmatched_items}")


def load_incentive_for_treeview0_choice(treeview, default_month=None):
    """
    장려금 불러오기 방식을 선택하는 함수.
    DB 방식 vs 엑셀 업로드 방식을 묻는 팝업을 띄워서,
    DB를 선택하면 load_incentive_for_treeview0_db,
    엑셀 업로드를 선택하면 load_incentive_for_treeview0_excel 로 분기.
    default_month를 두 번째 인자로 받아, DB 로직에 전달할 수 있음.
    """
    method = ask_incentive_loading_method()  # 팝업에서 'db' 또는 'excel' 선택
    if method is None:
        return  # 사용자가 창 닫거나 취소

    if method == 'db':
        # DB 조회 로직
        load_incentive_for_treeview0_db(treeview, default_month)
    else:
        # 엑셀 업로드 로직
        load_incentive_for_treeview0_excel(treeview)



def load_incentive_for_treeview0_db(treeview, default_month=None):
    """
    '장려금 불러오기' DB조회 로직
    :param default_month: (옵션) 기본 표시할 년/월 값 (예: '2025/02')
    """
    from datetime import datetime, timedelta
    import tkinter.simpledialog as sd

    # (A) 기존 '이전 달' 기본값 계산
    today = datetime.now()
    first_day_this_month = today.replace(day=1)
    previous_month_date = first_day_this_month - timedelta(days=1)
    fallback_month = previous_month_date.strftime("%Y/%m")

    # (B) popup에서 사용될 기본값 결정
    #     - default_month가 있다면 그것 사용, 없으면 fallback_month
    if default_month:
        init_value = default_month
    else:
        init_value = fallback_month

    # 팝업 부모: 트리뷰의 최상위 윈도우
    parent_win = treeview.winfo_toplevel()

    # ────────── (A) 부모창을 최상위로 올려서 팝업이 포커스 잘 잡히도록 ──────────
    parent_win.lift()
    parent_win.attributes("-topmost", True)
    parent_win.update()

    reference_month = sd.askstring(
        "장려금 불러오기",
        "조회할 년월을 입력하세요 (예: 2025/01):",
        parent=parent_win,         # 팝업을 parent_win 앞에 표시
        initialvalue=init_value
    )
    if not reference_month:
        return
      

    # 다시 원상 복구
    parent_win.attributes("-topmost", False)

    # (B) 사용자가 취소하거나 ESC 누른 경우 None
    if not reference_month:
        return
    

    # 이하 기존 DB조회 로직 그대로
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT item_code, incentive
            FROM mds_incentive_result
            WHERE reference_month = %s
        """, (reference_month,))
        incentive_data_rows = cursor.fetchall()

        if not incentive_data_rows:
            messagebox.showinfo("정보", f"{reference_month}의 장려금 데이터가 존재하지 않습니다.")
            return

        # 트리뷰 매칭
        treeview_items = treeview.get_children()
        item_code_to_item_id = {}
        for item_id in treeview_items:
            item_values = treeview.item(item_id, 'values')
            item_code = item_values[0]
            item_code_to_item_id[item_code] = item_id

        unmatched_codes = []
        columns = treeview['columns']
        col_indices = {col: idx for idx, col in enumerate(columns)}
        inc_idx = col_indices['incentive_amount']

        for (item_code, incentive) in incentive_data_rows:
            try:
                incentive_val = float(incentive) if incentive not in [None, ''] else 0.0
            except ValueError:
                incentive_val = 0.0

            if item_code in item_code_to_item_id:
                row_id = item_code_to_item_id[item_code]
                values = list(treeview.item(row_id, 'values'))
                values[inc_idx] = f"{incentive_val:,.2f}"
                treeview.item(row_id, values=values)
            else:
                unmatched_codes.append(item_code)

        # 후처리
        update_total_row(treeview)
        reapply_row_tags(treeview)
        format_numeric_columns(treeview, numeric_columns_treeview0)

        if unmatched_codes:
            print("매칭되지 않은 item_code:", unmatched_codes)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def load_incentive_for_treeview0_excel(treeview):
    """
    새로 추가된: 엑셀파일로 장려금 데이터를 불러오는 함수.
    - 장려금 컬럼을 먼저 초기화한 뒤,
    - 엑셀데이터를 트리뷰에 매칭/업데이트.
    """
    # 파일 선택
    file_path = filedialog.askopenfilename(
        title="장려금 엑셀파일을 선택하세요",
        filetypes=[("Excel files", "*.xlsx *.xls")],
        initialdir="."
    )
    if not file_path:
        return  # 사용자가 선택 취소

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # [1] 트리뷰 전체 행에서 'incentive_amount' 컬럼값을 초기화
        columns = treeview['columns']
        col_indices = {col: idx for idx, col in enumerate(columns)}

        if 'incentive_amount' not in col_indices:
            raise KeyError("'incentive_amount' 컬럼을 찾을 수 없습니다.")

        # 트리뷰 모든 아이템 순회 → incentive_amount = 0 or ''
        for item_id in treeview.get_children():
            values = list(treeview.item(item_id, 'values'))
            # 원하는 초기값 (예: 0 / 빈 문자열 / "0.00" 등)
            values[col_indices['incentive_amount']] = '0'
            treeview.item(item_id, values=values)

        # [2] 트리뷰의 품목코드->item_id 매핑 딕셔너리 생성
        treeview_items = treeview.get_children()
        item_code_to_item_id = {}
        for item_id in treeview_items:
            item_values = treeview.item(item_id, 'values')
            item_code = item_values[0]  # '품목코드'는 첫 번째 컬럼
            item_code_to_item_id[item_code] = item_id

        unmatched_codes = []
        matched_count = 0

        # [3] 엑셀 데이터 읽어와서 트리뷰 업데이트
        # 엑셀의 1~2행은 헤더, 3행부터 실제 데이터.
        # 마지막 행도 '합계'라고 무조건 제외하지 않고,
        # '품목코드'가 비어 있으면 제외
        for row_idx in range(3, ws.max_row + 1):
            excel_item_code = ws.cell(row=row_idx, column=1).value
            excel_incentive = ws.cell(row=row_idx, column=4).value

            # 품목코드가 비어있으면 업로드에서 제외
            if not excel_item_code:
                continue

            # 문자열 처리
            excel_item_code = str(excel_item_code).strip()
            try:
                excel_incentive = float(excel_incentive) if excel_incentive not in [None, ''] else 0.0
            except:
                excel_incentive = 0.0

            # 트리뷰 매칭
            if excel_item_code in item_code_to_item_id:
                this_item_id = item_code_to_item_id[excel_item_code]
                values = list(treeview.item(this_item_id, 'values'))

                # 'incentive_amount' 컬럼에 업데이트
                values[col_indices['incentive_amount']] = f"{excel_incentive:,.2f}"
                treeview.item(this_item_id, values=values)
                matched_count += 1
            else:
                unmatched_codes.append(excel_item_code)

        # [4] 매칭되지 않은 코드가 있다면, 엑셀로 저장
        if unmatched_codes:
            wb_unmatched = openpyxl.Workbook()
            ws_unmatched = wb_unmatched.active
            ws_unmatched.title = "Unmatched Item Codes"
            ws_unmatched.append(["품목코드"])
            for code in unmatched_codes:
                ws_unmatched.append([code])

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"엑셀업로드_매칭되지_않은_품목코드_{timestamp}.xlsx"
            wb_unmatched.save(filename)
            os.startfile(filename)

            messagebox.showinfo(
                "결과",
                f"총 {matched_count}개 품목코드를 업데이트했습니다.\n"
                f"매칭되지 않은 품목코드가 {len(unmatched_codes)}개 있습니다.\n"
                f"파일: {filename}"
            )
        else:
            messagebox.showinfo(
                "결과",
                f"엑셀 업로드 완료!\n"
                f"총 {matched_count}개 품목코드를 업데이트했습니다.\n"
                f"매칭되지 않은 품목코드는 없습니다."
            )

        # [5] 마무리(합계, 음영, 숫자포맷 등)
        update_total_row(treeview)
        reapply_row_tags(treeview)
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("에러", f"엑셀 업로드 중 오류가 발생했습니다: {e}")


# =====================================================================
# (3) 장려금 불러오기 방식 선택 팝업 
# =====================================================================
def ask_incentive_loading_method():
    method_window = tk.Toplevel()
    method_window.title("장려금 불러오기 방법 선택")
    method_window.geometry("300x120")

    # 팝업창이 실제 표시될 때까지 대기
    method_window.wait_visibility(method_window)

    # 최상위로 올림 + 포커스 강제
    method_window.lift()
    method_window.focus_force()

    # 팝업창 닫힐 때까지 부모 창을 누를 수 없게(모달)
    method_window.grab_set()

    # ESC 키 누르면 팝업창만 종료
    method_window.bind("<Escape>", lambda e: method_window.destroy())

    var_method = tk.StringVar(value='')

    label = tk.Label(method_window, text="장려금 불러올 방식을 선택하세요.")
    label.pack(pady=10)

    def select_db():
        var_method.set('db')
        method_window.destroy()

    def select_excel():
        var_method.set('excel')
        method_window.destroy()

    btn_db = tk.Button(method_window, text="기존로직", command=select_db, width=10)
    btn_db.pack(side="left", padx=20, pady=10)

    btn_excel = tk.Button(method_window, text="엑셀업로드", command=select_excel, width=10)
    btn_excel.pack(side="right", padx=20, pady=10)
    
    # ───────────── [1] Enter 키 → "기존로직" ─────────────
    method_window.bind("<Return>", lambda e: select_db())
    

    method_window.wait_window()  # 이 창이 닫힐 때까지 대기

    result = var_method.get()
    return result if result else None


def save_unmatched_codes_to_excel(unmatched_codes, prefix="매칭되지_않은_품목코드"):
    """
    매칭되지 않은 품목코드를 엑셀로 저장 후 파일명을 리턴
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unmatched Item Codes"
    ws.append(["품목코드"])
    for code in unmatched_codes:
        ws.append([code])

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{prefix}_{timestamp}.xlsx"
    wb.save(filename)
    os.startfile(filename)
    return filename


def load_transfer_in_data(treeview, work_month=None):
    """
    '이체입고 불러오기' 버튼 클릭 시 실행되는 함수입니다.
    """
    today = datetime.now()
    first_day_of_this_month = today.replace(day=1)
    previous_month_date = first_day_of_this_month - timedelta(days=1)
    default_month = previous_month_date.strftime('%Y/%m')

    if work_month is None:
        selected_month = simpledialog.askstring("이체입고 불러오기", "조회할 년월을 입력하세요 (예: 2025/01):", initialvalue=default_month)
        if not selected_month:
            return
        reference_month = selected_month
    else:
        reference_month = work_month

    try:
        selected_year_month = reference_month.replace('/', '')
        if len(selected_year_month) != 6:
            raise ValueError("올바른 형식의 년월을 입력하세요.")
    except Exception as e:
        messagebox.showerror("에러", f"년월 입력 형식이 잘못되었습니다: {e}")
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        cursor.execute("""
            SELECT item_code, substitution_type, quantity
            FROM mds_account_substitution_input
            WHERE SUBSTRING(input_number, 4, 6) = %s
            AND substitution_type IN ('무상지원', '코드변경')
        """, (selected_year_month,))
        transfer_in_rows = cursor.fetchall()

        if not transfer_in_rows:
            messagebox.showinfo("정보", f"{reference_month}의 이체입고 데이터가 존재하지 않습니다.")
            return

        # --- 수정 부분: 해당 열값 초기화 ---
        # 트리뷰의 모든 행(합계행 제외)에서 'transfer_in_free_quantity'와 'transfer_in_code_change_quantity' 열 값을 0으로 초기화합니다.
        columns = treeview['columns']
        col_indices = {col: idx for idx, col in enumerate(columns)}
        for item_id in treeview.get_children():
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue
            values = list(treeview.item(item_id, 'values'))
            for col_name in ['transfer_in_free_quantity', 'transfer_in_code_change_quantity']:
                if col_name in col_indices:
                    idx = col_indices[col_name]
                    values[idx] = format_numeric_value(0)
            treeview.item(item_id, values=values)
        # --- 초기화 끝 ---

        # 데이터 집계
        data_dict = {}
        for item_code, substitution_type, quantity in transfer_in_rows:
            key = (item_code, substitution_type)
            data_dict[key] = data_dict.get(key, 0) + quantity

        treeview_items = treeview.get_children()
        item_code_to_item_id = {}
        for item_id in treeview_items:
            item_values = treeview.item(item_id, 'values')
            item_code_in_tree = item_values[0]
            item_code_to_item_id[item_code_in_tree] = item_id

        unmatched_codes = []
        for (item_code, substitution_type), total_quantity in data_dict.items():
            if item_code in item_code_to_item_id:
                item_id = item_code_to_item_id[item_code]
                if 'totalrow' in treeview.item(item_id, 'tags'):
                    continue

                values = list(treeview.item(item_id, 'values'))
                # substitution_type에 따라 적용할 열 지정
                if substitution_type == '무상지원':
                    col_name = 'transfer_in_free_quantity'
                elif substitution_type == '코드변경':
                    col_name = 'transfer_in_code_change_quantity'
                else:
                    continue

                if col_name in col_indices:
                    idx = col_indices[col_name]
                    existing_value_str = values[idx]
                    existing_value = safe_float_from_string(existing_value_str) or 0.0
                    new_value = existing_value + total_quantity

                    values[idx] = format_numeric_value(new_value)
                    treeview.item(item_id, values=values)
                else:
                    print(f"컬럼 '{col_name}'을 찾을 수 없습니다.")
            else:
                unmatched_codes.append(item_code)

        update_total_row(treeview)
        reapply_row_tags(treeview)
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



def load_shipment_data(treeview, work_month=None):
    """
    '출고내역 불러오기' 버튼 클릭 시 실행되는 함수입니다.
    (개선: mds_shipment_status 테이블에 없는 품목도 계산식 적용)
    """
    # 현재 날짜를 기준으로 전월을 기본값으로 설정
    today = datetime.now()
    first_day_of_this_month = today.replace(day=1)
    previous_month_date = first_day_of_this_month - timedelta(days=1)
    default_month = previous_month_date.strftime('%Y/%m')

    if work_month is None:
        selected_month = simpledialog.askstring("출고내역 불러오기", 
                                                "조회할 년월을 입력하세요 (예: 2025/01):", 
                                                initialvalue=default_month)
        if not selected_month:
            return
        reference_month = selected_month
    else:
        reference_month = work_month

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 1) mds_shipment_status 테이블에서 선택한 월의 item_code, shipment_quantity 불러오기
        cursor.execute("""
            SELECT item_code, shipment_quantity
            FROM mds_shipment_status
            WHERE reference_month = %s
        """, (reference_month,))
        shipment_rows = cursor.fetchall()

        # 2) dict 형태로 저장 (테이블에 없는 품목은 dict에 없음)
        shipment_dict = {}
        for item_code, quantity in shipment_rows:
            shipment_dict[item_code] = quantity

        # 트리뷰 모든 아이템에 대해 로직 적용
        treeview_items = treeview.get_children()
        columns = treeview['columns']
        col_indices = {col: idx for idx, col in enumerate(columns)}

        # 편의 함수: 특정 컬럼에서 float 값 얻기
        def get_numeric_value_from_column(values_list, col_name):
            idx = col_indices.get(col_name)
            if idx is not None and idx < len(values_list):
                value_str = values_list[idx]
                return safe_float_from_string(value_str) or 0.0
            return 0.0

        unmatched_codes = []

        for item_id in treeview_items:
            # 합계행 제외
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue

            values = list(treeview.item(item_id, 'values'))
            if not values:
                continue

            item_code_in_tree = values[0]  # '품목코드'는 첫 번째 컬럼

            # 3) shipment_quantity = dict에서 가져오거나 0
            shipment_quantity = shipment_dict.get(item_code_in_tree, 0)

            # 4) 출고내역_수량 업데이트
            col_name_quantity = 'outgoing_quantity'
            if col_name_quantity in col_indices:
                idx_quantity = col_indices[col_name_quantity]
                values[idx_quantity] = format_numeric_value(shipment_quantity)
            else:
                print(f"컬럼 '{col_name_quantity}'을 찾을 수 없습니다.")

            # 5) 필요한 컬럼값 가져오기
            beginning_amount               = get_numeric_value_from_column(values, 'beginning_amount')
            incoming_amount                = get_numeric_value_from_column(values, 'incoming_amount')
            misc_profit_amount             = get_numeric_value_from_column(values, 'misc_profit_amount')
            incentive_amount               = get_numeric_value_from_column(values, 'incentive_amount')
            # 기존 코드: transfer_in_free_amount 삭제
            transfer_in_code_change_amount = get_numeric_value_from_column(values, 'transfer_in_code_change_amount')

            beginning_quantity             = get_numeric_value_from_column(values, 'beginning_quantity')
            incoming_quantity              = get_numeric_value_from_column(values, 'incoming_quantity')
            transfer_in_free_quantity      = get_numeric_value_from_column(values, 'transfer_in_free_quantity')
            transfer_in_code_change_quantity = get_numeric_value_from_column(values, 'transfer_in_code_change_quantity')

            # 6) 계산 로직
            numerator = (beginning_amount + incoming_amount + misc_profit_amount
                         - incentive_amount + transfer_in_code_change_amount)

            denominator = (beginning_quantity + incoming_quantity
                           + transfer_in_free_quantity + transfer_in_code_change_quantity)

            if denominator != 0:
                outgoing_unit_price = round(numerator / denominator, 5)  # 소수점 5자리
            else:
                outgoing_unit_price = 0.0

            # '출고내역_단가' 업데이트
            col_name_unit_price = 'outgoing_unit_price'
            if col_name_unit_price in col_indices:
                idx_unit_price = col_indices[col_name_unit_price]
                values[idx_unit_price] = format_numeric_value(outgoing_unit_price)
            else:
                print(f"컬럼 '{col_name_unit_price}'을 찾을 수 없습니다.")

            # '출고내역_금액' 계산 = 단가 × 출고수량
            outgoing_amount = outgoing_unit_price * shipment_quantity
            col_name_amount = 'outgoing_amount'
            if col_name_amount in col_indices:
                idx_amount = col_indices[col_name_amount]
                values[idx_amount] = format_numeric_value(outgoing_amount)
            else:
                print(f"컬럼 '{col_name_amount}'을 찾을 수 없습니다.")

            # 7) 출고수량 ≠ 0 & 단가가 0이면 이월재고_단가 사용, 없으면 priority 태그
            if shipment_quantity != 0:
                new_out_uprice_val = get_numeric_value_from_column(values, 'outgoing_unit_price')
                if new_out_uprice_val == 0.0:
                    # 이월재고_단가
                    beginning_unit_price = get_numeric_value_from_column(values, 'beginning_unit_price')
                    if beginning_unit_price:
                        # 대체
                        if col_name_unit_price in col_indices:
                            values[idx_unit_price] = format_numeric_value(beginning_unit_price)
                        # 금액 재계산
                        new_amount = beginning_unit_price * shipment_quantity
                        if col_name_amount in col_indices:
                            values[idx_amount] = format_numeric_value(new_amount)
                    else:
                        # priority 태그
                        current_tags = list(treeview.item(item_id, 'tags'))
                        if 'priority' not in current_tags:
                            current_tags.append('priority')
                        treeview.item(item_id, tags=current_tags)

            # 8) 업데이트
            treeview.item(item_id, values=values)

            # 만약 shipment_dict에 없음 → unmatched_codes 추가
            if item_code_in_tree not in shipment_dict:
                unmatched_codes.append(item_code_in_tree)

        # 중복 제거
        unmatched_codes = list(set(unmatched_codes))

        # # unmatched_codes 결과 저장
        # if unmatched_codes:
        #     import openpyxl
        #     from openpyxl import Workbook
        #     import os

        #     wb = Workbook()
        #     ws = wb.active
        #     ws.title = "Unmatched Item Codes"
        #     ws.append(["품목코드"])
        #     for code in unmatched_codes:
        #         ws.append([code])

        #     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        #     filename = f"출고내역_매칭되지_않은_품목코드_{timestamp}.xlsx"
        #     wb.save(filename)

        #     os.startfile(filename)
        #     messagebox.showinfo("완료", 
        #         f"트리뷰에 {len(treeview_items)}개 품목이 있었으며, "
        #         f"이 중 mds_shipment_status에 없는 품목이 {len(unmatched_codes)}개입니다.\n"
        #         f"파일: {filename}")
        # else:
        #     messagebox.showinfo("완료", "출고내역 데이터 로딩이 완료되었습니다. (모두 매칭)")

        # print(f"출고내역 불러오기 진행 완료. 매칭되지 않은 품목코드: {len(unmatched_codes)}개")

        # priority 태그가 있는 아이템을 상단으로 이동
        priority_items = [i for i in treeview.get_children('') if 'priority' in treeview.item(i, 'tags')]
        for idx, pid in enumerate(priority_items):
            treeview.move(pid, '', idx)

        # 데이터 로드 완료 후 합계 행 업데이트
        update_total_row(treeview)
        # 음영 재적용
        reapply_row_tags(treeview)
        # 숫자 포맷팅
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("데이터 로드 중 오류 발생", f"{e}")
        print(f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



def load_transfer_out_data(treeview, work_month=None):
    """
    '이체출고 불러오기' 버튼 클릭 시 실행되는 함수입니다.
    (출고 + 입고 양쪽 모두에 없는 품목도 union으로 처리)
    """
    today = datetime.now()
    first_day_of_this_month = today.replace(day=1)
    previous_month_date = first_day_of_this_month - timedelta(days=1)
    default_month = previous_month_date.strftime('%Y/%m')

    if work_month is None:
        selected_month = simpledialog.askstring(
            "이체출고 불러오기", "조회할 년월을 입력하세요 (예: 2025/01):",
            initialvalue=default_month
        )
        if not selected_month:
            return
        reference_month = selected_month
    else:
        reference_month = work_month

    try:
        # 입력된 년월의 형식 검증
        selected_year_month = reference_month.replace('/', '')
        if len(selected_year_month) != 6:
            raise ValueError("올바른 형식의 년월을 입력하세요.")
    except Exception as e:
        messagebox.showerror("에러", f"년월 입력 형식이 잘못되었습니다: {e}")
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()




# 예시: 이체출고에 활용되는 컬럼 전부
        TRANSFER_OUT_COLUMNS_TO_RESET = [
    'transfer_out_donation_quantity', 'transfer_out_donation_amount',
    'transfer_out_free_quantity', 'transfer_out_free_amount',
    'transfer_out_internal_use_quantity', 'transfer_out_internal_use_amount',
    'transfer_out_sample_quantity', 'transfer_out_sample_amount',
    'transfer_out_employee_gift_quantity', 'transfer_out_employee_gift_amount',
    'transfer_out_code_change_quantity', 'transfer_out_code_change_amount',
    'transfer_out_loss_quantity', 'transfer_out_loss_amount',
    'transfer_out_account_substitution_quantity', 'transfer_out_account_substitution_amount',
    'transfer_out_accident_compensation_quantity', 'transfer_out_accident_compensation_amount',
    'transfer_out_expired_quantity', 'transfer_out_expired_amount',
    'transfer_out_inventory_adjustment_quantity', 'transfer_out_inventory_adjustment_amount',
    'transfer_out_regular_inventory_check_quantity', 'transfer_out_regular_inventory_check_amount',
    'transfer_out_claim_processing_quantity', 'transfer_out_claim_processing_amount'
]
        # --- [수정 부분] 모든 'transfer_out' 컬럼 초기화 ---
        columns = treeview['columns']
        col_indices = {col: idx for idx, col in enumerate(columns)}

        # 위에서 정의한 리스트 (TRANSFER_OUT_COLUMNS_TO_RESET)를 불러와서 0으로 세팅
        for item_id in treeview.get_children():
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue  # 합계행 제외

            values = list(treeview.item(item_id, 'values'))
            for col_name in TRANSFER_OUT_COLUMNS_TO_RESET:
                if col_name in col_indices:
                    idx = col_indices[col_name]
                    # 숫자 포맷 함수가 있다면 사용(없으면 그냥 '0'으로 대입해도 됨)
                    values[idx] = format_numeric_value(0)
            treeview.item(item_id, values=values)
        # --- [초기화 끝] ---

        # 이체출고 대상 substitution_type 목록
        substitution_types = [
            '기부', '사내소비', '샘플', '직원선물대', 'LOSS',
            '계정대체유형(ERP-iU기본)', '사고보상건(분실, 파손 등)', '유통기한경과',
            '재고조정', '정기재고실사', '클레임처리', '무상지원', '코드변경'
        ]

        # (1) output_rows
        cursor.execute(f"""
            SELECT item_code, substitution_type, SUM(quantity)
            FROM mds_account_substitution_output
            WHERE SUBSTRING(output_number, 4, 6) = %s
              AND substitution_type IN %s
            GROUP BY item_code, substitution_type
        """, (selected_year_month, tuple(substitution_types)))
        output_rows = cursor.fetchall()

        # (2) input_rows
        cursor.execute(f"""
            SELECT item_code, substitution_type, SUM(quantity)
            FROM mds_account_substitution_input
            WHERE SUBSTRING(input_number, 4, 6) = %s
              AND substitution_type IN %s
            GROUP BY item_code, substitution_type
        """, (selected_year_month, tuple(substitution_types)))
        input_rows = cursor.fetchall()

        # 출고/입고 데이터를 딕셔너리로 (key = (item_code, substitution_type))
        output_data = {}
        for item_code, substitution_type, quantity in output_rows:
            key = (item_code, substitution_type)
            output_data[key] = output_data.get(key, 0) + quantity

        input_data = {}
        for item_code, substitution_type, quantity in input_rows:
            key = (item_code, substitution_type)
            input_data[key] = input_data.get(key, 0) + quantity

        # 출고 + 입고 합집합 키
        all_keys = set(output_data.keys()) | set(input_data.keys())

        data_dict = {}
        for key in all_keys:
            item_code, substitution_type = key
            output_qty = output_data.get(key, 0)
            input_qty  = input_data.get(key, 0)
            net_qty    = output_qty - input_qty  # 일반적인 경우

            # '무상지원', '코드변경'은 input_qty를 무시하고 output_qty만 사용
            if substitution_type in ['무상지원', '코드변경']:
                data_dict[key] = output_qty
            else:
                # 나머지 유형: net_qty가 0이 아닐 때만 저장
                if net_qty != 0:
                    data_dict[key] = net_qty

        if not data_dict:
            messagebox.showinfo("정보", f"{reference_month}의 이체출고 데이터가 존재하지 않습니다.")
            return

        # 트리뷰에 반영
        treeview_items = treeview.get_children()
        item_code_to_item_id = {}
        for item_id in treeview_items:
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue
            item_values = treeview.item(item_id, 'values')
            item_code_in_tree = item_values[0]
            item_code_to_item_id[item_code_in_tree] = item_id

        unmatched_codes = []
        db_debug_rows = []

        for (item_code, substitution_type), total_quantity in data_dict.items():
            matched = False
            reason = ""

            if item_code in item_code_to_item_id:
                item_id = item_code_to_item_id[item_code]
                values = list(treeview.item(item_id, 'values'))

                col_indices = {col: idx for idx, col in enumerate(treeview['columns'])}
                # 치환 (사고보상건 등)
                substitution_type_processed = substitution_type.replace("사고보상건(분실, 파손 등)", "사고보상건(분실,파손등)")
                substitution_type_processed = substitution_type_processed.replace("ERP-iU기본", "erp_iu_basic")
                substitution_type_processed = (substitution_type_processed
                                               .replace(' ', '_')
                                               .replace('(', '')
                                               .replace(')', '')
                                               .replace(',', '')
                                               .replace('-', '_')
                                               .replace('/', '_')
                                               .lower())

                mapping_dict = {
                    '기부': 'donation',
                    '무상지원': 'free',
                    '사내소비': 'internal_use',
                    '샘플': 'sample',
                    '직원선물대': 'employee_gift',
                    '코드변경': 'code_change',
                    'loss': 'loss',
                    '계정대체유형erp_iu_basic': 'account_substitution',
                    '사고보상건분실파손등': 'accident_compensation',
                    '유통기한경과': 'expired',
                    '재고조정': 'inventory_adjustment',
                    '정기재고실사': 'regular_inventory_check',
                    '클레임처리': 'claim_processing'
                }

                col_name_value = mapping_dict.get(substitution_type_processed, substitution_type_processed)
                col_name_quantity = f"transfer_out_{col_name_value}_quantity"
                col_name_amount   = f"transfer_out_{col_name_value}_amount"

                if col_name_quantity in col_indices:
                    idx_quantity = col_indices[col_name_quantity]
                    existing_value_str = values[idx_quantity]
                    existing_value = safe_float_from_string(existing_value_str) or 0.0
                    new_value = existing_value + total_quantity
                    values[idx_quantity] = format_numeric_value(new_value)

                    # 출고내역 단가 가져오기
                    idx_outgoing_unit_price = col_indices.get('outgoing_unit_price')
                    if idx_outgoing_unit_price is not None:
                        outgoing_unit_price_str = values[idx_outgoing_unit_price]
                        outgoing_unit_price = safe_float_from_string(outgoing_unit_price_str) or 0.0
                    else:
                        outgoing_unit_price = 0.0

                    amount = outgoing_unit_price * new_value
                    if col_name_amount in col_indices:
                        idx_amount = col_indices[col_name_amount]
                        existing_amount_str = values[idx_amount]
                        existing_amount = safe_float_from_string(existing_amount_str) or 0.0
                        new_amount = existing_amount + amount
                        values[idx_amount] = format_numeric_value(new_amount)

                        treeview.item(item_id, values=values)
                        matched = True
                    else:
                        reason = f"컬럼 '{col_name_amount}' 없음"
                else:
                    reason = f"컬럼 '{col_name_quantity}' 없음"
            else:
                reason = "트리뷰에 해당 item_code 없음"

            if not matched:
                db_debug_rows.append([item_code, substitution_type, total_quantity, reason])
                unmatched_codes.append(item_code)

        # if db_debug_rows:
        #     import openpyxl
        #     from openpyxl import Workbook
        #     import os

        #     wb = Workbook()
        #     ws = wb.active
        #     ws.title = "Unprocessed Rows"
        #     ws.append(["품목코드", "substitution_type", "총수량", "이유"])
        #     for row in db_debug_rows:
        #         ws.append(row)

        #     timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        #     filename = f"이체출고_업데이트되지_않은_데이터_{timestamp}.xlsx"
        #     wb.save(filename)
        #     os.startfile(filename)

        #     msg = f"업데이트되지 않은 항목이 {len(db_debug_rows)}개 있습니다. 파일: {filename}"
        #     messagebox.showinfo("완료", msg)
        # else:
        #     messagebox.showinfo("완료", "이체출고 데이터 로딩 및 업데이트가 모두 완료되었습니다.")

        # print(f"이체출고 불러오기 진행 완료. 매칭되지 않은 품목코드: {len(unmatched_codes)}개")

        # 합계, 음영, 숫자포맷
        update_total_row(treeview)
        reapply_row_tags(treeview)
        format_numeric_columns(treeview, numeric_columns_treeview0)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()




def calculate_current_inventory(treeview):
    """
    '현재고 계산하기' 버튼 클릭 시 실행되는 함수입니다.
    변경사항:
    1) 현재고_단가가 0/빈 값일 때 이월재고_단가를 가져오는 로직.
    2) 이월재고_단가도 0이면 입고내역_단가를 사용하도록 변경.
    """
    item_ids = treeview.get_children()
    unmatched_items = 0

    columns = treeview['columns']
    col_indices = {col: idx for idx, col in enumerate(columns)}

    total_row_id = None
    for item_id in item_ids:
        if 'totalrow' in treeview.item(item_id, 'tags'):
            total_row_id = item_id
            break

    for item_id in item_ids:
        if item_id == total_row_id:
            continue

        values = list(treeview.item(item_id, 'values'))
        try:
            idx_outgoing_unit_price = col_indices.get('outgoing_unit_price')
            idx_current_unit_price = col_indices.get('current_unit_price')
            idx_beginning_unit_price = col_indices.get('beginning_unit_price')  # 이월재고 단가 인덱스
            idx_incoming_unit_price = col_indices.get('incoming_unit_price')    # 입고내역 단가 인덱스

            if idx_outgoing_unit_price is not None and idx_current_unit_price is not None:
                outgoing_unit_price_str = values[idx_outgoing_unit_price]
                values[idx_current_unit_price] = outgoing_unit_price_str
            else:
                unmatched_items += 1
                continue

            def safe_float(val):
                if val is None or val == '':
                    return 0.0
                return float(str(val).replace(',', '').strip())

            def get_idx(col_name):
                return col_indices.get(col_name)

            # 수량 컬럼 인덱스들
            idx_beginning_quantity = get_idx('beginning_quantity')
            idx_incoming_quantity = get_idx('incoming_quantity')
            idx_transfer_in_free_quantity = get_idx('transfer_in_free_quantity')
            idx_transfer_in_code_change_quantity = get_idx('transfer_in_code_change_quantity')

            idx_outgoing_quantity = get_idx('outgoing_quantity')
            idx_transfer_out_loss_quantity = get_idx('transfer_out_loss_quantity')
            idx_transfer_out_account_substitution_quantity = get_idx('transfer_out_account_substitution_quantity')
            idx_transfer_out_donation_quantity = get_idx('transfer_out_donation_quantity')
            idx_transfer_out_free_quantity = get_idx('transfer_out_free_quantity')
            idx_transfer_out_accident_compensation_quantity = get_idx('transfer_out_accident_compensation_quantity')
            idx_transfer_out_internal_use_quantity = get_idx('transfer_out_internal_use_quantity')
            idx_transfer_out_sample_quantity = get_idx('transfer_out_sample_quantity')
            idx_transfer_out_expired_quantity = get_idx('transfer_out_expired_quantity')
            idx_transfer_out_inventory_adjustment_quantity = get_idx('transfer_out_inventory_adjustment_quantity')
            idx_transfer_out_regular_inventory_check_quantity = get_idx('transfer_out_regular_inventory_check_quantity')
            idx_transfer_out_employee_gift_quantity = get_idx('transfer_out_employee_gift_quantity')
            idx_transfer_out_code_change_quantity = get_idx('transfer_out_code_change_quantity')
            idx_transfer_out_claim_processing_quantity = get_idx('transfer_out_claim_processing_quantity')

            required_indices = [
                idx_beginning_quantity, idx_incoming_quantity,
                idx_transfer_in_free_quantity, idx_transfer_in_code_change_quantity,
                idx_outgoing_quantity, idx_transfer_out_loss_quantity,
                idx_transfer_out_account_substitution_quantity, idx_transfer_out_donation_quantity,
                idx_transfer_out_free_quantity, idx_transfer_out_accident_compensation_quantity,
                idx_transfer_out_internal_use_quantity, idx_transfer_out_sample_quantity,
                idx_transfer_out_expired_quantity, idx_transfer_out_inventory_adjustment_quantity,
                idx_transfer_out_regular_inventory_check_quantity, idx_transfer_out_employee_gift_quantity,
                idx_transfer_out_code_change_quantity, idx_transfer_out_claim_processing_quantity,
                idx_beginning_unit_price,  # 이월재고단가 인덱스 추가
                idx_incoming_unit_price    # 입고내역 단가 인덱스 추가
            ]

            if None in required_indices:
                unmatched_items += 1
                continue

            # 수량 값 가져오기
            beginning_quantity = safe_float(values[idx_beginning_quantity])
            incoming_quantity = safe_float(values[idx_incoming_quantity])
            transfer_in_free_quantity = safe_float(values[idx_transfer_in_free_quantity])
            transfer_in_code_change_quantity = safe_float(values[idx_transfer_in_code_change_quantity])

            outgoing_quantity = safe_float(values[idx_outgoing_quantity])
            transfer_out_loss_quantity = safe_float(values[idx_transfer_out_loss_quantity])
            transfer_out_account_substitution_quantity = safe_float(values[idx_transfer_out_account_substitution_quantity])
            transfer_out_donation_quantity = safe_float(values[idx_transfer_out_donation_quantity])
            transfer_out_free_quantity = safe_float(values[idx_transfer_out_free_quantity])
            transfer_out_accident_compensation_quantity = safe_float(values[idx_transfer_out_accident_compensation_quantity])
            transfer_out_internal_use_quantity = safe_float(values[idx_transfer_out_internal_use_quantity])
            transfer_out_sample_quantity = safe_float(values[idx_transfer_out_sample_quantity])
            transfer_out_expired_quantity = safe_float(values[idx_transfer_out_expired_quantity])
            transfer_out_inventory_adjustment_quantity = safe_float(values[idx_transfer_out_inventory_adjustment_quantity])
            transfer_out_regular_inventory_check_quantity = safe_float(values[idx_transfer_out_regular_inventory_check_quantity])
            transfer_out_employee_gift_quantity = safe_float(values[idx_transfer_out_employee_gift_quantity])
            transfer_out_code_change_quantity = safe_float(values[idx_transfer_out_code_change_quantity])
            transfer_out_claim_processing_quantity = safe_float(values[idx_transfer_out_claim_processing_quantity])

            # 현재고_수량 계산
            current_quantity = (
                beginning_quantity + incoming_quantity +
                transfer_in_free_quantity + transfer_in_code_change_quantity -
                outgoing_quantity - transfer_out_loss_quantity -
                transfer_out_account_substitution_quantity - transfer_out_donation_quantity -
                transfer_out_free_quantity - transfer_out_accident_compensation_quantity -
                transfer_out_internal_use_quantity - transfer_out_sample_quantity -
                transfer_out_expired_quantity - transfer_out_inventory_adjustment_quantity -
                transfer_out_regular_inventory_check_quantity - transfer_out_employee_gift_quantity -
                transfer_out_code_change_quantity - transfer_out_claim_processing_quantity
            )

            idx_current_quantity = col_indices.get('current_quantity')
            if idx_current_quantity is not None:
                values[idx_current_quantity] = format_numeric_value(current_quantity)
            else:
                unmatched_items += 1
                continue

            # 현재고_단가 가져오기
            current_unit_price = safe_float(values[idx_current_unit_price])

            # 추가 로직: 현재고_단가가 0이면 이월재고_단가 확인
            if current_unit_price == 0:
                beginning_unit_price = safe_float(values[idx_beginning_unit_price])
                if beginning_unit_price != 0:
                    current_unit_price = beginning_unit_price
                    values[idx_current_unit_price] = format_numeric_value(current_unit_price)
                else:
                    # 이월재고단가도 0이면 입고내역_단가 확인
                    incoming_unit_price = safe_float(values[idx_incoming_unit_price])
                    if incoming_unit_price != 0:
                        current_unit_price = incoming_unit_price
                        values[idx_current_unit_price] = format_numeric_value(current_unit_price)
                    # 모두 0이면 current_unit_price 그대로 0 유지

            # 현재고_금액 계산
            current_amount = round(current_unit_price * current_quantity, 0)
            idx_current_amount = col_indices.get('current_amount')
            if idx_current_amount is not None:
                values[idx_current_amount] = format_numeric_value(current_amount)
            else:
                unmatched_items += 1
                continue

            treeview.item(item_id, values=values)

        except Exception as e:
            unmatched_items += 1
            print(f"아이템 처리 중 오류 발생 (ID: {item_id}): {e}")
            continue

    update_total_row(treeview)
    reapply_row_tags(treeview)
    format_numeric_columns(treeview, numeric_columns_treeview0)
    # messagebox.showinfo("완료", "현재고 계산이 완료되었습니다.")
    # print(f"현재고 계산 완료. 처리되지 않은 항목 수: {unmatched_items}")



def perform_verification(treeview):
    """
    '검증' 버튼 클릭 시 실행되는 함수
    - ver_amount가 0이 아닌 모든 로우들에 대해 차이금액을 클레임처리에 반영
    """
    item_ids = treeview.get_children()
    unmatched_items = 0  # 처리되지 않은 항목 수를 세기 위한 변수

    # 컬럼 인덱스 매핑
    columns = treeview['columns']
    col_indices = {col: idx for idx, col in enumerate(columns)}

    # TRUE/FALSE 검증 컬럼 목록
    boolean_verification_columns = [
        'verification_inventory',
        'verification_negative_inventory',
        'verification_outgoing',
        'verification_return',
        'verification_negative_stock_check',
        'verification_incentive',  # 새로 추가
        'verification_free_support',  # 새로 추가
        'verification_inventory_unit_price'
    ]

    # 합계 행 제거
    for item_id in item_ids:
        if 'totalrow' in treeview.item(item_id, 'tags'):
            treeview.delete(item_id)
    # 합계 행 삭제 후 아이템 목록 갱신
    item_ids = treeview.get_children()

    # 모든 TRUE/FALSE 검증 컬럼의 최종 결과(True/False)를 저장할 딕셔너리
    boolean_verification_results = {col: True for col in boolean_verification_columns}

    # 미리 함수 정의(인덱스 얻기 등)
    def get_idx(col_name):
        idx = col_indices.get(col_name)
        if idx is None:
            print(f"컬럼 '{col_name}'의 인덱스를 찾을 수 없습니다.")
        return idx

    def get_numeric_value(values_list, idx):
        """문자열 -> float 변환(실패시 0.0)."""
        if idx is None:
            return 0.0
        val_str = values_list[idx]
        return safe_float_from_string(val_str) or 0.0

    # 자주 쓰는 컬럼 인덱스들 미리
    idx_ver_amount = get_idx('verification_amount')
    idx_transfer_out_claim_processing_amount = get_idx('transfer_out_claim_processing_amount')

    # 이체출고 금액 컬럼들
    transfer_out_amount_cols = [
        'transfer_out_loss_amount',
        'transfer_out_account_substitution_amount',
        'transfer_out_donation_amount',
        'transfer_out_free_amount',
        'transfer_out_accident_compensation_amount',
        'transfer_out_internal_use_amount',
        'transfer_out_sample_amount',
        'transfer_out_expired_amount',
        'transfer_out_inventory_adjustment_amount',
        'transfer_out_regular_inventory_check_amount',
        'transfer_out_employee_gift_amount',
        'transfer_out_code_change_amount',
        'transfer_out_claim_processing_amount'
    ]

    # -- (A) 한 로우의 검증 로직을 별도 함수로 --
    def verify_item(values):
        """
        values: 트리뷰 한 row의 values(list)
        검증 로직을 수행하고, 변경된 values를 리턴.
        """
         # 기존 인덱스들
        idx_current_quantity = get_idx('current_quantity')
        idx_current_amount = get_idx('current_amount')
        idx_outgoing_quantity = get_idx('outgoing_quantity')
        idx_outgoing_amount = get_idx('outgoing_amount')
        idx_outgoing_unit_price = get_idx('outgoing_unit_price')
        idx_current_unit_price = get_idx('current_unit_price')

         # 검증 컬럼 인덱스들
        idx_ver_stock = get_idx('verification_inventory')
        idx_ver_negative_stock = get_idx('verification_negative_inventory')
        idx_ver_outgoing = get_idx('verification_outgoing')
        idx_ver_return = get_idx('verification_return')
        idx_ver_negative_stock_check = get_idx('verification_negative_stock_check')
        idx_ver_incentive = get_idx('verification_incentive')  # 새로 추가
        idx_ver_free_support = get_idx('verification_free_support')  # 새로 추가
        idx_ver_quantity = get_idx('verification_quantity')
        idx_ver_amount = get_idx('verification_amount')
        idx_ver_unit_price = get_idx('verification_inventory_unit_price')
        # (idx_ver_amount 는 바깥에서 이미 있음)

        # 새로운 검증을 위한 추가 인덱스들
        idx_incentive_amount = get_idx('incentive_amount')
        idx_incoming_quantity = get_idx('incoming_quantity')
        idx_transfer_in_free_quantity = get_idx('transfer_in_free_quantity')
        idx_beginning_quantity = get_idx('beginning_quantity')


        if None in [
            idx_current_quantity, idx_current_amount,
            idx_outgoing_quantity, idx_outgoing_amount,
            idx_outgoing_unit_price, idx_current_unit_price,
            idx_ver_stock, idx_ver_negative_stock, idx_ver_outgoing,
            idx_ver_return, idx_ver_negative_stock_check,
            idx_ver_incentive, idx_ver_free_support,  # 새로 추가된 컬럼들
            idx_ver_quantity, idx_ver_amount, idx_ver_unit_price
        ]:
            return values  # 필요한 인덱스 하나라도 없으면 그대로 반환

        # 숫자값 읽기
        current_quantity = get_numeric_value(values, idx_current_quantity)
        current_amount = get_numeric_value(values, idx_current_amount)
        outgoing_quantity = get_numeric_value(values, idx_outgoing_quantity)
        outgoing_amount = get_numeric_value(values, idx_outgoing_amount)
        outgoing_unit_price = get_numeric_value(values, idx_outgoing_unit_price)
        current_unit_price = get_numeric_value(values, idx_current_unit_price)

        # 새로운 검증을 위한 값들
        incentive_amount = get_numeric_value(values, idx_incentive_amount)
        incoming_quantity = get_numeric_value(values, idx_incoming_quantity)
        transfer_in_free_quantity = get_numeric_value(values, idx_transfer_in_free_quantity)
        beginning_quantity = get_numeric_value(values, idx_beginning_quantity)


        # 검증_재고검증
        if current_quantity > 0:
            ver_stock = 'TRUE' if current_amount > 0 else 'FALSE'
        else:
            ver_stock = 'TRUE'

        # 검증_음수재고검증
        if current_quantity == 0:
            ver_negative_stock = 'TRUE' if current_amount == 0 else 'FALSE'
        else:
            ver_negative_stock = 'TRUE'

        # 검증_출고검증
        if outgoing_quantity > 0:
            ver_outgoing = 'TRUE' if outgoing_amount > 0 else 'FALSE'
        else:
            ver_outgoing = 'TRUE'

        # 검증_반품검증
        if outgoing_quantity < 0:
            ver_return = 'TRUE' if outgoing_amount < 0 else 'FALSE'
        else:
            ver_return = 'TRUE'

        # 검증_재고음수체크
        ver_negative_stock_check = 'TRUE' if current_quantity >= 0 else 'FALSE'

        # === 새로운 검증 로직들 ===
        # 장려금검증 = 장려금 금액이 0이 아닌경우 and 입고내역수량이 0 인 경우 면 FALSE, 아니면 TRUE
        if incentive_amount != 0 and incoming_quantity == 0:
            ver_incentive = 'FALSE'
        else:
            ver_incentive = 'TRUE'

        # 무상지원 조정 = 무상지원 수량이 0이 아닌경우 and 이월재고수량+입고내역수량이 0인경우 면 FALSE, 아니면 TRUE
        if transfer_in_free_quantity != 0 and (beginning_quantity + incoming_quantity) == 0:
            ver_free_support = 'FALSE'
        else:
            ver_free_support = 'TRUE'

        # --- 검증_수량검증 ---
        idx_beginning_quantity = get_idx('beginning_quantity')
        idx_incoming_quantity = get_idx('incoming_quantity')
        idx_transfer_in_free_quantity = get_idx('transfer_in_free_quantity')
        idx_transfer_in_code_change_quantity = get_idx('transfer_in_code_change_quantity')

        transfer_out_quantity_cols = [
            'transfer_out_loss_quantity',
            'transfer_out_account_substitution_quantity',
            'transfer_out_donation_quantity',
            'transfer_out_free_quantity',
            'transfer_out_accident_compensation_quantity',
            'transfer_out_internal_use_quantity',
            'transfer_out_sample_quantity',
            'transfer_out_expired_quantity',
            'transfer_out_inventory_adjustment_quantity',
            'transfer_out_regular_inventory_check_quantity',
            'transfer_out_employee_gift_quantity',
            'transfer_out_code_change_quantity',
            'transfer_out_claim_processing_quantity'
        ]

        sum_quantities = (
            get_numeric_value(values, idx_beginning_quantity) +
            get_numeric_value(values, idx_incoming_quantity) +
            get_numeric_value(values, idx_transfer_in_free_quantity) +
            get_numeric_value(values, idx_transfer_in_code_change_quantity)
        )

        subtract_quantities = outgoing_quantity
        for col_name in transfer_out_quantity_cols:
            idx_col = get_idx(col_name)
            subtract_quantities += get_numeric_value(values, idx_col)

        ver_quantity = sum_quantities - subtract_quantities - current_quantity

        # --- 검증_금액검증 ---
        idx_beginning_amount = get_idx('beginning_amount')
        idx_incoming_amount = get_idx('incoming_amount')
        idx_misc_profit_amount = get_idx('misc_profit_amount')
        idx_incentive_amount = get_idx('incentive_amount')
        idx_transfer_in_code_change_amount = get_idx('transfer_in_code_change_amount')

        transfer_out_amount_cols = [
            'transfer_out_loss_amount',
            'transfer_out_account_substitution_amount',
            'transfer_out_donation_amount',
            'transfer_out_free_amount',
            'transfer_out_accident_compensation_amount',
            'transfer_out_internal_use_amount',
            'transfer_out_sample_amount',
            'transfer_out_expired_amount',
            'transfer_out_inventory_adjustment_amount',
            'transfer_out_regular_inventory_check_amount',
            'transfer_out_employee_gift_amount',
            'transfer_out_code_change_amount',
            'transfer_out_claim_processing_amount'
        ]

        sum_amounts = (
            get_numeric_value(values, idx_beginning_amount) +
            get_numeric_value(values, idx_incoming_amount) +
            get_numeric_value(values, idx_misc_profit_amount) -
            get_numeric_value(values, idx_incentive_amount) +
            get_numeric_value(values, idx_transfer_in_code_change_amount)
        )

        subtract_amounts = outgoing_amount
        for col_name in transfer_out_amount_cols:
            idx_col = get_idx(col_name)
            subtract_amounts += get_numeric_value(values, idx_col)

        ver_amount = sum_amounts - subtract_amounts - current_amount

        # 검증_재고단가검증
        if round(outgoing_unit_price, 2) == round(current_unit_price, 2):
            ver_unit_price = 'TRUE'
        else:
            ver_unit_price = 'FALSE'

        # 값 저장
        values[idx_ver_stock] = ver_stock
        values[idx_ver_negative_stock] = ver_negative_stock
        values[idx_ver_outgoing] = ver_outgoing
        values[idx_ver_return] = ver_return
        values[idx_ver_negative_stock_check] = ver_negative_stock_check
        values[idx_ver_incentive] = ver_incentive  # 새로 추가
        values[idx_ver_free_support] = ver_free_support  # 새로 추가
        values[idx_ver_quantity] = format_numeric_value(ver_quantity)
        values[idx_ver_amount] = format_numeric_value(ver_amount)
        values[idx_ver_unit_price] = ver_unit_price

        return values
    
    # 검증 실행 및 결과 처리 (기존 로직과 동일)
    idx_ver_amount = get_idx('verification_amount')
    idx_transfer_out_claim_processing_amount = get_idx('transfer_out_claim_processing_amount')

    # -- (B) 1차 루프: 전 아이템 검증 실행, 차이금액이 0이 아닌 경우 기록 --
    lines_to_fix = []  # (item_id, ver_amount_val) 형태로 저장
    for item_id in item_ids:
        values = list(treeview.item(item_id, 'values'))
        try:
            # 먼저 검증
            new_values = verify_item(values)
            # 트리뷰 아이템에 반영
            treeview.item(item_id, values=new_values)

            # TRUE/FALSE 체크
            for col_name in boolean_verification_columns:
                idx_col = get_idx(col_name)
                if idx_col is not None and new_values[idx_col] == 'FALSE':
                    boolean_verification_results[col_name] = False

            # ver_amount 확인 (0이 아닌 모든 값)
            if idx_ver_amount is not None:
                ver_amount_val = safe_float_from_string(new_values[idx_ver_amount]) or 0.0
                # *** 수정된 부분: 모든 0이 아닌 차이금액을 처리 대상으로 ***
                if ver_amount_val != 0:
                    lines_to_fix.append((item_id, ver_amount_val))

        except Exception as e:
            unmatched_items += 1
            print(f"아이템 처리 중 오류 발생 (ID: {item_id}): {e}")
            continue

    # -- (C) 차이금액이 있는 아이템들이 있다면, 클레임처리에 반영할지 확인 --
    if lines_to_fix:
        # 차이금액 총합 계산
        total_difference = sum(diff_val for _, diff_val in lines_to_fix)
        
        # ±2원 이하와 그 이상으로 분리
        small_differences = []  # ±2원 이하
        large_differences = []  # ±2원 초과
        
        for item_id, diff_val in lines_to_fix:
            if abs(diff_val) <= 2:
                small_differences.append((item_id, diff_val))
            else:
                large_differences.append((item_id, diff_val))
        
        # 메시지 구성
        message_parts = [
            "금액검증과정에서 차이금액이 발견되었습니다.\n",
            f"총 {len(lines_to_fix)}개 항목에서 차이금액: {int(total_difference):,}원\n"
        ]
        
        # ±2원 이하 합산 표시
        if small_differences:
            small_total = sum(diff_val for _, diff_val in small_differences)
            message_parts.append(f"단수차이(±2원 이하): {len(small_differences)}개 항목, 합계 {int(small_total):,}원\n")
        
        # ±2원 초과 개별 리스트 표시
        if large_differences:
            message_parts.append(f"\n주요 차이금액(±2원 초과):")
            for item_id, diff_val in large_differences:
                # 해당 행의 품목코드와 품목명 가져오기
                item_values = treeview.item(item_id, 'values')
                item_code = item_values[0] if len(item_values) > 0 else "Unknown"
                item_name = item_values[1] if len(item_values) > 1 else "Unknown"
                message_parts.append(f"• {item_code} ({item_name}): {int(diff_val):,}원")
        
        message_parts.append(f"\n모든 차이금액을 해당 항목의 이체출고 클레임처리 금액으로 반영하시겠습니까?")
        
        # 최종 메시지 조합
        final_message = "\n".join(message_parts)
        
        answer = messagebox.askyesno("금액차이 발견", final_message)
        if answer:
            # (C-1) 예(Yes) -> 모든 line에 대해서 클레임처리금액 += ver_amount
            for (item_id, diff_val) in lines_to_fix:
                vals = list(treeview.item(item_id, 'values'))
                if idx_transfer_out_claim_processing_amount is not None:
                    old_claim_amt = safe_float_from_string(vals[idx_transfer_out_claim_processing_amount]) or 0.0
                    new_claim_amt = old_claim_amt + diff_val
                    vals[idx_transfer_out_claim_processing_amount] = str(new_claim_amt)
                # 다시 검증(2차) → ver_amount 갱신
                vals_after = verify_item(vals)
                # 저장
                treeview.item(item_id, values=vals_after)

    # -- (D) 최종적으로 합계 행 추가 및 마무리 --
    update_total_row(treeview)

    # 합계 행에 TRUE/FALSE 결과 업데이트
    total_row_id = None
    for item_id in treeview.get_children():
        if 'totalrow' in treeview.item(item_id, 'tags'):
            total_row_id = item_id
            break

    if total_row_id:
        total_values = list(treeview.item(total_row_id, 'values'))
        for col_name in boolean_verification_columns:
            idx_col = col_indices.get(col_name)
            if idx_col is not None:
                total_values[idx_col] = 'TRUE' if boolean_verification_results[col_name] else 'FALSE'
        treeview.item(total_row_id, values=total_values)

    # 음영 재적용
    reapply_row_tags(treeview)

    # 숫자 포맷팅 적용
    format_numeric_columns(treeview, numeric_columns_treeview0)

    # 정렬 기능 적용 (트리뷰 정렬 함수 호출)
    for col in columns:
        treeview.heading(col, command=lambda _col=col: treeview_sort_column(treeview, _col, False))

    # 완료 메시지
    # messagebox.showinfo("완료", "검증 계산이 완료되었습니다.")
    # print(f"검증 계산 완료. 처리되지 않은 항목 수: {unmatched_items}")




def treeview_sort_column(tv, col, reverse):
    l = [(tv.set(k, col), k) for k in tv.get_children('')]
    try:
        l.sort(key=lambda t: float(t[0].replace(',', '')), reverse=reverse)
    except ValueError:
        l.sort(reverse=reverse)
    for index, (val, k) in enumerate(l):
        tv.move(k, '', index)
    tv.heading(col, command=lambda: treeview_sort_column(tv, col, not reverse))


def save_results(treeview):
    """
    트리뷰의 데이터를 데이터베이스에 저장하는 함수 (psycopg2.extras.execute_values로 대량 삽입 최적화).
    """
    from datetime import datetime, timedelta

    def get_previous_month():
        today = datetime.today()
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return last_month.strftime("%Y/%m")

    # 기준월 입력
    default_month = get_previous_month()
    reference_month = simpledialog.askstring("기준월 선택", "저장할 기준월을 입력하세요 (예: 2025/01):", initialvalue=default_month)
    if not reference_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    global columns  # columns 리스트 전역 사용
    column_names = [col[0] for col in columns]

    # 'verification_inventory'가 TRUE인지 체크
    col_indices = {col: idx for idx, col in enumerate(column_names)}
    verification_idx = col_indices.get('verification_inventory')
    if verification_idx is None:
        messagebox.showerror("오류", "'verification_inventory' 컬럼을 찾을 수 없습니다.")
        return

    has_true = False
    for item_id in treeview.get_children():
        values = treeview.item(item_id, 'values')
        if len(values) != len(column_names):
            messagebox.showerror(
                "오류", 
                f"컬럼 수와 데이터 수가 일치하지 않습니다. "
                f"컬럼 수: {len(column_names)}, 데이터 수: {len(values)}"
            )
            return
        if values[verification_idx] == 'TRUE':
            has_true = True
            break

    if not has_true:
        messagebox.showinfo("알림", "현재 검증과정을 거치지 않았습니다. 검증 후 저장해주세요.")
        return

    try:
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return

        cursor = conn.cursor()

        # 접미사 규칙에 따라 numeric 컬럼 식별
        numeric_columns = [
            col[0] for col in columns
            if (
                col[0].endswith('_quantity') or
                col[0].endswith('_unit_price') or
                col[0].endswith('_amount') or
                col[0].endswith('_price') or
                col[0].endswith('_inventory') or
                col[0].endswith('_difference_quantity') or
                col[0].startswith('inventory_inspection_')
            )
            and col[0] not in [
                'verification_inventory', 'verification_negative_inventory', 'verification_outgoing',
                'verification_return', 'verification_negative_stock_check', 'verification_inventory_unit_price'
            ]
        ]
        numeric_columns.extend(['verification_quantity', 'verification_amount'])  # 추가로 숫자 처리

        # 테이블 생성 (없으면 생성)
        column_definitions = []
        for col_name in column_names:
            if col_name in numeric_columns:
                column_definitions.append(f"{col_name} NUMERIC")
            else:
                column_definitions.append(f"{col_name} VARCHAR(255)")
        column_definitions.append("reference_month VARCHAR(7)")

        create_table_query = f"""
            CREATE TABLE IF NOT EXISTS mds_monthly_inventory_transactions (
                {', '.join(column_definitions)}
            )
        """
        cursor.execute(create_table_query)

        # 기존 데이터 삭제 여부 확인
        cursor.execute(
            "SELECT COUNT(*) FROM mds_monthly_inventory_transactions WHERE reference_month = %s",
            (reference_month,)
        )
        existing_count = cursor.fetchone()[0]
        if existing_count > 0:
            response = messagebox.askyesno(
                "데이터 존재 확인",
                f"{reference_month}의 기존 데이터가 있습니다. 삭제하고 진행하시겠습니까?"
            )
            if response:
                cursor.execute(
                    "DELETE FROM mds_monthly_inventory_transactions WHERE reference_month = %s",
                    (reference_month,)
                )
                conn.commit()
                print(f"{reference_month} 기존 데이터 삭제 완료.")
            else:
                messagebox.showinfo("취소", "데이터 저장이 취소되었습니다.")
                return

        # (1) INSERT 쿼리 준비
        #     execute_values 사용 시 "VALUES %s" 형태를 쓰고, 실제 값들은 파라미터로 넘김
        insert_cols = ", ".join(column_names + ["reference_month"])
        insert_query = f"""
            INSERT INTO mds_monthly_inventory_transactions ({insert_cols})
            VALUES %s
        """

        # (2) 트리뷰에서 행 추출 → data_to_insert (list of tuples)
        data_to_insert = []
        for item_id in treeview.get_children():
            if 'totalrow' in treeview.item(item_id, 'tags'):
                continue  # 합계 행 제외

            row_values = list(treeview.item(item_id, 'values'))
            if len(row_values) != len(column_names):
                messagebox.showerror(
                    "오류",
                    f"컬럼 수와 데이터 수가 일치하지 않습니다. "
                    f"컬럼 수: {len(column_names)}, 데이터 수: {len(row_values)}"
                )
                return

            processed_values = []
            for idx, col_name in enumerate(column_names):
                raw_val = row_values[idx]
                if col_name in numeric_columns:
                    if not raw_val:
                        val = None
                    else:
                        try:
                            clean_str = str(raw_val).replace(',', '').replace(' ', '').replace('None', '0')
                            val = float(clean_str)
                        except ValueError:
                            val = None
                    processed_values.append(val)
                else:
                    processed_values.append(raw_val if raw_val != '' else None)

            # reference_month 추가
            processed_values.append(reference_month)
            data_to_insert.append(tuple(processed_values))

        # (3) execute_values로 대량 삽입
        from psycopg2.extras import execute_values
        # page_size=1000 → 1,000행씩 묶음
        execute_values(cursor, insert_query, data_to_insert, page_size=1000)
        conn.commit()

        messagebox.showinfo("성공", f"{reference_month} 데이터가 저장되었습니다.")
        print(f"{reference_month} 데이터 저장 완료. 저장된 항목 수: {len(data_to_insert)}")

    except Exception as e:
        if conn:
            conn.rollback()
        messagebox.showerror("오류", f"데이터 저장 중 오류가 발생했습니다: {e}")
        print(f"오류 발생: {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



def load_results(treeview, loaded_month_label):
    """
    DB에서 mds_monthly_inventory_transactions 데이터를 불러와
    트리뷰에 로딩하는 함수 (최적화 버전).
    - 쿼리에서 ORDER BY item_code -> Python 정렬 생략
    - 루프 내 숫자포맷만 적용 -> 별도 format_numeric_columns 생략
    - 합계는 루프에서 누적 후 한 번만 insert
    """
    from datetime import datetime, timedelta

    def get_previous_month():
        today = datetime.today()
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return last_month.strftime("%Y/%m")

    default_month = get_previous_month()
    reference_month = simpledialog.askstring(
        "기준월 선택", 
        "불러올 기준월을 입력하세요 (예: 2025/01):", 
        initialvalue=default_month
    )
    if not reference_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    try:
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "데이터베이스 연결 실패")
            return
        cursor = conn.cursor()

        # 데이터 존재 여부 확인
        cursor.execute("""
            SELECT COUNT(*) 
            FROM mds_monthly_inventory_transactions 
            WHERE reference_month = %s
        """, (reference_month,))
        count = cursor.fetchone()[0]
        if count == 0:
            messagebox.showinfo("알림", f"{reference_month} 데이터가 존재하지 않습니다.")
            return

        # 컬럼 및 숫자컬럼
        columns = treeview['columns']
        numeric_columns = numeric_columns_treeview0  # 전역 numeric 리스트

        # 쿼리에서 item_code 오름차순 정렬 => Python 정렬 스킵 가능
        cols_str = ', '.join(columns)
        select_query = f"""
            SELECT {cols_str}
            FROM mds_monthly_inventory_transactions
            WHERE reference_month = %s
            ORDER BY item_code
        """
        cursor.execute(select_query, (reference_month,))
        rows = cursor.fetchall()

        # 기존 행 초기화
        treeview.delete(*treeview.get_children())

        # ----- (A) 합계 계산을 위한 딕셔너리 -----
        # numeric 컬럼별로 합계 누적
        sum_dict = {col: 0.0 for col in numeric_columns}

        # ----- (B) 행 삽입 -----
        for idx_row, row in enumerate(rows):
            row_values = []
            for idx_col, val in enumerate(row):
                col_id = columns[idx_col]

                if col_id in numeric_columns:
                    # 숫자 포맷
                    display_val = format_numeric_value(val) if val is not None else ''
                    row_values.append(display_val)

                    # 합계 누적
                    if val not in (None, ''):
                        try:
                            sum_dict[col_id] += float(val)
                        except ValueError:
                            pass
                else:
                    row_values.append(val if val is not None else '')

            tag = 'evenrow' if idx_row % 2 == 0 else 'oddrow'
            treeview.insert('', 'end', values=row_values, tags=(tag,))

        # 태그별 색상
        treeview.tag_configure('oddrow', background='lightgray')
        treeview.tag_configure('evenrow', background='white')

        # ----- (C) 합계행 삽입 (맨 앞이나 맨 뒤 중 선택) -----
        sum_values = []
        for col_id in columns:
            if col_id in numeric_columns:
                sum_values.append(format_numeric_value(sum_dict[col_id]))
            else:
                sum_values.append('')  # 합계행 아닌 컬럼은 빈칸
        # 맨 위(인덱스 0)나 맨 뒤('end')에 삽입
        treeview.insert('', 0, values=sum_values, tags=('totalrow',))
        treeview.tag_configure('totalrow', background='yellow', font=('Arial', 10, 'bold'))

        # 로딩된 달 표시
        loaded_month_label.config(text=f"{reference_month} 데이터가 로드되었습니다. ({len(rows)} 행)")
        
        treeview.all_data = rows

        messagebox.showinfo("성공", f"{reference_month} 데이터가 로드되었습니다. 총 {len(rows)} 행")
        print(f"{reference_month} 데이터 로드 완료. 로드된 항목 수: {len(rows)}")

    except Exception as e:
        messagebox.showerror("오류", f"데이터 로드 중 오류가 발생했습니다: {e}")
        print(f"오류 발생: {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


# ====================================================================
# 7. load_ending_inventory_for_treeview0 함수 수정 (새로운 재고실사 컬럼들 계산)
# ====================================================================

def load_ending_inventory_for_treeview0(treeview, work_month=None):
    """
    Treeview0에 기말재고 데이터를 로드하여
    - '재고실사' (창고별 수량) 컬럼에 반영
    - 창고별 금액(=수량 * 현재고_단가) 컬럼도 계산
    - 합계금액 및 차이금액을 최종 계산

    :param treeview: 트리뷰0 위젯
    :param work_month: (옵션) 기본 표시할 기준월 (ex: "2025/01")
    """
# 기준월 입력
    if work_month:
        init_value = work_month
    else:
        init_value = ""

    reference_month = simpledialog.askstring(
        "기말재고 불러오기",
        "기말재고를 불러올 기준월을 입력하세요 (예: 2025/01):",
        initialvalue=init_value
    )
    if not reference_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 데이터 존재 여부 확인
        cursor.execute(
            "SELECT COUNT(*) FROM mds_monthly_inventory_status WHERE reference_month = %s",
            (reference_month,)
        )
        count = cursor.fetchone()[0]
        if count == 0:
            messagebox.showinfo("알림", f"{reference_month}에 해당하는 기말재고 데이터가 없습니다.")
            return

        # 기말재고 데이터 조회 (수량 합계)
        query = """
        SELECT
            item_code,
            
            -- 차산점 관련
            COALESCE(차산점_수량, 0) AS 차산점_수량,
            COALESCE(차산점a_수량, 0) AS 차산점a_수량,
            COALESCE("수입창고(차산)_수량", 0) AS 수입창고_차산_수량,
            COALESCE("수입창고(보관)_수량", 0) AS 수입창고_보관_수량,
            COALESCE(차산점반품_수량, 0) AS 차산점반품_수량,
            
            -- 청량리점 관련
            COALESCE(청량리점_수량, 0) AS 청량리점_수량,
            COALESCE(청량리반품_수량, 0) AS 청량리반품_수량,
            
            -- 이천점 관련
            COALESCE(이천점_수량, 0) AS 이천점_수량,
            COALESCE(케이터링_수량, 0) AS 케이터링_수량,
            COALESCE(이커머스_수량, 0) AS 이커머스_수량,
            COALESCE(이천점반품_수량, 0) AS 이천점반품_수량,
            
            -- 기타
            COALESCE(하남점_수량, 0) AS 하남점_수량,
            COALESCE(선매입창고_수량, 0) AS 선매입창고_수량
            
        FROM mds_monthly_inventory_status
        WHERE reference_month = %s
        """
        cursor.execute(query, (reference_month,))
        ending_inventory_data = cursor.fetchall()

        # 딕셔너리에 각 item_code별 수치 저장
        ending_inventory_dict = {}
        for row in ending_inventory_data:
            item_code = row[0]
            quantities = {
                '차산점': row[1],
                '차산점a': row[2],
                '수입창고_차산': row[3],
                '수입창고_보관': row[4],
                '차산점반품': row[5],
                '청량리점': row[6],
                '청량리반품': row[7],
                '이천점': row[8],
                '케이터링': row[9],
                '이커머스': row[10],
                '이천점반품': row[11],
                '하남점': row[12],
                '선매입창고': row[13]
            }
            ending_inventory_dict[item_code] = quantities

        # 컬럼 인덱스 매핑
        columns_list = [col[0] for col in columns]
        column_name_to_index = {col[0]: idx for idx, col in enumerate(columns)}

        # 트리뷰 업데이트
        for item_id in treeview.get_children():
            item_values = list(treeview.item(item_id)['values'])
            if not item_values:
                continue

            item_code_index = column_name_to_index.get('item_code')
            if item_code_index is None:
                continue

            this_item_code = item_values[item_code_index]
            quantities = ending_inventory_dict.get(this_item_code)
            if not quantities:
                continue

            # 현재고 단가 가져오기
            current_unit_price_idx = column_name_to_index.get('current_unit_price')
            current_unit_price = 0.0
            if current_unit_price_idx is not None and current_unit_price_idx < len(item_values):
                current_unit_price = safe_float_from_string(item_values[current_unit_price_idx]) or 0.0

            # === 개별 창고 수량 및 금액 설정 ===
            # 차산점
            if column_name_to_index.get('inventory_inspection_chasan') is not None:
                item_values[column_name_to_index['inventory_inspection_chasan']] = format_numeric_value(quantities['차산점'])
                item_values[column_name_to_index['inventory_inspection_chasan_amount']] = format_amount_value(quantities['차산점'] * current_unit_price)
            
            # 차산점A
            if column_name_to_index.get('inventory_inspection_chasan_a_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_chasan_a_quantity']] = format_numeric_value(quantities['차산점a'])
                item_values[column_name_to_index['inventory_inspection_chasan_a_amount']] = format_amount_value(quantities['차산점a'] * current_unit_price)
            
            # 수입창고(차산)
            if column_name_to_index.get('inventory_inspection_import_warehouse_chasan_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_import_warehouse_chasan_quantity']] = format_numeric_value(quantities['수입창고_차산'])
                item_values[column_name_to_index['inventory_inspection_import_warehouse_chasan_amount']] = format_amount_value(quantities['수입창고_차산'] * current_unit_price)
            
            # 수입창고(보관)
            if column_name_to_index.get('inventory_inspection_import_warehouse_storage_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_import_warehouse_storage_quantity']] = format_numeric_value(quantities['수입창고_보관'])
                item_values[column_name_to_index['inventory_inspection_import_warehouse_storage_amount']] = format_amount_value(quantities['수입창고_보관'] * current_unit_price)
            
            # 차산점반품
            if column_name_to_index.get('inventory_inspection_chasan_return_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_chasan_return_quantity']] = format_numeric_value(quantities['차산점반품'])
                item_values[column_name_to_index['inventory_inspection_chasan_return_amount']] = format_amount_value(quantities['차산점반품'] * current_unit_price)

            # === 차산점 합계 계산 ===
            chasan_sum_qty = quantities['차산점'] + quantities['수입창고_차산'] + quantities['수입창고_보관'] + quantities['차산점반품'] + quantities['차산점a']
            chasan_sum_amount = chasan_sum_qty * current_unit_price
            if column_name_to_index.get('inventory_inspection_chasan_sum_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_chasan_sum_quantity']] = format_numeric_value(chasan_sum_qty)
                item_values[column_name_to_index['inventory_inspection_chasan_sum_amount']] = format_amount_value(chasan_sum_amount)

            # 청량리점
            if column_name_to_index.get('inventory_inspection_cheongnyangni') is not None:
                item_values[column_name_to_index['inventory_inspection_cheongnyangni']] = format_numeric_value(quantities['청량리점'])
                item_values[column_name_to_index['inventory_inspection_cheongnyangni_amount']] = format_amount_value(quantities['청량리점'] * current_unit_price)
            
            # 청량리점반품
            if column_name_to_index.get('inventory_inspection_cheongnyangni_return_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_cheongnyangni_return_quantity']] = format_numeric_value(quantities['청량리반품'])
                item_values[column_name_to_index['inventory_inspection_cheongnyangni_return_amount']] = format_amount_value(quantities['청량리반품'] * current_unit_price)

            # === 청량리점 합계 계산 ===
            cheongnyangni_sum_qty = quantities['청량리점'] + quantities['청량리반품']
            cheongnyangni_sum_amount = cheongnyangni_sum_qty * current_unit_price
            if column_name_to_index.get('inventory_inspection_cheongnyangni_sum_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_cheongnyangni_sum_quantity']] = format_numeric_value(cheongnyangni_sum_qty)
                item_values[column_name_to_index['inventory_inspection_cheongnyangni_sum_amount']] = format_amount_value(cheongnyangni_sum_amount)

            # 이천점
            if column_name_to_index.get('inventory_inspection_icheon') is not None:
                item_values[column_name_to_index['inventory_inspection_icheon']] = format_numeric_value(quantities['이천점'])
                item_values[column_name_to_index['inventory_inspection_icheon_amount']] = format_amount_value(quantities['이천점'] * current_unit_price)
            
            # 케이터링
            if column_name_to_index.get('inventory_inspection_catering_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_catering_quantity']] = format_numeric_value(quantities['케이터링'])
                item_values[column_name_to_index['inventory_inspection_catering_amount']] = format_amount_value(quantities['케이터링'] * current_unit_price)
            
            # 이커머스
            if column_name_to_index.get('inventory_inspection_ecommerce_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_ecommerce_quantity']] = format_numeric_value(quantities['이커머스'])
                item_values[column_name_to_index['inventory_inspection_ecommerce_amount']] = format_amount_value(quantities['이커머스'] * current_unit_price)
            
            # 이천점반품
            if column_name_to_index.get('inventory_inspection_icheon_return_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_icheon_return_quantity']] = format_numeric_value(quantities['이천점반품'])
                item_values[column_name_to_index['inventory_inspection_icheon_return_amount']] = format_amount_value(quantities['이천점반품'] * current_unit_price)

            # === 이천점 합계 계산 ===
            icheon_sum_qty = quantities['이천점'] + quantities['케이터링'] + quantities['이커머스'] + quantities['이천점반품']
            icheon_sum_amount = icheon_sum_qty * current_unit_price
            if column_name_to_index.get('inventory_inspection_icheon_sum_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_icheon_sum_quantity']] = format_numeric_value(icheon_sum_qty)
                item_values[column_name_to_index['inventory_inspection_icheon_sum_amount']] = format_amount_value(icheon_sum_amount)

            # 하남점
            if column_name_to_index.get('inventory_inspection_hanam') is not None:
                item_values[column_name_to_index['inventory_inspection_hanam']] = format_numeric_value(quantities['하남점'])
                item_values[column_name_to_index['inventory_inspection_hanam_amount']] = format_amount_value(quantities['하남점'] * current_unit_price)
            
            # 선매입창고
            if column_name_to_index.get('inventory_inspection_prepurchase_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_prepurchase_quantity']] = format_numeric_value(quantities['선매입창고'])
                item_values[column_name_to_index['inventory_inspection_prepurchase_amount']] = format_amount_value(quantities['선매입창고'] * current_unit_price)

            # === 전체 합계 계산 ===
            total_qty = chasan_sum_qty + cheongnyangni_sum_qty + icheon_sum_qty + quantities['하남점'] + quantities['선매입창고']
            total_amount = total_qty * current_unit_price
            
            if column_name_to_index.get('inventory_inspection_total_quantity') is not None:
                item_values[column_name_to_index['inventory_inspection_total_quantity']] = format_numeric_value(total_qty)
                item_values[column_name_to_index['inventory_inspection_total_amount']] = format_amount_value(total_amount)

            # === 차이수량/금액 계산 (현재고 - 실사수량) ===
            current_quantity_idx = column_name_to_index.get('current_quantity')
            current_amount_idx = column_name_to_index.get('current_amount')
            
            if current_quantity_idx is not None and column_name_to_index.get('inventory_inspection_difference_quantity') is not None:
                current_quantity = safe_float_from_string(item_values[current_quantity_idx]) or 0.0
                difference_quantity = current_quantity - total_qty
                item_values[column_name_to_index['inventory_inspection_difference_quantity']] = format_numeric_value(difference_quantity)
            
            if current_amount_idx is not None and column_name_to_index.get('inventory_inspection_difference_amount') is not None:
                current_amount = safe_float_from_string(item_values[current_amount_idx]) or 0.0
                difference_amount = current_amount - total_amount
                item_values[column_name_to_index['inventory_inspection_difference_amount']] = format_amount_value(difference_amount)

            # 최종 업데이트
            treeview.item(item_id, values=item_values)

        # 9) 모든 아이템 업데이트 후 합계행 갱신
        update_total_row(treeview)

    except Exception as e:
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



def load_inventory_evaluation_for_treeview0(treeview):
    """
    재고평가를 불러와 '합계수량(inventory_inspection_total_quantity)' 컬럼에 반영하고 '차이수량(inventory_inspection_difference_quantity)'을 재계산하는 함수.
    """
    # 재고실사 관련 컬럼 목록
    inspection_columns = [
        'inventory_inspection_chasan',
        'inventory_inspection_icheon',
        'inventory_inspection_hanam',
        'inventory_inspection_cheongnyangni',
        'inventory_inspection_total_quantity',
        'inventory_inspection_difference_quantity'
    ]

    columns_list = [col[0] for col in columns]
    column_name_to_index = {col[0]: idx for idx, col in enumerate(columns)}

    # 재고실사 컬럼을 모두 초기화 (원하면 다른 처리 가능)
    for item_id in treeview.get_children():
        item_values = list(treeview.item(item_id)['values'])
        if not item_values:
            continue
        for col_name in inspection_columns:
            idx = column_name_to_index.get(col_name)
            if idx is not None:
                item_values[idx] = ''  
        treeview.item(item_id, values=item_values)

    reference_month = simpledialog.askstring("기준월 선택", "재고평가를 불러올 기준월을 입력하세요 (예: 2025/01):")
    if not reference_month:
        messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 기준월 존재 여부 확인
        cursor.execute(
            "SELECT COUNT(*) FROM mds_inventory_evaluation WHERE reference_month = %s",
            (reference_month,)
        )
        count = cursor.fetchone()[0]
        if count == 0:
            messagebox.showinfo("알림", f"{reference_month}에 해당하는 재고평가 데이터가 없습니다.")
            return

        # 재고평가 데이터 가져오기 (item_code와 inventory_quantity)
        query = """
        SELECT item_code, inventory_quantity
        FROM mds_inventory_evaluation
        WHERE reference_month = %s
        """
        cursor.execute(query, (reference_month,))
        eval_data = cursor.fetchall()

        # 딕셔너리: item_code -> inventory_quantity
        eval_dict = {row[0]: row[1] for row in eval_data}

        # 컬럼 인덱스
        inv_total_idx = column_name_to_index.get('inventory_inspection_total_quantity')
        difference_idx = column_name_to_index.get('inventory_inspection_difference_quantity')
        current_idx = column_name_to_index.get('current_quantity')
        item_code_idx = column_name_to_index.get('item_code')

        for item_id in treeview.get_children():
            item_values = list(treeview.item(item_id)['values'])
            if not item_values:
                continue
            if item_code_idx is None:
                continue

            this_item_code = item_values[item_code_idx]
            if this_item_code in eval_dict:
                inv_qty = eval_dict[this_item_code]
                # 합계수량 컬럼에 inventory_quantity 설정
                if inv_total_idx is not None:
                    item_values[inv_total_idx] = format_numeric_value(inv_qty)

                # 차이수량 재계산 (current_quantity - 합계수량)
                if current_idx is not None and difference_idx is not None:
                    current_quantity = safe_float_from_string(item_values[current_idx]) or 0.0
                    total_quantity = safe_float_from_string(item_values[inv_total_idx]) or 0.0
                    diff = current_quantity - total_quantity
                    item_values[difference_idx] = format_numeric_value(diff)

                treeview.item(item_id, values=item_values)

        # 합계행 업데이트
        update_total_row(treeview)

    except Exception as e:
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def create_treeview1(root):
    """
    창고별 재고 수불 현황을 조회하는 트리뷰 생성.
    """
    global loaded_month_label_treeview1  # 로딩된 기준월을 표시할 레이블 변수 선언
    global total_label_treeview1  # 총계 라벨 변수 선언

    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품목명'),
        ('specification', '규격'),
        ('chasanjum', '차산점'),
        ('chasanjum_a', '차산점A'),
        # 추가된 2개 컬럼
        ('import_warehouse_chasan', '수입창고(차산)'),
        ('import_warehouse_storage', '수입창고(보관)'),
        ('cheongnyangni', '청량리점'),
        ('icheon', '이천점'),
        ('catering', '케이터링'),
        ('hanam', '하남점'),
        ('ecommerce', '이커머스'),
        ('prepurchase', '선매입창고'),
        ('total_quantity', '합계수량'),
        ('chasanjum_return', '차산점반품'),
        ('cheongnyangni_return', '청량리반품'),
        ('icheon_return', '이천점반품'),
        ('hanam_return', '하남점반품'),
        ('total_return_quantity', '반품합계수량'),
        ('chasanjum_discard', '차산점폐기'),
        ('icheon_discard', '이천점폐기'),
        ('total_discard_quantity', '폐기합계수량'),
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    # 버튼 프레임
    button_frame = ttk.Frame(treeview_frame)
    button_frame.pack(side='top', fill='x')

    # 로딩된 기준월 레이블
    loaded_month_label_treeview1 = ttk.Label(button_frame, text="")
    loaded_month_label_treeview1.pack(side='left', padx=5, pady=5)

    # 트리뷰 생성
    treeview_subframe = ttk.Frame(treeview_frame)
    treeview_subframe.pack(fill='both', expand=True)

    treeview = ttk.Treeview(treeview_subframe, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 수직 스크롤바
    y_scroll = ttk.Scrollbar(treeview_subframe, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 컬럼 설정 + 정렬 기능 연결
    for col, name in columns:
        treeview.heading(col, text=name, command=lambda _col=col: sort_treeview_column(
            treeview, _col, numeric_columns_treeview1, reverse=False
        ))
        if col == 'no':
            treeview.column(col, width=50, anchor='center')
        elif col == 'item_code':
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')
        elif col == 'specification':
            treeview.column(col, width=100, anchor='center')
        else:
            treeview.column(col, width=80, anchor='center')

    # 태그 스타일
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')
    treeview.tag_configure('negative_value', foreground='red')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 더블 클릭 이벤트
    treeview.bind("<Double-1>", on_treeview_double_click)

    # 총계 라벨
    total_label_treeview1 = ttk.Label(treeview_frame, text="")
    total_label_treeview1.pack(side='bottom', fill='x')

    return treeview



def create_treeview2(root):
    """
    계정대체출고현황을 조회하는 트리뷰 생성.
    """
    # 기존과 동일한 columns 정의
    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품명'),
        ('specification', '규격'),
        ('quantity', '양품출고량'),
        ('substitution_type', '대체유형'),
        ('warehouse', '창고'),
        ('output_number', '출고번호'),
        ('request_number', '의뢰번호'),
        ('department', '담당부서'),
        ('manager', '담당자'),
        ('customer_code', '거래처코드'),
        ('customer_name', '거래처명'),
        ('unit_price', '단가'),
        ('amount', '금액'),
        ('foreign_currency_amount', '외화금액'),
        ('weight_unit', '중량단위'),
        ('account_type', '계정구분'),
        ('requesting_department', '요청부서명'),
        ('header_note', '헤더비고'),
        ('line_note', '라인비고'),
        ('unit_weight', '단위중량'),
    ]

    # === (1) 숫자 컬럼 리스트 정의 (트리뷰2 전용) ===
    numeric_columns_treeview2 = [
        'quantity',         # 양품출고량
        'unit_price',       # 단가
        'amount',           # 금액
        'foreign_currency_amount',  # 외화금액
        'unit_weight'       # 단위중량
        # 필요시 추가
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(treeview_frame, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 수직 스크롤바 추가
    y_scroll = ttk.Scrollbar(treeview_frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 컬럼 설정 및 정렬 기능 연결
    for col, name in columns:
        # === (2) 수정: sort_treeview_column 호출부 ===
        # 기존: command=lambda _col=col: sort_treeview_column(treeview, _col, False)
        # 변경: numeric_columns_treeview2 를 세 번째 인자로, reverse=None 을 네 번째 인자로 전달
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(
                treeview,
                _col,
                numeric_columns_treeview2,  # 세 번째: 숫자 컬럼 리스트
                None                        # 네 번째: 정렬 방향 (None → 처음 클릭 시 오름/다음 클릭 내림)
            )
        )

        # 각 컬럼 너비 및 정렬 설정
        if col == 'no':
            treeview.column(col, width=50, anchor='center')  # 'NO' 열은 50px 고정
        elif col in ['item_code', 'output_number', 'request_number', 'customer_code']:
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')  # 품목명은 왼쪽 정렬
        else:
            treeview.column(col, width=120, anchor='center')  # 기타 컬럼

    # 태그 스타일 설정
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 트리뷰2에서 더블 클릭 이벤트 바인딩
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview


def create_treeview3(root):
    """
    계정대체입고현황을 조회하는 트리뷰 생성.
    """
    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품명'),
        ('specification', '규격'),
        ('quantity', '입고량'),
        ('unit_price', '단가'),
        ('amount', '금액'),
        ('substitution_type', '대체유형명'),
        ('warehouse', '창고'),
        ('input_number', '입고번호'),
        ('request_number', '의뢰번호'),
        ('department', '담당부서'),
        ('manager', '담당자'),
        ('header_note', '비고'),
        ('line_note', '비고(라인)'),
    ]

    ### [수정] 이 트리뷰에서 숫자로 정렬해야 할 컬럼들의 리스트를 정의
    numeric_columns_treeview3 = [
        'quantity',    # 입고량
        'unit_price',  # 단가
        'amount',      # 금액
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(treeview_frame, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 수직 스크롤바 추가
    y_scroll = ttk.Scrollbar(treeview_frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 컬럼 설정 및 정렬 기능 연결
    for col, name in columns:
        ### [수정] 세 번째 인자: numeric_columns_treeview3, 네 번째 인자: None
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(treeview, _col, numeric_columns_treeview3, None)
        )
        
        # 각 컬럼 너비 및 정렬 설정
        if col == 'no':
            treeview.column(col, width=50, anchor='center')  # 'NO' 열은 50px 고정
        elif col in ['item_code', 'input_number', 'request_number']:
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')  # 품목명은 왼쪽 정렬
        else:
            treeview.column(col, width=120, anchor='center')  # 기타 컬럼

    # 태그 스타일 설정
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 트리뷰3에서 더블 클릭 이벤트 바인딩
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview



def create_treeview4(root):
    """
    출하현황을 조회하는 트리뷰 생성.
    """
    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품목명'),
        ('specification', '규격'),
        ('shipment_quantity', '출하수량'),
        ('unit_price', '단가'),
        ('amount', '금액'),
        ('won_amount_shipment', '원화금액(출하)'),
        ('vat_shipment', '부가세(출하)'),
        ('won_amount_sales', '원화금액(매출)'),
        ('vat_sales', '부가세(매출)'),
        ('total_amount_shipment', '총금액(출하)'),
        ('total_amount_sales', '총금액(매출)'),
        ('weight', '중량'),
        ('reference_month', '기준월')
    ]

    ### [수정] 숫자 컬럼 리스트 정의
    numeric_columns_treeview4 = [
        'shipment_quantity',   # 출하수량
        'unit_price',          # 단가
        'amount',              # 금액
        'won_amount_shipment', # 원화금액(출하)
        'vat_shipment',        # 부가세(출하)
        'won_amount_sales',    # 원화금액(매출)
        'vat_sales',           # 부가세(매출)
        'total_amount_shipment', 
        'total_amount_sales',
        'weight'               # 중량
        # 필요시 추가
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(treeview_frame, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 수직 스크롤바 추가
    y_scroll = ttk.Scrollbar(treeview_frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 컬럼 설정 및 정렬 기능 연결
    for col, name in columns:
        ### [수정] 세 번째 인자로 numeric_columns_treeview4, 네 번째 인자로 None
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(treeview, _col, numeric_columns_treeview4, None)
        )

        if col == 'no':
            treeview.column(col, width=50, anchor='center')  
        elif col == 'item_code':
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')  
        else:
            treeview.column(col, width=120, anchor='center')

    # 태그 스타일 설정
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 더블 클릭 이벤트 바인딩
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview


def create_treeview5(root):
    """
    [구매입고현황] 트리뷰
    새 구조: (품목코드, 품목명, 규격, 부가세, 총금액,
            관리수량, 원화금액, 거래처명, 거래처코드, 기준월)
    """

    # (1) 트리뷰5 컬럼(프로그램 내부명, 화면표시명)
    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품목명'),
        ('specification', '규격'),
        ('vat', '부가세'),
        ('total_amount', '총금액'),
        ('management_quantity', '관리수량'),  # (입고수량 → 관리수량)
        ('won_amount', '원화금액'),          # (입고금액 → 원화금액)
        ('supplier_name', '거래처명'),       # (주거래처 → 거래처명)
        ('supplier_code', '거래처코드'),    # 신규
        ('reference_month', '기준월')
    ]

    # (2) 숫자 컬럼
    numeric_columns_treeview5 = [
        'vat', 'total_amount', 'management_quantity', 'won_amount'
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    # (3) TreeView 생성
    treeview = ttk.Treeview(treeview_frame, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 스크롤바
    y_scroll = ttk.Scrollbar(treeview_frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # (4) 각 컬럼 설정
    for col, name in columns:
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(treeview, _col, numeric_columns_treeview5, None)
        )
        if col == 'no':
            treeview.column(col, width=50, anchor='center')
        elif col == 'item_code':
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')
        else:
            treeview.column(col, width=120, anchor='center')

    # 태그 스타일
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 더블클릭 이벤트
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview



def create_treeview6(root):
    """
    재고평가를 조회하는 트리뷰 생성.
    """
    columns = [
        ('no', 'NO'),
        ('item_code', '품목'),
        ('item_name', '품목명'),
        ('specification', '규격'),
        ('beginning_quantity', '기초수량'),
        ('beginning_unit_price', '기초단가'),
        ('beginning_amount', '기초금액'),
        ('receipt_quantity', '입고수량'),
        ('receipt_amount', '입고금액'),
        ('substitution_quantity', '대체수량'),
        ('substitution_amount', '대체금액'),
        ('shipment_quantity', '출고수량'),
        ('shipment_amount', '출고금액'),
        ('inventory_quantity', '재고수량'),
        ('inventory_unit_price', '재고단가'),
        ('inventory_amount', '재고금액'),
        ('reference_month', '기준월'),
    ]

    ### [수정] 트리뷰6에서 숫자로 처리해야 할 컬럼 목록
    numeric_columns_treeview6 = [
        'beginning_quantity', 'beginning_unit_price', 'beginning_amount',
        'receipt_quantity', 'receipt_amount',
        'substitution_quantity', 'substitution_amount',
        'shipment_quantity', 'shipment_amount',
        'inventory_quantity', 'inventory_unit_price', 'inventory_amount',
    ]

    treeview_frame = ttk.Frame(root)
    treeview_frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(treeview_frame, columns=[col[0] for col in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 수직 스크롤바 추가
    y_scroll = ttk.Scrollbar(treeview_frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 컬럼 설정 및 정렬 기능 연결
    for col, name in columns:
        ### [수정] 세 번째 인자에 numeric_columns_treeview6, 네 번째 인자에 None
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(treeview, _col, numeric_columns_treeview6, None)
        )

        if col == 'no':
            treeview.column(col, width=50, anchor='center')  
        elif col == 'item_code':
            treeview.column(col, width=100, anchor='center')
        elif col == 'item_name':
            treeview.column(col, width=250, anchor='w')  
        else:
            treeview.column(col, width=120, anchor='center')

    # 태그 스타일 설정
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))

    # 더블 클릭 이벤트 바인딩
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview


def create_treeview7(parent):
    columns = [
        ('no', 'NO'),
        ('item_code', '품목코드'),
        ('item_name', '품명'),
        ('specification', '규격'),
        ('unit', '단위'),
        ('category', '분류'),
        ('beginning_unit_price', '단가'),
        ('beginning_quantity', '수량'),
        ('beginning_amount', '금액')
    ]

    ### [수정] 트리뷰7에서 숫자로 처리해야 할 컬럼 목록
    numeric_columns_treeview7 = [
        'beginning_unit_price', 'beginning_quantity', 'beginning_amount'
    ]

    # 트리뷰와 스크롤바를 담을 프레임 생성
    frame = tk.Frame(parent)
    frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(frame, columns=[col[0] for col in columns], show='headings')

    # 컬럼 설정
    for col, name in columns:
        ### [수정] 세 번째 인자에 numeric_columns_treeview7, 네 번째 인자에 None
        treeview.heading(
            col,
            text=name,
            command=lambda _col=col: sort_treeview_column(treeview, _col, numeric_columns_treeview7, None)
        )
        treeview.column(col, width=100, anchor='center')

    # 스크롤바 생성
    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=treeview.yview)
    treeview.configure(yscrollcommand=scrollbar.set)

    # 트리뷰와 스크롤바 배치
    treeview.grid(row=0, column=0, sticky='nsew')
    scrollbar.grid(row=0, column=1, sticky='ns')

    # 프레임 그리드 행열 크기 설정
    frame.grid_rowconfigure(0, weight=1)
    frame.grid_columnconfigure(0, weight=1)

    # 스크롤 시 음영 재적용
    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))
    treeview.bind('<ButtonRelease-1>', lambda event: tag_alternate_rows(treeview))

    # 더블 클릭 이벤트 바인딩
    treeview.bind("<Double-1>", on_treeview_double_click)

    # 태그 설정 추가
    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    return treeview



   
def clear_treeview(treeview):
    """
    트리뷰의 모든 항목을 삭제하는 함수
    """
    treeview.delete(*treeview.get_children())



def update_total_row(treeview):
    """
    트리뷰의 데이터를 기반으로 합계를 계산하고 합계 행을 업데이트합니다.
    """
    # 기존의 합계 행이 있으면 삭제
    total_row = [k for k in treeview.get_children('') if 'totalrow' in treeview.item(k, 'tags')]
    if total_row:
        treeview.delete(total_row[0])

    # 단가 컬럼 목록 정의 (컬럼 ID를 사용)
    unit_price_columns = [col_id for col_id in treeview['columns'] if 'unit_price' in col_id]

    # 합계 계산을 위한 딕셔너리 초기화 (단가 컬럼 제외)
    total_values = {col_id: 0 for col_id in numeric_columns_treeview0 if col_id not in unit_price_columns}

    # 각 행의 값을 합산
    for item in treeview.get_children(''):
        if 'totalrow' in treeview.item(item, 'tags'):
            continue  # 합계 행은 제외
        item_values = treeview.item(item, 'values')
        for idx, col_id in enumerate(treeview['columns']):
            if col_id in total_values:
                value = item_values[idx]
                # 쉼표 제거 등 숫자 값으로 변환
                numeric_value = safe_float_from_string(value) or 0.0
                if numeric_value is not None:
                    total_values[col_id] += numeric_value

    # 합계 행 생성
    total_values_formatted = {}
    for idx, col_id in enumerate(treeview['columns']):
        if col_id in total_values:
            formatted_value = format_amount_value(total_values[col_id])
            total_values_formatted[col_id] = formatted_value
        else:
            total_values_formatted[col_id] = ''  # 합계가 아닌 컬럼은 빈 문자열

    # 합계 행을 트리뷰의 첫 번째 위치에 삽입
    total_values_list = [total_values_formatted[col_id] for col_id in treeview['columns']]
    treeview.insert('', 0, values=total_values_list, tags=('totalrow',))

    # 합계 행의 스타일 설정
    treeview.tag_configure('totalrow', background='yellow', font=('Arial', 10, 'bold'))


def create_treeview8(parent):
    """
    트리뷰8: 장려금 배분 결과를 표시.
    열 구성 (순서):
      1) no (NO)
      2) supplier_code (거래처코드)
      3) supplier_name (거래처명)
      4) item_code (품목코드)
      5) item_name (품목명)
      6) sum_won_amount (합계 : 원화금액)
      7) sum_vat_amount (합계 : 부가세)
      8) sum_total_amount (합계 : 총금액)
      9) ratio (비율)
      10) incentive (장려금)
    """

    columns = [
        ('no', 'NO'),
        ('supplier_code', '거래처코드'),
        ('supplier_name', '거래처명'),
        ('item_code', '품목코드'),
        ('item_name', '품목명'),
        ('sum_won_amount', '합계 : 원화금액'),
        ('sum_vat_amount', '합계 : 부가세'),
        ('sum_total_amount', '합계 : 총금액'),
        ('ratio', '비율'),
        ('incentive', '장려금')
    ]

    frame = ttk.Frame(parent)
    frame.pack(fill='both', expand=True)

    treeview = ttk.Treeview(frame, columns=[c[0] for c in columns], show='headings')
    treeview.pack(side='left', fill='both', expand=True)

    # 스크롤바
    y_scroll = ttk.Scrollbar(frame, orient='vertical', command=treeview.yview)
    treeview.configure(yscrollcommand=y_scroll.set)
    y_scroll.pack(side='right', fill='y')

    # 열 설정
    for col_id, col_text in columns:
        treeview.heading(col_id, text=col_text,
                         command=lambda _col=col_id: sort_treeview_column(treeview, _col, numeric_columns_treeview8))
        if col_id == 'no':
            treeview.column(col_id, width=50, anchor='center')
        elif col_id == 'item_name':
            treeview.column(col_id, width=200, anchor='w')
        else:
            treeview.column(col_id, width=120, anchor='center')

    treeview.tag_configure('evenrow', background='#f0f0f0')
    treeview.tag_configure('oddrow', background='#ffffff')

    treeview.bind('<Motion>', lambda event: tag_alternate_rows(treeview))
    treeview.bind("<Double-1>", on_treeview_double_click)

    return treeview





def load_treeview0_data(treeview):
    """
    트리뷰에 데이터를 로드하는 함수
    """
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT item_code, item_name, specification, unit, category
            FROM master
            WHERE item_code NOT IN (SELECT item_code FROM exclude_item_codes)
        """)
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())

        # 데이터 삽입
        for idx, row in enumerate(rows):
            # 데이터 포맷팅
            values = {
                'item_code': row[0],
                'item_name': row[1],
                'specification': row[2],
                'unit': row[3],
                'category': row[4],
                # 필요한 다른 컬럼들을 추가하세요
            }

            # 모든 컬럼에 대해 빈 값을 설정
            for col_id in treeview['columns']:
                if col_id not in values:
                    values[col_id] = ''

            # 트리뷰에 데이터 삽입 시 컬럼 순서에 맞게 리스트로 변환
            values_list = [values[col_id] for col_id in treeview['columns']]
            

            # 행의 음영 처리를 위한 태그 설정
            tag = 'evenrow' if idx % 2 == 0 else 'oddrow'

            # 트리뷰에 데이터 삽입
            treeview.insert('', 'end', values=values_list, tags=(tag,))

        # 태그에 따른 스타일 설정
        treeview.tag_configure('oddrow', background='lightgray')
        treeview.tag_configure('evenrow', background='white')

        # 데이터 로드 후 숫자 포맷팅 적용
        format_numeric_columns(treeview, numeric_columns_treeview0)

        # 합계 행 업데이트
        update_total_row(treeview)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def load_treeview1_data(treeview, selected_month):
    """
    mds_monthly_inventory_status 테이블에서 'reference_month'가 selected_month인 데이터만 조회하여
    트리뷰1에 로드하는 함수.
    괄호가 포함된 컬럼명을 사용하기 위해, COALESCE(...) 구문에서 컬럼명을 큰따옴표로 감쌌으며,
    조회월과 총계를 각각 다른 줄에 표시합니다.
    """
    global total_label_treeview1

    # (A) 조회월 형식 검증 (예: "2025/03")
    if not selected_month or len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    def quote_identifier(identifier: str) -> str:
        escaped = identifier.replace('"', '""')
        return f'"{escaped}"'

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 트리뷰 초기화
        clear_treeview(treeview)

        # (1) 창고 수량 컬럼 정의
        warehouse_columns = [
            ('차산점_수량', 'chasanjum'),
            ('차산점a_수량', 'chasanjum_a'),
            ('수입창고(차산)_수량', 'import_warehouse_chasan'),
            ('수입창고(보관)_수량', 'import_warehouse_storage'),
            ('청량리점_수량', 'cheongnyangni'),
            ('이천점_수량', 'icheon'),
            ('케이터링_수량', 'catering'),
            ('하남점_수량', 'hanam'),
            ('이커머스_수량', 'ecommerce'),
            ('선매입창고_수량', 'prepurchase'),
        ]

        return_columns = [
            ('차산점반품_수량', 'chasanjum_return'),
            ('청량리반품_수량', 'cheongnyangni_return'),
            ('이천점반품_수량', 'icheon_return'),
            ('하남점반품_수량', 'hanam_return'),
        ]

        discard_columns = [
            ('차산점폐기_수량', 'chasanjum_discard'),
            ('이천점폐기_수량', 'icheon_discard'),
        ]

        # (2) 괄호 포함 컬럼: COALESCE(...)
        def make_sum_expression(col_tuples):
            parts = []
            for col in col_tuples:
                quoted_col = quote_identifier(col[0])
                parts.append(f"COALESCE({quoted_col}, 0)")
            return " + ".join(parts)

        warehouse_sum = make_sum_expression(warehouse_columns)
        return_sum = make_sum_expression(return_columns)
        discard_sum = make_sum_expression(discard_columns)

        # (3) SELECT용 컬럼 (큰따옴표로 감싸기)
        select_warehouse_cols = ", ".join([quote_identifier(col[0]) for col in warehouse_columns])
        select_return_cols = ", ".join([quote_identifier(col[0]) for col in return_columns])
        select_discard_cols = ", ".join([quote_identifier(col[0]) for col in discard_columns])

        query = f"""
        SELECT
            item_code,
            item_name,
            specification,
            {select_warehouse_cols},
            ({warehouse_sum}) AS total_quantity,
            {select_return_cols},
            ({return_sum}) AS total_return_quantity,
            {select_discard_cols},
            ({discard_sum}) AS total_discard_quantity,
            reference_month
        FROM mds_monthly_inventory_status
        WHERE reference_month = %s
          AND item_code NOT IN (
              SELECT item_code FROM exclude_item_codes
          )
          AND ({warehouse_sum} + {return_sum} + {discard_sum}) <> 0
        """

        cursor.execute(query, (selected_month,))
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("정보", f"{selected_month} 조회 결과가 없습니다.")
            total_label_treeview1.config(text="")
            return

        # (4) 숫자 열 헤더
        numeric_columns = [
            '차산점', '차산점A', '수입창고(차산)', '수입창고(보관)', '청량리점', '이천점',
            '케이터링', '하남점', '이커머스', '선매입창고', '합계수량',
            '차산점반품', '청량리반품', '이천점반품', '하남점반품', '반품합계수량',
            '차산점폐기', '이천점폐기', '폐기합계수량'
        ]

        # (5) 그룹 매핑
        group_mappings = {
            '차산점_수량': '차산점',
            '차산점a_수량': '차산점',
            '수입창고(차산)_수량': '차산점',
            '수입창고(보관)_수량': '차산점',
            '차산점반품_수량': '차산점',
            '차산점폐기_수량': '차산점',

            '청량리점_수량': '청량리점',
            '청량리반품_수량': '청량리점',

            '이천점_수량': '이천점',
            '케이터링_수량': '이천점',
            '이커머스_수량': '이천점',
            '이천점반품_수량': '이천점',
            '이천점폐기_수량': '이천점',

            '하남점_수량': '기타창고',
            '하남점반품_수량': '기타창고',
            '선매입창고_수량': '기타창고',
        }

        group_warehouse_names = {
            '차산점': ['차산점', '차산점A', '수입창고(차산)', '수입창고(보관)', '차산점반품', '차산점폐기'],
            '청량리점': ['청량리점', '청량리반품'],
            '이천점': ['이천점', '케이터링', '이커머스', '이천점반품', '이천점폐기'],
            '기타창고': ['하남점', '하남점반품', '선매입창고']
        }

        # (6) all_warehouse_columns
        all_warehouse_columns = [col[0] for col in warehouse_columns + return_columns + discard_columns]

        # (7) 총계 초기화
        totals = {'차산점': 0, '청량리점': 0, '이천점': 0, '기타창고': 0, '그외창고': 0}

        # (8) 컬럼 인덱스 매핑 (reference_month 포함)
        column_names = (
            ['item_code', 'item_name', 'specification']
            + [wc[0] for wc in warehouse_columns]
            + ['total_quantity']
            + [rc[0] for rc in return_columns]
            + ['total_return_quantity']
            + [dc[0] for dc in discard_columns]
            + ['total_discard_quantity', 'reference_month']
        )
        col_index_to_name = {idx: name for idx, name in enumerate(column_names)}

        # (9) 트리뷰에 데이터 삽입
        for idx, row in enumerate(rows):
            values = [idx + 1]  # NO
            for col_index, val in enumerate(row):
                if col_index < 3:
                    # item_code, item_name, specification
                    values.append(val)
                else:
                    if col_index == len(row) - 1:
                        # 마지막은 reference_month
                        values.append(val)
                    else:
                        # 숫자 포맷
                        if val is not None and val != '':
                            try:
                                numeric_value = float(val)
                                if numeric_value.is_integer():
                                    formatted_value = f"{int(numeric_value):,}"
                                else:
                                    formatted_value = f"{numeric_value:,.5f}"
                            except ValueError:
                                formatted_value = val
                        else:
                            formatted_value = '0'
                        values.append(formatted_value)

                        # 총계 계산
                        column_name = col_index_to_name[col_index]
                        if column_name in all_warehouse_columns:
                            group_name = group_mappings.get(column_name, '그외창고')
                            numeric_val = safe_float_from_string(val) or 0.0
                            totals[group_name] += numeric_val

            treeview.insert('', 'end', values=values)

        # (10) 음영 처리 및 상태 초기화
        tag_alternate_rows(treeview)
        reset_treeview_state(treeview)

        # (11) 총계 라벨: 줄바꿈으로 조회월과 총계 구분
        total_texts = []
        for group_name in ['차산점', '청량리점', '이천점', '기타창고']:
            total_value = totals[group_name]
            if total_value == 0:
                total_display = '없음'
            else:
                total_display = format_numeric_value(total_value)
            if group_name != '그외창고':
                warehouse_list = '+'.join(group_warehouse_names[group_name])
            else:
                warehouse_list = '누락'
            total_texts.append(f"{group_name} ({warehouse_list}): {total_display}")

        # 2줄로 표시: 1) 조회월, 2) 총계
        total_label_treeview1.config(
            text=f"총계 - {' | '.join(total_texts)}"
        )

    except Exception as e:
        logging.exception("load_treeview1_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()



# def load_ending_inventory(treeview):
#     """
#     기말재고 데이터를 로드하여 treeview1에 표시하는 함수.
#     """
#     global loaded_month_label_treeview1  # 로딩된 기준월을 표시할 레이블 변수
#     global total_label_treeview1  # 총계 라벨 변수

#     # 기준월 입력 받기
#     reference_month = simpledialog.askstring("기준월 선택", "기말재고를 불러올 기준월을 입력하세요 (예: 2025/01):")
#     if not reference_month:
#         messagebox.showerror("오류", "기준월이 입력되지 않았습니다.")
#         return

#     conn = get_postgres_connection()
#     if not conn:
#         messagebox.showerror("Error", "데이터베이스 연결 실패")
#         return

#     try:
#         cursor = conn.cursor()

#         # 기준월에 해당하는 데이터가 있는지 확인
#         cursor.execute("SELECT COUNT(*) FROM mds_monthly_inventory_status WHERE reference_month = %s", (reference_month,))
#         count = cursor.fetchone()[0]
#         if count == 0:
#             messagebox.showinfo("알림", f"{reference_month}에 해당하는 기말재고 데이터가 없습니다.")
#             return

#         # 트리뷰 초기화
#         clear_treeview(treeview)

#         # 데이터 로드
#         query = """
#         SELECT
#             item_code, item_name, specification,
#             COALESCE(차산점_수량, 0) AS 차산점_수량,
#             COALESCE(차산점a_수량, 0) AS 차산점a_수량,
#             COALESCE(수입창고_수량, 0) AS 수입창고_수량,
#             COALESCE(청량리점_수량, 0) AS 청량리점_수량,
#             COALESCE(이천점_수량, 0) AS 이천점_수량,
#             COALESCE(케이터링_수량, 0) AS 케이터링_수량,
#             COALESCE(하남점_수량, 0) AS 하남점_수량,
#             COALESCE(이커머스_수량, 0) AS 이커머스_수량,
#             COALESCE(선매입창고_수량, 0) AS 선매입창고_수량,
#             COALESCE(차산점반품_수량, 0) AS 차산점반품_수량,
#             COALESCE(청량리반품_수량, 0) AS 청량리반품_수량,
#             COALESCE(이천점반품_수량, 0) AS 이천점반품_수량,
#             COALESCE(하남점반품_수량, 0) AS 하남점반품_수량,
#             COALESCE(차산점폐기_수량, 0) AS 차산점폐기_수량,
#             COALESCE(이천점폐기_수량, 0) AS 이천점폐기_수량
#         FROM mds_monthly_inventory_status
#         WHERE reference_month = %s
#         """
#         cursor.execute(query, (reference_month,))
#         rows = cursor.fetchall()

#         # 합계 수량 계산을 위한 컬럼들
#         warehouse_columns = [
#             ('차산점_수량', 'chasanjum'),
#             ('차산점a_수량', 'chasanjum_a'),
#             ('수입창고_수량', 'import_warehouse'),
#             ('청량리점_수량', 'cheongnyangni'),
#             ('이천점_수량', 'icheon'),
#             ('케이터링_수량', 'catering'),
#             ('하남점_수량', 'hanam'),
#             ('이커머스_수량', 'ecommerce'),
#             ('선매입창고_수량', 'prepurchase'),
#         ]

#         return_columns = [
#             ('차산점반품_수량', 'chasanjum_return'),
#             ('청량리반품_수량', 'cheongnyangni_return'),
#             ('이천점반품_수량', 'icheon_return'),
#             ('하남점반품_수량', 'hanam_return'),
#         ]

#         discard_columns = [
#             ('차산점폐기_수량', 'chasanjum_discard'),
#             ('이천점폐기_수량', 'icheon_discard'),
#         ]

#         # 그룹 매핑 정의
#         group_mappings = {
#             '차산점_수량': '차산점',
#             '차산점a_수량': '차산점',
#             '수입창고_수량': '차산점',
#             '선매입창고_수량': '차산점',
#             '차산점반품_수량': '차산점',
#             '차산점폐기_수량': '차산점',
#             '청량리점_수량': '청량리점',
#             '청량리반품_수량': '청량리점',
#             '이천점_수량': '이천점',
#             '케이터링_수량': '이천점',
#             '이커머스_수량': '이천점',
#             '이천점반품_수량': '이천점',
#             '이천점폐기_수량': '이천점',
#             '하남점_수량': '하남점',
#             '하남점반품_수량': '하남점',
#         }

#         # 그룹별 창고명 매핑
#         group_warehouse_names = {
#             '차산점': ['차산점', '차산점A', '수입창고', '선매입창고', '차산점반품', '차산점폐기'],
#             '청량리점': ['청량리점', '청량리반품'],
#             '이천점': ['이천점', '케이터링', '이커머스', '이천점반품', '이천점폐기'],
#             '하남점': ['하남점', '하남점반품']
#         }

#         # 총계 초기화
#         totals = {'차산점': 0, '청량리점': 0, '이천점': 0, '하남점': 0, '그외창고': 0}

#         # 모든 창고 컬럼 리스트
#         all_warehouse_columns = [col[0] for col in warehouse_columns + return_columns + discard_columns]

#         # 컬럼 인덱스와 컬럼명 매핑
#         column_names = ['item_code', 'item_name', 'specification'] + all_warehouse_columns
#         col_index_to_name = {idx: name for idx, name in enumerate(column_names)}

#         # 트리뷰의 컬럼 순서에 맞게 데이터 삽입
#         for idx, row in enumerate(rows):
#             values = [idx + 1]  # 'NO' 컬럼 값

#             # 각 열의 값을 리스트에 추가
#             for col_index, value in enumerate(row):
#                 if col_index < 3:
#                     # 품목코드, 품목명, 규격은 그대로 추가
#                     values.append(value)
#                 else:
#                     # 숫자 열은 천 단위 구분기호로 포맷팅
#                     if value is not None and value != '':
#                         try:
#                             numeric_value = float(value)
#                             if numeric_value.is_integer():
#                                 formatted_value = f"{int(numeric_value):,}"
#                             else:
#                                 formatted_value = f"{numeric_value:,.5f}"
#                         except ValueError:
#                             formatted_value = value
#                     else:
#                         formatted_value = '0'
#                     values.append(formatted_value)

#                     # 총계 계산
#                     column_name = col_index_to_name[col_index]
#                     group_name = group_mappings.get(column_name, '그외창고')
#                     numeric_value = safe_float_from_string(value) or 0.0
#                     totals[group_name] += numeric_value

#             # 합계수량 계산
#             total_quantity = sum([safe_float_from_string(row[column_names.index(col[0])]) for col in warehouse_columns]) or 0.0
#             total_return_quantity = sum([safe_float_from_string(row[column_names.index(col[0])]) for col in return_columns]) or 0.0
#             total_discard_quantity = sum([safe_float_from_string(row[column_names.index(col[0])]) for col in discard_columns]) or 0.0

#             # 합계수량 포맷팅하여 추가
#             values.insert(len(values) - len(return_columns + discard_columns),
#                           format_numeric_value(total_quantity))
#             values.insert(len(values) - len(return_columns + discard_columns) + len(return_columns),
#                           format_numeric_value(total_return_quantity))
#             values.append(format_numeric_value(total_discard_quantity))

#             treeview.insert('', 'end', values=values)

#         # 음영 처리 적용
#         tag_alternate_rows(treeview)

#         # 트리뷰 상태 초기화
#         reset_treeview_state(treeview)

#         # 총계 라벨 업데이트
#         total_texts = []
#         for group_name in ['차산점', '청량리점', '이천점', '하남점', '그외창고']:
#             total_value = totals[group_name]
#             if total_value == 0:
#                 total_display = '없음'
#             else:
#                 total_display = format_numeric_value(total_value)
#             # 그룹에 포함된 창고명 표시
#             if group_name != '그외창고':
#                 warehouse_list = '+'.join(group_warehouse_names[group_name])
#             else:
#                 warehouse_list = '기타'
#             total_texts.append(f"{group_name} ({warehouse_list}): {total_display}")
#         total_label_treeview1.config(text=f"총계 - {' | '.join(total_texts)}")

#         # 로딩된 기준월을 표시
#         loaded_month_label_treeview1.config(text=f"{reference_month} 기말재고 데이터가 로딩되었습니다.")

#     except Exception as e:
#         messagebox.showerror("Error", f"오류 발생: {e}")

#     finally:
#         if cursor:
#             cursor.close()
#         if conn:
#             conn.close()





def load_treeview2_data(treeview, selected_month):
    """
    mds_account_substitution_output 테이블에서 데이터를 가져와 트리뷰2에 로드하는 함수.
    """
    global total_label_treeview2  # 총계 라벨 변수 선언
    logging.info(f"load_treeview2_data 함수 시작. 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    selected_month_str = selected_month.replace('/', '')
    if len(selected_month_str) != 6:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 트리뷰 초기화
        clear_treeview(treeview)
        logging.debug("트리뷰 초기화 완료")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, quantity, substitution_type,
            warehouse, output_number, request_number, department, manager,
            customer_code, customer_name, unit_price, amount, foreign_currency_amount,
            weight_unit, account_type, requesting_department, header_note, line_note,
            unit_weight
        FROM mds_account_substitution_output
        WHERE output_number LIKE %s
        AND item_code NOT IN %s
        """
        like_pattern = f"TGO{selected_month_str}%"
        cursor.execute(query, (like_pattern, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("조회 결과가 없습니다.")
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview2.config(text="")  # 총계 라벨 초기화
            return

        numeric_columns = ['양품출고량', '단가', '금액', '외화금액', '단위중량']

        # 합계를 계산할 컬럼들
        sum_columns = {
            '양품출고량': 0.0,
            '금액': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'no' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']
                
                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview2.config(text=total_text)

    except Exception as e:
        logging.exception("load_treeview2_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")





def load_treeview3_data(treeview, selected_month):
    """
    mds_account_substitution_input 테이블에서 데이터를 가져와 트리뷰3에 로드하는 함수.
    만약 selected_month가 비어있다면 기본값으로 지난달(YYYY/MM)을 사용.
    """
    global total_label_treeview3  # 총계 라벨 변수 선언
    logging.info(f"load_treeview3_data 함수 시작. 조회월: {selected_month}")

    # 기준월이 제공되지 않은 경우 기본값(지난달)으로 설정
    if not selected_month or selected_month.strip() == "":
        from datetime import datetime, timedelta
        today = datetime.today()
        first_day_of_current_month = today.replace(day=1)
        last_day_previous_month = first_day_of_current_month - timedelta(days=1)
        selected_month = last_day_previous_month.strftime("%Y/%m")
        logging.info(f"조회월이 입력되지 않아 기본값(지난달) {selected_month}으로 설정됨.")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    selected_month_str = selected_month.replace('/', '')
    if len(selected_month_str) != 6:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 트리뷰 초기화
        clear_treeview(treeview)
        logging.debug("트리뷰 초기화 완료")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, quantity, unit_price, amount,
            substitution_type, warehouse, input_number, request_number, department,
            manager, header_note, line_note
        FROM mds_account_substitution_input
        WHERE input_number LIKE %s
          AND item_code NOT IN %s
        """
        like_pattern = f"TGI{selected_month_str}%"
        cursor.execute(query, (like_pattern, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("조회 결과가 없습니다.")
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview3.config(text="")  # 총계 라벨 초기화
            return

        numeric_columns = ['입고량', '단가', '금액']

        # 합계를 계산할 컬럼들
        sum_columns = {
            '입고량': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'no' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']
                
                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview3.config(text=total_text)

    except Exception as e:
        logging.exception("load_treeview3_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")






def load_treeview4_data(treeview, selected_month):
    """
    mds_shipment_status 테이블에서 데이터를 가져와 트리뷰4에 로드하는 함수.
    """
    global total_label_treeview4  # 총계 라벨 변수 선언
    logging.info(f"load_treeview4_data 함수 시작. 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 트리뷰 초기화
        clear_treeview(treeview)
        logging.debug("트리뷰 초기화 완료")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, shipment_quantity, unit_price, amount,
            won_amount_shipment, vat_shipment, won_amount_sales, vat_sales,
            total_amount_shipment, total_amount_sales, weight, reference_month
        FROM mds_shipment_status
        WHERE reference_month = %s
        AND item_code NOT IN %s
        ORDER BY item_code
        """
        cursor.execute(query, (selected_month, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("조회 결과가 없습니다.")
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview4.config(text="")  # 총계 라벨 초기화
            return

        numeric_columns = [
            '출하수량', '단가', '금액', '원화금액(출하)', '부가세(출하)',
            '원화금액(매출)', '부가세(매출)', '총금액(출하)', '총금액(매출)', '중량'
        ]

        # 합계를 계산할 컬럼들
        sum_columns = {
            '출하수량': 0.0,
            '금액': 0.0,
            '원화금액(출하)': 0.0,
            '부가세(출하)': 0.0,
            '원화금액(매출)': 0.0,
            '부가세(매출)': 0.0,
            '총금액(출하)': 0.0,
            '총금액(매출)': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'no' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']
                
                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview4.config(text=total_text)

    except Exception as e:
        logging.exception("load_treeview4_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")




def load_treeview5_data(treeview, selected_month):
    """
    mds_purchase_receipt_status 테이블에서 데이터를 가져와 트리뷰5에 로드.
    """
    global total_label_treeview5  # 총계 라벨

    logging.info(f"load_treeview5_data 함수 시작. 조회월: {selected_month}")
    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # 1) 트리뷰 초기화
        clear_treeview(treeview)

        # 2) 제외 품목코드 목록 (선택사항)
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 3) DB 조회 (새 테이블명: mds_purchase_receipt_status)
        query = """
            SELECT
                item_code,
                item_name,
                specification,
                vat,
                total_amount,
                management_quantity,
                won_amount,
                supplier_name,
                supplier_code,
                reference_month
            FROM mds_purchase_receipt_status
            WHERE reference_month = %s
              AND item_code NOT IN %s
            ORDER BY item_code
        """
        cursor.execute(query, (selected_month, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행 수: {len(rows)}")

        if not rows:
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview5.config(text="")  # 총계 라벨 초기화
            return

        # 4) 합계를 낼 컬럼들 (트리뷰 한글명과 맞춰서 관리 가능)
        sum_columns = {
            '부가세': 0.0,
            '총금액': 0.0,
            '관리수량': 0.0,
            '원화금액': 0.0
        }
        # numeric_db_cols: DB 컬럼명
        numeric_db_cols = ['vat','total_amount','management_quantity','won_amount']

        # 5) 행 삽입
        for idx, row in enumerate(rows):
            # row = (item_code, item_name, specification, vat, total_amount,
            #        management_quantity, won_amount, supplier_name, supplier_code, reference_month)
            values = [idx+1]  # 'NO' 컬럼

            for col_index, db_value in enumerate(row):
                tree_col_id = treeview['columns'][col_index+1]  # +1은 NO 칼럼
                heading_text = treeview.heading(tree_col_id)['text']  # 예: '품목코드','관리수량' 등

                if heading_text in sum_columns:
                    numeric_val = safe_float_from_string(db_value) or 0.0
                    sum_columns[heading_text] += numeric_val
                    values.append(format_numeric_value(db_value))
                else:
                    values.append(db_value)

            treeview.insert('', 'end', values=values)

        # 6) 음영 처리
        tag_alternate_rows(treeview)
        reset_treeview_state(treeview)

        # 7) 총계 라벨
        total_text = ', '.join([f"{k} 합계: {format_numeric_value(v)}" for k, v in sum_columns.items()])
        total_label_treeview5.config(text=total_text)

    except Exception as e:
        logging.exception("load_treeview5_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()







def load_treeview6_data(treeview, selected_month):
    """
    mds_inventory_evaluation 테이블에서 데이터를 가져와 트리뷰6에 로드하는 함수.
    """
    global total_label_treeview6  # 총계 라벨 변수 선언
    logging.info(f"load_treeview6_data 함수 시작. 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 트리뷰 초기화
        clear_treeview(treeview)
        logging.debug("트리뷰 초기화 완료")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, beginning_quantity, beginning_unit_price, beginning_amount,
            receipt_quantity, receipt_amount, substitution_quantity, substitution_amount, shipment_quantity,
            shipment_amount, inventory_quantity, inventory_unit_price, inventory_amount, reference_month
        FROM mds_inventory_evaluation
        WHERE reference_month = %s
        AND item_code NOT IN %s
        ORDER BY item_code
        """
        cursor.execute(query, (selected_month, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("조회 결과가 없습니다.")
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview6.config(text="")  # 총계 라벨 초기화
            return

        numeric_columns = [
            '기초수량', '기초단가', '기초금액',
            '입고수량', '입고금액', '대체수량', '대체금액',
            '출고수량', '출고금액', '재고수량', '재고단가', '재고금액'
        ]

        # 합계를 계산할 컬럼들
        sum_columns = {
            '기초수량': 0.0,
            '기초금액': 0.0,
            '입고수량': 0.0,
            '입고금액': 0.0,
            '대체수량': 0.0,
            '대체금액': 0.0,
            '출고수량': 0.0,
            '출고금액': 0.0,
            '재고수량': 0.0,
            '재고금액': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'no' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']
                
                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview6.config(text=total_text)

    except Exception as e:
        logging.exception("load_treeview6_data 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")

def load_treeview7_data(treeview, selected_year):
    global total_label_treeview7  # 총계 라벨 변수 선언
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT item_code, item_name, specification, unit, category,
                   beginning_unit_price, beginning_quantity, beginning_amount
            FROM mds_basic_data
            WHERE reference_year = %s
        """, (selected_year,))
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("정보", "조회 결과가 없습니다.")
            total_label_treeview7.config(text="")  # 총계 라벨 초기화
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())

        # 합계 계산을 위한 초기화 (단가 합계 제거)
        sum_columns = {
            # 'beginning_unit_price': 0.0,  # 단가 합계 계산 제거
            'beginning_quantity': 0.0,
            'beginning_amount': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1] + list(row)
            treeview.insert('', 'end', values=values)

            # 합계 계산 (단가 합계 계산 제거)
            # sum_columns['beginning_unit_price'] += safe_float_from_string(row[5])
            sum_columns['beginning_quantity'] += safe_float_from_string(row[6]) or 0.0
            sum_columns['beginning_amount'] += safe_float_from_string(row[7]) or 0.0

        # 총계 라벨 업데이트 (단가 합계 제거)
        total_text = (
            f"수량 합계: {format_numeric_value(sum_columns['beginning_quantity'])}, "
            f"금액 합계: {format_numeric_value(sum_columns['beginning_amount'])}"
        )
        total_label_treeview7.config(text=total_text)

        # 숫자 열 포맷팅 적용
        numeric_columns = ['단가', '수량', '금액']
        format_numeric_columns(treeview, numeric_columns)

        # 음영 처리 적용
        tag_alternate_rows(treeview)

    except Exception as e:
        messagebox.showerror("에러", f"데이터 로드 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def load_treeview8_data(treeview, data_list):
    """
    data_list: 
      [
        (
          no,               # 1
          supplier_code,    # 2
          supplier_name,    # 3
          item_code,        # 4
          item_name,        # 5
          sum_won_amount,   # 6
          sum_vat_amount,   # 7
          sum_total_amount, # 8
          ratio,            # 9 (퍼센트 문자열)
          incentive         # 10 (정수)
        ),
        ...
      ]
    """

    # 1) 기존 데이터 삭제
    for item in treeview.get_children():
        treeview.delete(item)

    # 2) 백업
    treeview.all_data = data_list[:]
    inserted_items = []

    # 3) 행 삽입
    for idx, row in enumerate(data_list):
        tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
        item_id = treeview.insert('', 'end', values=row, tags=(tag,))
        inserted_items.append((row, item_id))

    # 4) 원래 순서 백업
    treeview.original_order = inserted_items

    # 5) 숫자 열 포맷팅
    # ratio는 퍼센트, incentive는 정수금액, sum_won_amount/sum_vat_amount/sum_total_amount는 정수금액
    format_numeric_columns(treeview, numeric_columns_treeview8)

    # 6) 음영
    tag_alternate_rows(treeview)


def show_ledger_total(reference_month):
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("오류", "DB 연결 실패")
        return
    try:
        cursor = conn.cursor()
        query = """
            SELECT SUM(대변)
            FROM mds_account_ledger
            WHERE 회계일자 LIKE %s || '%%'
        """
        cursor.execute(query, (reference_month,))
        total_debit = cursor.fetchone()[0]
        if total_debit is None:
            total_debit = 0.0

        messagebox.showinfo("회계월 대변 총합계",
                            f"{reference_month}의 대변 총합계: {total_debit:,.2f}")
    except Exception as e:
        messagebox.showerror("오류", f"회계월 대변 총합계 계산 중 오류: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def distribute_incentive_for_treeview8(treeview8, label_current_month_8):
    """
    [장려금 배분 로직 - 거래처코드 기준, 대소문자 구분 없이 매칭, GROUP BY를 통해 item_code 중복 합산]

    1) mds_account_ledger:
       - (회계일자 LIKE %selected_month%) + 거래처코드 IS NOT NULL
       - 대변 합계
    2) mds_purchase_receipt_status:
       - reference_month = selected_month
       - lower(supplier_code) = lower(ledger.거래처코드)  (대소문자 구분 없이)
       - GROUP BY item_code, SUM(won_amount), SUM(vat), SUM(total_amount)
         + MAX(item_name), MAX(supplier_name) 등 (중복행 합산)
    3) won_amount>0 품목만 배분(비율 = won_amount / 전체합)
    4) 소수점 반올림 → diff 보정
    5) 최종 (NO, supplier_code, supplier_name, item_code, item_name, won_amount, vat, total_amount, ratio, incentive) → load_treeview8_data
    """
    from datetime import datetime, timedelta

    # 1) 기본값(이전달) 설정
    today = datetime.today()
    first = today.replace(day=1)
    last_month = first - timedelta(days=1)
    default_month = last_month.strftime("%Y/%m")  # 예: 2025/02

    selected_month = simpledialog.askstring(
        "장려금배분",
        "년/월을 입력하세요 (예: 2025/01):",
        initialvalue=default_month
    )
    if not selected_month:
        return

    # 선택한 월 대변합계 표시(기존 show_ledger_total 등 함수가 있다면 호출)
    show_ledger_total(selected_month)

    # 라벨에 표시
    label_current_month_8.config(text=f"현재 작업 기준월: {selected_month}")

    try:
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "DB 연결 실패")
            return
        cursor = conn.cursor()

        # 2) mds_account_ledger: 거래처코드 + 월 기준 대변합계
        query_ledger = """
            SELECT 
                거래처코드,
                SUM(대변) AS total_debit
            FROM mds_account_ledger
            WHERE 회계일자 LIKE %s || '%%'
              AND 거래처코드 IS NOT NULL
            GROUP BY 거래처코드
        """
        cursor.execute(query_ledger, (selected_month,))
        ledger_rows = cursor.fetchall()  # [(ledger_code, total_debit), ...]

        if not ledger_rows:
            messagebox.showinfo("정보", f"{selected_month}월 데이터가 mds_account_ledger에 없습니다.")
            cursor.close()
            conn.close()
            return

        final_data = []  # [(no, supplier_code, supplier_name, item_code, item_name, w_val, v_val, t_val, ratio, incentive)]
        no_counter = 1

        for ledger_code, total_debit in ledger_rows:
            if not ledger_code:
                continue
            total_debit_float = float(total_debit) if total_debit else 0.0

            # (A) mds_purchase_receipt_status: 
            #     supplier_code=ledger_code (대소문자 무시), reference_month=selected_month
            #     GROUP BY item_code => SUM(won_amount), SUM(vat), SUM(total_amount), 
            #                            MAX(item_name), MAX(supplier_name)
            query_purchase = """
                SELECT
                    item_code,
                    MAX(item_name) AS item_name,
                    MAX(supplier_name) AS supplier_name,
                    SUM(won_amount) AS sum_won,
                    SUM(vat) AS sum_vat,
                    SUM(total_amount) AS sum_total
                FROM mds_purchase_receipt_status
                WHERE reference_month = %s
                  AND LOWER(supplier_code) = LOWER(%s)
                GROUP BY item_code
            """
            cursor.execute(query_purchase, (selected_month, ledger_code))
            purchase_rows = cursor.fetchall()
            # [(ic, i_name, s_name, sum_won, sum_vat, sum_total), ...]

            if not purchase_rows:
                continue

            # (B) won_amount>0 대상만 배분
            sum_of_won_positive = 0.0
            item_list = []
            for ic, i_name, s_name, sum_won, sum_vat, sum_total in purchase_rows:
                w_val = float(sum_won) if sum_won else 0.0
                v_val = float(sum_vat) if sum_vat else 0.0
                t_val = float(sum_total) if sum_total else 0.0
                i_name = i_name or ""
                s_name = s_name or ""

                item_list.append((ic, i_name, s_name, w_val, v_val, t_val))
                if w_val > 0:
                    sum_of_won_positive += w_val

            if sum_of_won_positive <= 0:
                # 배분 불가
                continue

            # (C) 품목별 비율 + 장려금
            row_data_list = []
            for ic, i_name, s_name, w_val, v_val, t_val in item_list:
                if w_val > 0:
                    ratio_float = (w_val / sum_of_won_positive) * 100.0
                    incentive_float = total_debit_float * (ratio_float / 100.0)
                else:
                    ratio_float = 0.0
                    incentive_float = 0.0

                row_data_list.append({
                    "ic": ic,
                    "item_name": i_name,
                    "supplier_name": s_name,  # 모든 row가 동일 s_name일 가능성
                    "won_val": w_val,
                    "vat_val": v_val,
                    "total_val": t_val,
                    "ratio_float": ratio_float,
                    "incentive_float": incentive_float
                })

            # (D) 소수점 반올림 + diff 보정
            incentives_rounded = [int(round(r["incentive_float"])) for r in row_data_list]
            sum_incentives = sum(incentives_rounded)
            diff = sum_incentives - int(round(total_debit_float))

            adjust_item_idx = None
            if diff != 0 and len(incentives_rounded) > 0:
                max_idx = max(range(len(incentives_rounded)), key=lambda i: incentives_rounded[i])
                adjust_item_idx = max_idx
                incentives_rounded[max_idx] -= diff

            # (E) 최종 final_data 누적 (10개 항목)
            for i, row_item in enumerate(row_data_list):
                ratio_str = f"{row_item['ratio_float']:.2f}%"
                final_incentive = incentives_rounded[i]

                final_data.append((
                    no_counter,             # no
                    ledger_code,            # supplier_code
                    row_item["supplier_name"],  # supplier_name
                    row_item["ic"],         # item_code
                    row_item["item_name"],  # item_name
                    int(row_item["won_val"]),
                    int(row_item["vat_val"]),
                    int(row_item["total_val"]),
                    ratio_str,
                    final_incentive
                ))
                no_counter += 1

            # # (F) diff 안내 장려금 보전값 팝언으로 안내
            # if diff != 0 and adjust_item_idx is not None:
            #     changed_code = row_data_list[adjust_item_idx]["ic"]
            #     before_val = int(round(row_data_list[adjust_item_idx]["incentive_float"]))
            #     after_val = incentives_rounded[adjust_item_idx]
            #     msg_text = (
            #         f"[보정 안내]\n"
            #         f"거래처코드: {ledger_code}\n"
            #         f"품목: {changed_code}\n"
            #         f"조정 전 장려금: {before_val:,.0f}원\n"
            #         f"조정 후 장려금: {after_val:,.0f}원\n"
            #         f"(차이: {abs(diff):,.0f}원)"
            #     )
            #     messagebox.showinfo("보정 안내", msg_text)

        # (G) 트리뷰8 로딩
        load_treeview8_data(treeview8, final_data)

        messagebox.showinfo("완료", f"{selected_month} 장려금배분 완료.")
        cursor.close()
        conn.close()

    except Exception as e:
        messagebox.showerror("오류", f"장려금배분 중 오류: {e}")
        logging.exception("distribute_incentive_for_treeview8 예외")
        if conn:
            conn.close()




def save_incentive_result_for_treeview8(treeview8, label_current_month_8):
    """
    결과 저장하기 버튼을 눌렀을 때 실행.
    트리뷰8에 로딩된 모든 컬럼 + reference_month를 mds_incentive_result 테이블에 저장.
    """
    if not hasattr(treeview8, 'all_data') or not treeview8.all_data:
        messagebox.showinfo("알림", "저장할 데이터가 없습니다.")
        return

    # label_current_month_8 예: "현재 작업 기준월: 2025/02" -> "2025/02"
    current_text = label_current_month_8.cget("text")
    default_ref_month = current_text.replace("현재 작업 기준월: ", "").strip()

    reference_month = simpledialog.askstring("기준월 선택",
                                             "저장할 기준월을 입력하세요 (예: 2025/01):",
                                             initialvalue=default_ref_month)
    if not reference_month:
        return

    try:
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "DB 연결 실패")
            return
        cursor = conn.cursor()

        # 기존 해당 월 데이터 삭제
        cursor.execute("SELECT COUNT(*) FROM mds_incentive_result WHERE reference_month = %s", (reference_month,))
        cnt = cursor.fetchone()[0]
        if cnt > 0:
            if messagebox.askyesno("데이터 존재 확인", f"{reference_month} 데이터가 이미 존재합니다. 삭제 후 진행하시겠습니까?"):
                cursor.execute("DELETE FROM mds_incentive_result WHERE reference_month = %s", (reference_month,))
                conn.commit()
            else:
                return

        # 테이블 컬럼 순서와 정확히 매칭 (11개)
        insert_query = """
            INSERT INTO mds_incentive_result 
            (no, supplier_code, supplier_name, item_code, item_name, sum_won_amount, sum_vat_amount, sum_total_amount, ratio, incentive, reference_month)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        data_to_insert = []
        # treeview8.all_data: 각 행이 (no, supplier_code, supplier_name, item_code, item_name, sum_won_amount, sum_vat_amount, sum_total_amount, ratio, incentive)
        for row in treeview8.all_data:
            # row는 10개 튜플
            new_row = list(row)
            new_row.append(reference_month)  # 11번째 항목으로 reference_month 추가
            data_to_insert.append(new_row)

        batch_size = 1000
        total_rows = len(data_to_insert)
        for i in range(0, total_rows, batch_size):
            batch = data_to_insert[i:i+batch_size]
            cursor.executemany(insert_query, batch)
            conn.commit()

        messagebox.showinfo("완료", f"{reference_month} 결과 저장 완료. 총 {total_rows}행 저장.")
        cursor.close()
        conn.close()

    except Exception as e:
        messagebox.showerror("오류", f"결과 저장 중 오류: {e}")
        print(e)





def load_incentive_result_for_treeview8(treeview8, label_current_month_8):
    """
    저장된 결과를 불러오는 함수.
    기준월 입력받아 mds_incentive_result에서 해당월 데이터 불러와 treeview8에 로딩.
    테이블 컬럼: no, supplier_code, supplier_name, item_code, item_name,
                 sum_won_amount, sum_vat_amount, sum_total_amount, ratio, incentive, reference_month
    트리뷰8 컬럼: no, supplier_code, supplier_name, item_code, item_name,
                 sum_won_amount, sum_vat_amount, sum_total_amount, ratio, incentive
    """
    reference_month = simpledialog.askstring("기준월 선택", "불러올 기준월을 입력하세요 (예: 2025/01):")
    if not reference_month:
        return

    try:
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("오류", "DB 연결 실패")
            return
        cursor = conn.cursor()

        # mds_incentive_result 테이블에서, 트리뷰8에 필요한 10개 컬럼만 SELECT
        query = """
            SELECT 
                no,
                supplier_code,
                supplier_name,
                item_code,
                item_name,
                sum_won_amount,
                sum_vat_amount,
                sum_total_amount,
                ratio,
                incentive
            FROM mds_incentive_result
            WHERE reference_month = %s
            ORDER BY no
        """
        cursor.execute(query, (reference_month,))
        rows = cursor.fetchall()
        # rows 형식: [(no, supplier_code, supplier_name, item_code, item_name,
        #             sum_won_amount, sum_vat_amount, sum_total_amount, ratio, incentive), ...]

        if not rows:
            messagebox.showinfo("정보", f"{reference_month} 데이터가 없습니다.")
            cursor.close()
            conn.close()
            return

        # 트리뷰8에 로딩
        load_treeview8_data(treeview8, rows)

        # 현재 작업 기준월 표시
        label_current_month_8.config(text=f"현재 작업 기준월: {reference_month}")

        messagebox.showinfo("완료", f"{reference_month} 데이터 불러오기 완료.")
        cursor.close()
        conn.close()

    except Exception as e:
        messagebox.showerror("오류", f"결과 불러오기 중 오류: {e}")
        print(e)








def on_treeview_double_click(event):
    """
    트리뷰에서 더블 클릭 시 해당 셀의 데이터를 검색창에 입력하는 함수
    """
    global search_entry  # 전역 변수 선언

    tree = event.widget

    item_id = tree.identify_row(event.y)
    column_id = tree.identify_column(event.x)

    if not item_id or not column_id:
        messagebox.showinfo("정보", "선택된 항목이 없습니다.")
        return

    # 클릭된 컬럼의 인덱스를 가져옵니다.
    column_index = int(column_id.replace('#', '')) - 1  # 인덱스는 0부터 시작

    # 아이템의 값을 가져옵니다.
    item_values = tree.item(item_id, 'values')

    if 0 <= column_index < len(item_values):
        cell_value = item_values[column_index]
    else:
        messagebox.showinfo("정보", "선택된 셀의 값을 가져올 수 없습니다.")
        return

    # 검색 칸에 값 삽입
    search_entry.delete(0, tk.END)
    search_entry.insert(0, cell_value)


def search_current_treeview(event=None):
    global search_entry, current_treeview, treeview0, treeview1, treeview2, treeview3, treeview4, treeview5, treeview6, treeview7, treeview8
    global date_var_treeview2, date_var_treeview3, date_var_treeview4, date_var_treeview5, date_var_treeview6, date_var_treeview7

    keyword = search_entry.get().strip()
    logging.info(f"검색어: '{keyword}'")

    if not current_treeview:
        messagebox.showwarning("Warning", "현재 표시된 트리뷰가 없습니다.")
        return

    if current_treeview == 'treeview0':
        search_treeview0(treeview0, keyword)
    elif current_treeview == 'treeview1':
        search_treeview1(treeview1, keyword)
    elif current_treeview == 'treeview2':
        selected_month = date_var_treeview2.get()
        search_treeview2(treeview2, keyword, selected_month)
    elif current_treeview == 'treeview3':
        selected_month = date_var_treeview3.get()
        search_treeview3(treeview3, keyword, selected_month)
    elif current_treeview == 'treeview4':
        selected_month = date_var_treeview4.get()
        search_treeview4(treeview4, keyword, selected_month)
    elif current_treeview == 'treeview5':
        selected_month = date_var_treeview5.get()
        search_treeview5(treeview5, keyword, selected_month)
    elif current_treeview == 'treeview6':
        selected_month = date_var_treeview6.get()
        search_treeview6(treeview6, keyword, selected_month)
    elif current_treeview == 'treeview7':
        selected_year = date_var_treeview7.get()
        search_treeview7(treeview7, keyword, selected_year)
    elif current_treeview == 'treeview8':
        search_treeview8(treeview8, keyword)
    else:
        messagebox.showinfo("Information", "검색할 수 없는 트리뷰입니다.")



# def search_treeview0(treeview, keyword):
#     """
#     트리뷰0에서 '결과 불러오기'로 로드된 데이터 내에서 검색을 수행하는 함수
#     """
#     # # 데이터 로드 여부 확인
#     # if not hasattr(treeview, 'all_data') or not treeview.all_data:
#     #     messagebox.showinfo("알림", "먼저 '결과 불러오기' 버튼으로 데이터를 로딩해주세요.")
#     #     return

#     keyword = keyword.lower()

#     columns = treeview['columns']

#     # 숫자 컬럼 및 단가 컬럼 식별
#     numeric_columns = numeric_columns_treeview0  # 전역 변수 사용
#     unit_price_columns = [col_id for col_id in columns if 'unit_price' in col_id]

#     # 검색 결과를 저장할 리스트
#     matching_rows = []
#     for row in treeview.all_data:
#         if any(keyword in str(value).lower() for value in row):
#             matching_rows.append(row)

#     if not matching_rows:
#         messagebox.showinfo("정보", "검색 결과가 없습니다.")
#         return

#     # 트리뷰 초기화
#     treeview.delete(*treeview.get_children())

#     # 합계 계산을 위한 딕셔너리 초기화 (단가 컬럼 제외)
#     total_values = {col_id: 0 for col_id in numeric_columns if col_id not in unit_price_columns}

#     # 검색 결과를 트리뷰에 삽입
#     for idx, row in enumerate(matching_rows):
#         values = []
#         for idx_col, value in enumerate(row):
#             col = columns[idx_col]
#             if col in numeric_columns:
#                 if value is not None:
#                     # 숫자 포맷팅
#                     value_str = format_numeric_value(value)
#                 else:
#                     value_str = ''
#                 values.append(value_str)
#                 # 합계 계산 (단가 컬럼 제외)
#                 if col not in unit_price_columns:
#                     numeric_value = safe_float_from_string(value_str) or 0.0
#                     if numeric_value is not None:
#                         total_values[col] += numeric_value
#             else:
#                 values.append(value if value is not None else '')

#         # 행의 음영 처리를 위한 태그 설정
#         tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
#         treeview.insert('', 'end', values=values, tags=(tag,))

#     # 합계 행 생성
#     total_values_formatted = {}
#     for col_id in columns:
#         if col_id in total_values:
#             formatted_value = format_numeric_value(total_values[col_id])
#             total_values_formatted[col_id] = formatted_value
#         else:
#             total_values_formatted[col_id] = ''  # 합계가 아닌 컬럼은 빈 문자열

#     total_values_list = [total_values_formatted[col_id] for col_id in columns]
#     treeview.insert('', 0, values=total_values_list, tags=('totalrow',))

#     # 합계 행의 스타일 설정
#     treeview.tag_configure('totalrow', background='yellow', font=('Arial', 10, 'bold'))

#     # 태그에 따른 스타일 설정
#     treeview.tag_configure('oddrow', background='lightgray')
#     treeview.tag_configure('evenrow', background='white')

#     # 음영 재적용
#     reapply_row_tags(treeview)

#     # 정렬 상태 초기화 (필요에 따라)
#     if hasattr(treeview, '_sort_states'):
#         treeview._sort_states.clear()

#     # 정렬 함수 호출 (필요에 따라)
#     # sort_treeview_column(treeview, 'item_code', numeric_columns, reverse=False)

def search_treeview0(treeview, keyword):
    """
    트리뷰0 내에서 검색어(keyword)에 해당하는 모든 행을 찾아,
    만약 고정(=pinned) 행이 있다면 마지막 고정행 바로 아래,
    없으면 합계행(totalrow) 바로 아래에 해당 행들을 모두 이동시키고 선택합니다.
    """
    keyword = keyword.strip().lower()
    if not keyword:
        # 검색어 없이 엔터면 트리뷰 그대로 둠
        return

    # 현재 트리뷰의 모든 행 가져오기
    all_items = list(treeview.get_children())
    matching_items = []

    # 검색: totalrow(합계행)는 제외
    for item in all_items:
        tags = treeview.item(item, 'tags')
        if 'totalrow' in tags:
            continue
        values = treeview.item(item, 'values')
        if any(keyword in str(val).lower() for val in values):
            matching_items.append(item)

    if not matching_items:
        messagebox.showinfo("정보", "검색 결과가 없습니다.")
        return

    # 정렬고정 행과 합계행의 인덱스 찾기
    pinned_indexes = []
    totalrow_indexes = []
    for idx, child in enumerate(all_items):
        row_tags = treeview.item(child, 'tags')
        if 'totalrow' in row_tags:
            totalrow_indexes.append(idx)
        elif 'pinned' in row_tags:
            pinned_indexes.append(idx)
    
    # 삽입 위치 결정
    if pinned_indexes:
        insertion_index = pinned_indexes[-1] + 1
    elif totalrow_indexes:
        insertion_index = totalrow_indexes[-1] + 1
    else:
        insertion_index = 0

    # 매칭된 모든 행을 재배치
    # 역순으로 이동시켜 원래 순서 유지
    for item in reversed(matching_items):
        treeview.move(item, '', insertion_index)

    # 마지막 매칭 항목에 포커스
    if matching_items:
        treeview.selection_set(matching_items)  # 모든 매칭 항목 선택
        treeview.focus(matching_items[0])       # 첫 번째 항목에 포커스
        treeview.see(matching_items[0])         # 첫 번째 항목이 보이도록 스크롤
    
    reapply_row_tags(treeview)



def search_treeview1(treeview, keyword):
    """
    master 테이블에서 통합 검색을 수행하는 함수.
    """
    global total_label_treeview1  # 총계 라벨 변수 선언
    logging.info(f"search_treeview1 함수 시작. 검색어: {keyword}")

    # 컬럼 정의
    warehouse_columns = [
        ('차산점_수량', 'chasanjum'),
        ('차산점a_수량', 'chasanjum_a'),
        ('수입창고_수량', 'import_warehouse'),
        ('청량리점_수량', 'cheongnyangni'),
        ('이천점_수량', 'icheon'),
        ('케이터링_수량', 'catering'),
        ('하남점_수량', 'hanam'),
        ('이커머스_수량', 'ecommerce'),
        ('선매입창고_수량', 'prepurchase'),
    ]

    return_columns = [
        ('차산점반품_수량', 'chasanjum_return'),
        ('청량리반품_수량', 'cheongnyangni_return'),
        ('이천점반품_수량', 'icheon_return'),
        ('하남점반품_수량', 'hanam_return'),
    ]

    discard_columns = [
        ('차산점폐기_수량', 'chasanjum_discard'),
        ('이천점폐기_수량', 'icheon_discard'),
    ]

    # 합계 수량 계산
    warehouse_sum = " + ".join(
        [f"COALESCE({col[0]}, 0)" for col in warehouse_columns]
    )
    return_sum = " + ".join(
        [f"COALESCE({col[0]}, 0)" for col in return_columns]
    )
    discard_sum = " + ".join(
        [f"COALESCE({col[0]}, 0)" for col in discard_columns]
    )

    query = f"""
    SELECT
        item_code, item_name, specification,
        {', '.join([col[0] for col in warehouse_columns])},
        ({warehouse_sum}) AS total_quantity,
        {', '.join([col[0] for col in return_columns])},
        ({return_sum}) AS total_return_quantity,
        {', '.join([col[0] for col in discard_columns])},
        ({discard_sum}) AS total_discard_quantity
    FROM master
    WHERE item_code NOT IN (
        SELECT item_code FROM exclude_item_codes
    )
      AND (
        (LOWER(item_code) LIKE LOWER(%s)) OR
        (LOWER(item_name) LIKE LOWER(%s))
      )
    """

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute(query, (f"%{keyword}%", f"%{keyword}%"))
        rows = cursor.fetchall()

        if not rows:
            logging.info("검색 결과가 없습니다.")
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview1.config(text="")  # 총계 라벨 초기화
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())
        logging.debug("트리뷰 초기화 완료")

        # 숫자 열의 헤더 텍스트 리스트
        numeric_columns = [
            '차산점', '차산점A', '수입창고', '청량리점', '이천점',
            '케이터링', '하남점', '이커머스', '선매입창고', '합계수량',
            '차산점반품', '청량리반품', '이천점반품', '하남점반품',
            '반품합계수량', '차산점폐기', '이천점폐기', '폐기합계수량'
        ]

        # 그룹 매핑 정의
        group_mappings = {}
        for col in [
            '차산점_수량', '차산점a_수량', '수입창고_수량',
            '차산점반품_수량', '차산점폐기_수량'
        ]:
            group_mappings[col] = '차산점'
        for col in ['청량리점_수량', '청량리반품_수량']:
            group_mappings[col] = '청량리점'
        for col in [
            '이천점_수량', '케이터링_수량', '이커머스_수량',
            '이천점반품_수량', '이천점폐기_수량'
        ]:
            group_mappings[col] = '이천점'
        for col in '선매입창고_수량', ['하남점_수량', '하남점반품_수량']:
            group_mappings[col] = '기타창고'

        # 그룹별 창고명 매핑
        group_warehouse_names = {
            '차산점': [
                '차산점', '차산점A', '수입창고',
                '차산점반품', '차산점폐기'
            ],
            '청량리점': ['청량리점', '청량리반품'],
            '이천점': [
                '이천점', '케이터링', '이커머스',
                '이천점반품', '이천점폐기'
            ],
            '기타창고': ['선매입창고', '하남점', '하남점반품']
        }

        # 모든 창고 컬럼 리스트
        all_warehouse_columns = [
            col[0] for col in
            warehouse_columns + return_columns + discard_columns
        ]

        # 그룹에 포함되지 않은 컬럼 식별
        columns_in_groups = set(group_mappings.keys())
        other_columns = [
            col for col in all_warehouse_columns
            if col not in columns_in_groups
        ]

        # 총계 초기화
        totals = {
            '차산점': 0, '청량리점': 0, '이천점': 0,
            '기타창고': 0, '그외창고': 0
        }

        # 컬럼 인덱스와 컬럼명 매핑
        column_names = ['item_code', 'item_name', 'specification'] + \
            all_warehouse_columns + [
                'total_quantity', 'total_return_quantity',
                'total_discard_quantity'
            ]
        col_index_to_name = {
            idx: name for idx, name in enumerate(column_names)
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값

            for col_index, value in enumerate(row):
                if col_index < 3:
                    # 품목코드, 품목명, 규격은 그대로 추가
                    values.append(value)
                else:
                    # 숫자 열은 천 단위 구분기호로 포맷팅
                    if value is not None and value != '':
                        try:
                            numeric_value = float(value)
                            if numeric_value.is_integer():
                                formatted_value = f"{int(numeric_value):,}"
                            else:
                                formatted_value = f"{numeric_value:,.5f}"
                        except ValueError:
                            formatted_value = value
                    else:
                        formatted_value = '0'
                    values.append(formatted_value)

                    # 총계 계산
                    column_name = col_index_to_name[col_index]
                    if column_name in all_warehouse_columns:
                        # 그룹명 가져오기
                        group_name = group_mappings.get(
                            column_name, '그외창고'
                        )
                        # 숫자 값 파싱
                        numeric_value = safe_float_from_string(value) or 0.0
                        # 그룹에 합산
                        totals[group_name] += numeric_value

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_texts = []
        for group_name in [
            '차산점', '청량리점', '이천점', '기타창고', '그외창고'
        ]:
            total_value = totals[group_name]
            if total_value == 0:
                total_display = '없음'
            else:
                total_display = format_numeric_value(total_value)
            # 그룹에 포함된 창고명 표시
            if group_name != '그외창고':
                warehouse_list = '+'.join(
                    group_warehouse_names[group_name]
                )
            else:
                warehouse_list = '누락'
            total_texts.append(
                f"{group_name} ({warehouse_list}): {total_display}"
            )
        total_label_treeview1.config(
            text=f"총계 - {' | '.join(total_texts)}"
        )

        logging.debug("검색 결과 삽입 및 총계 계산 완료")

    except Exception as e:
        logging.exception("search_treeview1 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")



def search_treeview2(treeview, keyword, selected_month):
    """
    mds_account_substitution_output 테이블에서 통합 검색을 수행하는 함수.
    """
    global total_label_treeview2  # 총계 라벨 변수 선언
    logging.info(f"search_treeview2 함수 시작. 검색어: {keyword}, 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        logging.error("데이터베이스 연결 실패")
        return

    # 조회월 형식 확인 및 처리
    selected_month_str = selected_month.replace('/', '')
    if len(selected_month_str) != 6:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # LIKE 패턴 생성
        like_pattern = f"TGO{selected_month_str}%"

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, quantity, substitution_type,
            warehouse, output_number, request_number, department, manager,
            customer_code, customer_name, unit_price, amount, foreign_currency_amount,
            weight_unit, account_type, requesting_department, header_note, line_note,
            unit_weight
        FROM mds_account_substitution_output
        WHERE (LOWER(item_code) LIKE LOWER(%s) OR
               LOWER(item_name) LIKE LOWER(%s) OR
               LOWER(output_number) LIKE LOWER(%s))
        AND output_number LIKE %s
        AND item_code NOT IN %s
        """
        cursor.execute(query, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", like_pattern, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("검색 결과가 없습니다.")
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview2.config(text="")  # 총계 라벨 초기화
            return

        treeview.delete(*treeview.get_children())
        logging.debug("트리뷰 초기화 완료")

        numeric_columns = ['양품출고량', '단가', '금액', '외화금액', '단위중량']

        # 합계를 계산할 컬럼들
        sum_columns = {
            '양품출고량': 0.0,
            '금액': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'NO' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']

                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview2.config(text=total_text)

        logging.debug("검색 결과 삽입 완료")

    except Exception as e:
        logging.exception("search_treeview2 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")



def search_treeview3(treeview, keyword, selected_month):
    """
    mds_account_substitution_input 테이블에서 통합 검색을 수행하는 함수.
    """
    global total_label_treeview3  # 총계 라벨 변수 선언
    logging.info(f"search_treeview3 함수 시작. 검색어: {keyword}, 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        logging.error("데이터베이스 연결 실패")
        return

    # 조회월 형식 확인 및 처리
    selected_month_str = selected_month.replace('/', '')
    if len(selected_month_str) != 6:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # LIKE 패턴 생성
        like_pattern = f"TGI{selected_month_str}%"

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, quantity, unit_price, amount,
            substitution_type, warehouse, input_number, request_number, department,
            manager, header_note, line_note
        FROM mds_account_substitution_input
        WHERE (LOWER(item_code) LIKE LOWER(%s) OR
               LOWER(item_name) LIKE LOWER(%s) OR
               LOWER(input_number) LIKE LOWER(%s))
        AND input_number LIKE %s
        AND item_code NOT IN %s
        """
        cursor.execute(query, (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%", like_pattern, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("검색 결과가 없습니다.")
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview3.config(text="")  # 총계 라벨 초기화
            return

        treeview.delete(*treeview.get_children())
        logging.debug("트리뷰 초기화 완료")

        numeric_columns = ['입고량', '단가', '금액']

        # 합계를 계산할 컬럼들
        sum_columns = {
            '입고량': 0.0,
            '금액': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'NO' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']

                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview3.config(text=total_text)

        logging.debug("검색 결과 삽입 완료")

    except Exception as e:
        logging.exception("search_treeview3 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")



def search_treeview4(treeview, keyword, selected_month):
    """
    mds_shipment_status 테이블에서 통합 검색을 수행하는 함수.
    """
    global total_label_treeview4  # 총계 라벨 변수 선언
    logging.info(f"search_treeview4 함수 시작. 검색어: {keyword}, 조회월: {selected_month}")
    
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        logging.error("데이터베이스 연결 실패")
        return

    # 조회월 형식 확인
    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, shipment_quantity, unit_price, amount,
            won_amount_shipment, vat_shipment, won_amount_sales, vat_sales,
            total_amount_shipment, total_amount_sales, weight, reference_month
        FROM mds_shipment_status
        WHERE (LOWER(item_code) LIKE LOWER(%s) OR
               LOWER(item_name) LIKE LOWER(%s))
        AND reference_month = %s
        AND item_code NOT IN %s
        ORDER BY reference_month DESC, item_code
        """
        cursor.execute(query, (f"%{keyword}%", f"%{keyword}%", selected_month, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("검색 결과가 없습니다.")
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview4.config(text="")  # 총계 라벨 초기화
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())
        logging.debug("트리뷰 초기화 완료")

        numeric_columns = [
            '출하수량', '단가', '금액', '원화금액(출하)', '부가세(출하)',
            '원화금액(매출)', '부가세(매출)', '총금액(출하)', '총금액(매출)', '중량'
        ]

        # 합계를 계산할 컬럼들
        sum_columns = {
            '출하수량': 0.0,
            '금액': 0.0,
            '원화금액(출하)': 0.0,
            '부가세(출하)': 0.0,
            '원화금액(매출)': 0.0,
            '부가세(매출)': 0.0,
            '총금액(출하)': 0.0,
            '총금액(매출)': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'NO' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']

                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview4.config(text=total_text)

        logging.debug("검색 결과 삽입 완료")

    except Exception as e:
        logging.exception("search_treeview4 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")




def search_treeview5(treeview, keyword, selected_month):
    """
    [트리뷰5 통합 검색 함수]
      - mds_purchase_receipt_status 테이블 대상
      - (item_code, item_name) 기준으로 부분 검색 (keyword)
      - reference_month = selected_month
      - exclude_item_codes 테이블에서 제외 품목
      - 검색 결과를 트리뷰5에 표시, 합계(부가세·총금액·관리수량·원화금액) 계산
    """
    global total_label_treeview5

    logging.info(f"search_treeview5 함수 시작. 검색어='{keyword}', 조회월='{selected_month}'")

    # (1) DB 연결
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    # (2) 조회월 형식 확인
    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()

        # (A) 제외 품목코드 목록
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # (B) 검색 쿼리
        #   - item_code, item_name 중 하나라도 keyword를 포함(대소문자 구분 X)
        #   - reference_month = selected_month
        #   - item_code NOT IN exclude_codes
        query = """
            SELECT
                item_code,
                item_name,
                specification,
                vat,
                total_amount,
                management_quantity,
                won_amount,
                supplier_name,
                supplier_code,
                reference_month
            FROM mds_purchase_receipt_status
            WHERE reference_month = %s
              AND item_code NOT IN %s
              AND (
                   LOWER(item_code) LIKE LOWER(%s)
                   OR LOWER(item_name) LIKE LOWER(%s)
                  )
            ORDER BY item_code
        """
        # 파라미터: (selected_month, tuple(exclude_codes), f"%{keyword}%", f"%{keyword}%")
        cursor.execute(query, (selected_month, tuple(exclude_codes),
                               f"%{keyword}%", f"%{keyword}%"))
        rows = cursor.fetchall()
        logging.debug(f"검색 결과 행 수: {len(rows)}")

        if not rows:
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview5.config(text="")
            return

        # (C) 트리뷰 초기화
        clear_treeview(treeview)

        # (D) 합계 계산용 컬럼 (트리뷰 헤더 기준)
        sum_columns = {
            '부가세': 0.0,
            '총금액': 0.0,
            '관리수량': 0.0,
            '원화금액': 0.0
        }
        # DB 컬럼명 vs 트리뷰 헤더 매핑:
        #  vat    -> '부가세'
        #  total_amount -> '총금액'
        #  management_quantity -> '관리수량'
        #  won_amount -> '원화금액'

        # (E) 검색 결과 삽입
        for idx, row in enumerate(rows):
            # row: (item_code, item_name, specification, vat, total_amount,
            #       management_quantity, won_amount, supplier_name, supplier_code, reference_month)
            # 트리뷰에 삽입할 values: [NO, ...row...]
            values = [idx + 1]

            for col_index, db_value in enumerate(row):
                # treeview['columns']: ('no','item_code','item_name',...)
                # 'no'를 제외하면 col_index+1
                tree_col_id = treeview['columns'][col_index + 1]
                heading_text = treeview.heading(tree_col_id)['text']  # 예: '부가세','품목코드',...

                if heading_text in sum_columns:
                    # 숫자 변환 후 합계
                    numeric_val = safe_float_from_string(db_value) or 0.0
                    sum_columns[heading_text] += numeric_val
                    # 표시는 천단위 콤마
                    values.append(format_numeric_value(db_value))
                else:
                    values.append(db_value)

            treeview.insert('', 'end', values=values)

        # (F) 음영 처리
        tag_alternate_rows(treeview)
        reset_treeview_state(treeview)

        # (G) 총계 라벨
        total_text = ', '.join([f"{k} 합계: {format_numeric_value(v)}" for k, v in sum_columns.items()])
        total_label_treeview5.config(text=total_text)

        logging.info("search_treeview5 완료.")

    except Exception as e:
        logging.exception("search_treeview5 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        conn.close()



def search_treeview6(treeview, keyword, selected_month):
    """
    mds_inventory_evaluation 테이블에서 통합 검색을 수행하는 함수.
    """
    global total_label_treeview6  # 총계 라벨 변수 선언
    logging.info(f"search_treeview6 함수 시작. 검색어: {keyword}, 조회월: {selected_month}")

    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        logging.error("데이터베이스 연결 실패")
        return

    # 조회월 형식 확인
    if len(selected_month) != 7 or '/' not in selected_month:
        messagebox.showerror("Error", "올바른 조회월 형식이 아닙니다. 예: YYYY/MM")
        return

    try:
        cursor = conn.cursor()
        logging.debug("데이터베이스 커서 획득")

        # 제외할 품목코드 목록 가져오기
        cursor.execute("SELECT item_code FROM exclude_item_codes")
        exclude_codes = [row[0] for row in cursor.fetchall()]
        logging.debug(f"제외할 품목코드 목록: {exclude_codes}")

        # 쿼리 작성
        query = """
        SELECT
            item_code, item_name, specification, beginning_quantity, beginning_unit_price, beginning_amount,
            receipt_quantity, receipt_amount, substitution_quantity, substitution_amount, shipment_quantity,
            shipment_amount, inventory_quantity, inventory_unit_price, inventory_amount, reference_month
        FROM mds_inventory_evaluation
        WHERE (LOWER(item_code) LIKE LOWER(%s) OR
               LOWER(item_name) LIKE LOWER(%s))
        AND reference_month = %s
        AND item_code NOT IN %s
        ORDER BY reference_month DESC, item_code
        """
        cursor.execute(query, (f"%{keyword}%", f"%{keyword}%", selected_month, tuple(exclude_codes)))
        rows = cursor.fetchall()
        logging.debug(f"가져온 행의 수: {len(rows)}")

        if not rows:
            logging.info("검색 결과가 없습니다.")
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview6.config(text="")  # 총계 라벨 초기화
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())
        logging.debug("트리뷰 초기화 완료")

        numeric_columns = [
            '기초수량', '기초단가', '기초금액',
            '입고수량', '입고금액', '대체수량', '대체금액',
            '출고수량', '출고금액', '재고수량', '재고단가', '재고금액'
        ]

        # 합계를 계산할 컬럼들
        sum_columns = {
            '기초수량': 0.0,
            '기초금액': 0.0,
            '입고수량': 0.0,
            '입고금액': 0.0,
            '대체수량': 0.0,
            '대체금액': 0.0,
            '출고수량': 0.0,
            '출고금액': 0.0,
            '재고수량': 0.0,
            '재고금액': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1]  # 'NO' 컬럼 값 추가

            for col_index, value in enumerate(row):
                adjusted_col_index = col_index + 1  # 'NO' 컬럼으로 인한 인덱스 조정
                header_text = treeview.heading(treeview['columns'][adjusted_col_index])['text']

                if header_text in numeric_columns:
                    formatted_value = format_numeric_value(value)
                    values.append(formatted_value)
                    # 합계 계산 코드
                    if header_text in sum_columns:
                        numeric_value = safe_float_from_string(value) or 0.0
                        sum_columns[header_text] += numeric_value
                else:
                    values.append(value)

            treeview.insert('', 'end', values=values)
            logging.debug(f"행 {idx} 데이터 삽입 완료")

        # 음영 처리 적용
        tag_alternate_rows(treeview)

        # 트리뷰 상태 초기화
        reset_treeview_state(treeview)

        # 총계 라벨 업데이트
        total_text = ', '.join([f"{key} 합계: {format_numeric_value(value)}" for key, value in sum_columns.items()])
        total_label_treeview6.config(text=total_text)

        logging.debug("검색 결과 삽입 완료")

    except Exception as e:
        logging.exception("search_treeview6 함수 실행 중 예외 발생")
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
            logging.debug("데이터베이스 커서 닫음")
        if conn:
            conn.close()
            logging.debug("데이터베이스 연결 닫음")

def search_treeview7(treeview, keyword, selected_year):
    global total_label_treeview7  # 총계 라벨 변수 선언
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("에러", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT item_code, item_name, specification, unit, category,
                   beginning_unit_price, beginning_quantity, beginning_amount
            FROM mds_basic_data
            WHERE reference_year = %s AND
                  (item_code ILIKE %s OR item_name ILIKE %s)
        """, (selected_year, f"%{keyword}%", f"%{keyword}%"))
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("정보", "검색 결과가 없습니다.")
            total_label_treeview7.config(text="")  # 총계 라벨 초기화
            return

        # 트리뷰 초기화
        treeview.delete(*treeview.get_children())

        # 합계 계산을 위한 초기화 (단가 합계 제거)
        sum_columns = {
            # 'beginning_unit_price': 0.0,  # 단가 합계 계산 제거
            'beginning_quantity': 0.0,
            'beginning_amount': 0.0
        }

        for idx, row in enumerate(rows):
            values = [idx + 1] + list(row)
            treeview.insert('', 'end', values=values)

            # 합계 계산 (단가 합계 계산 제거)
            # sum_columns['beginning_unit_price'] += safe_float_from_string(row[5])
            sum_columns['beginning_quantity'] += safe_float_from_string(row[6]) or 0.0
            sum_columns['beginning_amount'] += safe_float_from_string(row[7]) or 0.0

        # 총계 라벨 업데이트 (단가 합계 제거)
        total_text = (
            f"수량 합계: {format_numeric_value(sum_columns['beginning_quantity'])}, "
            f"금액 합계: {format_numeric_value(sum_columns['beginning_amount'])}"
        )
        total_label_treeview7.config(text=total_text)

        # 숫자 열 포맷팅 적용
        numeric_columns = ['단가', '수량', '금액']
        format_numeric_columns(treeview, numeric_columns)

        # 음영 처리 적용
        tag_alternate_rows(treeview)

    except Exception as e:
        messagebox.showerror("에러", f"검색 중 오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def search_treeview8(treeview, keyword):
    if not hasattr(treeview, 'all_data') or not treeview.all_data:
        messagebox.showinfo("알림", "데이터를 먼저 로딩해주세요.")
        return

    filtered = []
    for row in treeview.all_data:
        # row: (no, supplier_code, supplier_name, item_code, item_name, sum_won, sum_vat, sum_total, ratio, incentive)
        if any(keyword.lower() in str(cell).lower() for cell in row):
            filtered.append(row)

    # 기존 데이터 삭제
    for item in treeview.get_children():
        treeview.delete(item)

    # 삽입
    for idx, row in enumerate(filtered):
        tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
        treeview.insert('', 'end', values=row, tags=(tag,))




def add_code(listbox):
    """
    새로운 품목코드를 추가하는 함수입니다.
    """
    new_code = simpledialog.askstring("추가", "추가할 품목코드를 입력하세요:")
    if new_code:
        listbox.insert(tk.END, new_code)

def edit_code(listbox):
    """
    선택한 품목코드를 수정하는 함수입니다.
    """
    selected_index = listbox.curselection()
    if selected_index:
        current_code = listbox.get(selected_index)
        new_code = simpledialog.askstring("수정", "품목코드를 수정하세요:", initialvalue=current_code)
        if new_code:
            listbox.delete(selected_index)
            listbox.insert(selected_index, new_code)
    else:
        messagebox.showwarning("선택 오류", "수정할 품목코드를 선택하세요.")

def delete_code(listbox):
    """
    선택한 품목코드를 삭제하는 함수입니다.
    """
    selected_index = listbox.curselection()
    if selected_index:
        listbox.delete(selected_index)
    else:
        messagebox.showwarning("선택 오류", "삭제할 품목코드를 선택하세요.")

def save_codes(listbox):
    """
    리스트박스의 품목코드들을 데이터베이스에 저장하는 함수입니다.
    기존 데이터를 모두 삭제하고 현재 리스트박스의 내용으로 대체합니다.
    """
    codes = listbox.get(0, tk.END)
    if messagebox.askyesno("저장 확인", "변경사항을 저장하시겠습니까?"):
        conn = get_postgres_connection()
        if not conn:
            messagebox.showerror("Error", "데이터베이스 연결 실패")
            return
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM exclude_item_codes")
            for code in codes:
                cursor.execute("INSERT INTO exclude_item_codes (item_code) VALUES (%s)", (code,))
            conn.commit()
            messagebox.showinfo("저장 완료", "변경사항이 저장되었습니다.")
        except Exception as e:
            conn.rollback()
            messagebox.showerror("오류", f"저장 중 오류 발생: {e}")
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()


def create_info_frame(parent):
    """
    정보 전달 프레임 생성.
    """
    # 글로벌 변수를 선언
    global master_label, warehouse_label, output_label, input_label

    # 하나의 프레임 생성 (Gray 스타일 적용)
    info_frame = ttk.Frame(parent, style='Gray.TLabelframe')
    info_frame.pack(fill='x', padx=5, pady=5)

    # 각종 레이블을 가로로 배치
    master_label = ttk.Label(info_frame, text="공장품목등록: ")
    master_label.pack(side='left', padx=10, pady=5)

    warehouse_label = ttk.Label(info_frame, text="창고별재고현황: ")
    warehouse_label.pack(side='left', padx=10, pady=5)

    output_label = ttk.Label(info_frame, text="계정대체출고현황: ")
    output_label.pack(side='left', padx=10, pady=5)

    input_label = ttk.Label(info_frame, text="계정대체입고현황: ")
    input_label.pack(side='left', padx=10, pady=5)

    # 정보 로드
    load_last_updated_info()


def load_last_updated_info():
    conn = get_postgres_connection()
    if not conn:
        messagebox.showerror("Error", "데이터베이스 연결 실패")
        return

    try:
        cursor = conn.cursor()

        # master 테이블의 registration_date의 최대값 가져오기
        cursor.execute("SELECT MAX(registration_date) FROM master")
        master_registration_date = format_time(cursor.fetchone()[0])  # 날짜 형식이므로 format_time 사용

        # master 테이블의 last_updated의 최대값 가져오기 (창고별재고현황에 사용)
        cursor.execute("SELECT MAX(last_updated) FROM master")
        master_last_updated = format_time(cursor.fetchone()[0])

        # mds_account_substitution_output 테이블의 last_updated의 최대값 가져오기
        cursor.execute("SELECT MAX(last_updated) FROM mds_account_substitution_output")
        output_last_updated = format_time(cursor.fetchone()[0])

        # mds_account_substitution_input 테이블의 last_updated의 최대값 가져오기
        cursor.execute("SELECT MAX(last_updated) FROM mds_account_substitution_input")
        input_last_updated = format_time(cursor.fetchone()[0])

        # 레이블 업데이트
        master_label.config(text=f"공장품목등록(최신 등록일): {master_registration_date}")
        warehouse_label.config(text=f"창고별재고현황: {master_last_updated}")
        output_label.config(text=f"계정대체출고현황: {output_last_updated}")
        input_label.config(text=f"계정대체입고현황: {input_last_updated}")

    except Exception as e:
        messagebox.showerror("Error", f"오류 발생: {e}")
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


def format_time(dt):
    """datetime 또는 date 객체를 적절한 형식으로 포맷"""
    if dt:
        if isinstance(dt, datetime):
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(dt, date):  # datetime.date가 아닌 date로 변경
            return dt.strftime('%Y-%m-%d')
    return "업데이트 기록 없음"

def main():
    global treeview0_frame, treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame
    global treeview0, treeview1, treeview2, treeview3, treeview4, treeview5, treeview6, treeview7, treeview8
    global date_frame_treeview0, date_frame_treeview1, date_frame_treeview2, date_frame_treeview3, date_frame_treeview4, date_frame_treeview5, date_frame_treeview6, date_frame_treeview7
    global total_frame_treeview2, total_frame_treeview3, total_frame_treeview4, total_frame_treeview5, total_frame_treeview6, total_frame_treeview7, total_frame_treeview8
    global total_label_treeview1, total_label_treeview2, total_label_treeview3, total_label_treeview4, total_label_treeview5, total_label_treeview6, total_label_treeview7
    global current_treeview, current_treeview_label
    global search_entry
    global root
    global config
    global date_var_treeview2, date_var_treeview3, date_var_treeview4, date_var_treeview5, date_var_treeview6, date_var_treeview7
    global filter_vars  # 필터 변수 딕셔너리 추가
    global header_height  # 헤더 높이 변수 추가
    global header_canvas  # 헤더 캔버스 추가
    global status_label_for_treeview0

    # 헤더 높이 설정
    header_height = 50  # 헤더의 높이를 픽셀 단위로 지정합니다.

    # 필터 변수 딕셔너리 초기화
    filter_vars = {}

    # 'config' 객체 초기화
    config = configparser.ConfigParser()

    # 메인 윈도우 생성
    root = tk.Tk()
    root.title("경영지원본부 재고수불현황 V.25-07-07_A by 시스템관리팀")
    root.geometry('1920x1080')

    # 스타일 설정
    style = ttk.Style(root)
    style.theme_use('clam')

    style.configure('Green.TButton', background='green', foreground='white')
    style.configure('Coral.TButton', background='coral', foreground='snow')
    style.configure('Yellow.TButton', background='yellow', foreground='black')

    # 새로운 버튼 스타일 정의: Custom.TButton
    style.configure('Custom.TButton',
                    background="#FFD700",       # 금색 배경
                    foreground="black",         # 검정 전경
                    font=('Helvetica', 12, 'bold'),  # 글꼴 설정
                    relief='solid',             # 실선 경계
                    borderwidth=2,              # 테두리 두께
                    padding=5)                  # 패딩

    style.map('Custom.TButton',
              background=[('active', '#FFC125')],  # 활성 상태의 배경색
              foreground=[('pressed', 'black'), ('active', 'black')])  # 활성/클릭 상태의 전경색

    # 새 스타일 정의: Gray.TLabelframe와 Gray.TCheckbutton
    style.configure("Gray.TLabelframe",
                    background="#D3D3D3",
                    foreground="black",
                    borderwidth=1,
                    relief="solid")
    style.configure("Gray.TLabelframe.Label",
                    background="#D3D3D3",
                    foreground="black")
    style.configure("Gray.TCheckbutton",
                    background="#D3D3D3",
                    foreground="black")


    # 메인 창 위치 불러오기
    try:
        if 'WINDOW' in config and 'position' in config['WINDOW']:
            position_value = config['WINDOW']['position']
            logging.info(f"불러온 창 위치: {position_value}")
            root.geometry(str(position_value))  # 문자열로 변환하여 전달
    except Exception as e:
        logging.error(f"창 위치 설정 중 오류 발생: {e}")

    def save_window_position():
        # 메인 창 위치 저장
        if 'WINDOW' not in config:
            config['WINDOW'] = {}
        config['WINDOW']['position'] = root.geometry()

        # 기존 데이터 유지하며 설정 파일 업데이트 (UTF-8 인코딩)
        with open(config_path, 'w', encoding='utf-8') as configfile:
            config.write(configfile)
        logging.info("메인 창 위치 저장 완료")

    # 창 닫기 이벤트 처리
    def on_closing_main():
        save_window_position()
        save_filter_settings(filter_vars)  # 필터 설정 저장 추가
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_closing_main)

    # 메인 프레임 생성
    main_frame = tk.Frame(root)
    main_frame.pack(fill='both', expand=True)

    # 좌우로 분할된 프레임 생성
    left_frame = ttk.Frame(main_frame)
    left_frame.pack(side='left', fill='y')

    right_frame = ttk.Frame(main_frame)
    right_frame.pack(side='right', fill='both', expand=True)

    # content_frame과 info_frame을 right_frame에 배치
    content_frame = ttk.Frame(right_frame, style='Gray.TLabelframe')
    content_frame.pack(side='top', fill='both', expand=True)

    info_frame = ttk.Frame(right_frame, style='Gray.TLabelframe')
    info_frame.pack(side='bottom', fill='x')

    # 검색 프레임 생성
    search_frame = ttk.Frame(content_frame, style='Gray.TLabelframe')
    search_frame.pack(side='top', fill='x', pady=5)

    # 검색 입력 필드
    search_entry = ttk.Entry(search_frame, width=40)  # width 옵션 추가
    search_entry.pack(side="left", padx=5, pady=2)

    # 검색 버튼 추가
    search_button = ttk.Button(
        search_frame,
        text="검색",
        command=search_current_treeview
    )
    search_button.pack(side='left', padx=5, pady=2)  # pady 줄임

    # 'xlsx 다운로드' 버튼 추가
    download_button = ttk.Button(
        search_frame,
        text="xlsx 다운로드",
        command=download_current_treeview
    )
    download_button.pack(side='left', padx=5, pady=2)

    # 엔터 키 바인딩
    search_entry.bind("<Return>", search_current_treeview)

    # 날짜 선택을 위한 함수
    def get_previous_month():
        today = datetime.today()
        first = today.replace(day=1)
        last_month = first - timedelta(days=1)
        return last_month.strftime("%Y/%m")

    # 날짜 변수 초기화
    date_var_treeview1 = tk.StringVar()
    date_var_treeview2 = tk.StringVar()
    date_var_treeview3 = tk.StringVar()
    date_var_treeview4 = tk.StringVar()
    date_var_treeview5 = tk.StringVar()
    date_var_treeview6 = tk.StringVar()
    date_var_treeview7 = tk.StringVar()

    previous_month = get_previous_month()
    date_var_treeview1.set(previous_month)
    date_var_treeview2.set(previous_month)
    date_var_treeview3.set(previous_month)
    date_var_treeview4.set(previous_month)
    date_var_treeview5.set(previous_month)
    date_var_treeview6.set(previous_month)
    date_var_treeview7.set(datetime.now().year)

    # tk.LabelFrame 사용
    treeview_frame = tk.LabelFrame(content_frame, text="", bg="SystemButtonFace")
    treeview_frame.pack(side="top", fill="both", expand=True, padx=2, pady=2)

    current_treeview_label = tk.Label(
        treeview_frame,
        text="현재 트리뷰: 없음",
        font=("맑은 고딕", 9),
        fg="blue",
        bg=treeview_frame.cget("background"),  # 부모 tk.LabelFrame 배경색
        anchor="w"
    )
    current_treeview_label.pack(side="top", fill="x", padx=5, pady=5)

    treeview_buttons_frame = tk.Frame(
        treeview_frame,
        bg=treeview_frame.cget("background")  # 부모 tk.LabelFrame 배경색
    )
    treeview_buttons_frame.pack(side="top", fill='x')

    treeview_display_frame = ttk.Frame(treeview_frame)
    treeview_display_frame.pack(side="top", fill='both', expand=True)

    # 트리뷰 프레임 생성
    treeview0_frame = tk.Frame(treeview_display_frame)
    treeview1_frame = tk.Frame(treeview_display_frame)
    treeview2_frame = tk.Frame(treeview_display_frame)
    treeview3_frame = tk.Frame(treeview_display_frame)
    treeview4_frame = tk.Frame(treeview_display_frame)
    treeview5_frame = tk.Frame(treeview_display_frame)
    treeview6_frame = tk.Frame(treeview_display_frame)
    treeview7_frame = tk.Frame(treeview_display_frame)
    treeview8_frame = tk.Frame(treeview_display_frame)


    # 날짜 선택 프레임 생성
    date_frame_treeview0 = ttk.Frame(treeview_display_frame)
    date_frame_treeview1 = ttk.Frame(treeview_display_frame)
    date_frame_treeview2 = ttk.Frame(treeview_display_frame)
    date_frame_treeview3 = ttk.Frame(treeview_display_frame)
    date_frame_treeview4 = ttk.Frame(treeview_display_frame)
    date_frame_treeview5 = ttk.Frame(treeview_display_frame)
    date_frame_treeview6 = ttk.Frame(treeview_display_frame)
    date_frame_treeview7 = ttk.Frame(treeview_display_frame)
   
    # 트리뷰1
    date_label_treeview1 = ttk.Label(date_frame_treeview1, text="조회월:")
    date_entry_treeview1 = ttk.Entry(date_frame_treeview1, textvariable=date_var_treeview1, width=10)
    date_label_treeview1.pack(side="left")
    date_entry_treeview1.pack(side="left")

    # 엔터키 바인딩 (load_treeview1_data에 조회월 전달)
    date_entry_treeview1.bind("<Return>", lambda event: load_treeview1_data(treeview1, date_var_treeview1.get()))

    # 트리뷰2
    date_label_treeview2 = ttk.Label(date_frame_treeview2, text="조회월:")
    date_entry_treeview2 = ttk.Entry(date_frame_treeview2, textvariable=date_var_treeview2, width=10)
    date_label_treeview2.pack(side="left")
    date_entry_treeview2.pack(side="left")

    # 엔터키 바인딩 추가
    date_entry_treeview2.bind("<Return>", lambda event: load_treeview2_data(treeview2, date_var_treeview2.get()))

    # 트리뷰3
    date_label_treeview3 = ttk.Label(date_frame_treeview3, text="조회월:")
    date_entry_treeview3 = ttk.Entry(date_frame_treeview3, textvariable=date_var_treeview3, width=10)
    date_label_treeview3.pack(side="left")
    date_entry_treeview3.pack(side="left")

    # 엔터키 바인딩 추가
    date_entry_treeview3.bind("<Return>", lambda event: load_treeview3_data(treeview3, date_var_treeview3.get()))

    # 트리뷰4
    date_label_treeview4 = ttk.Label(date_frame_treeview4, text="조회월:")
    date_entry_treeview4 = ttk.Entry(date_frame_treeview4, textvariable=date_var_treeview4, width=10)
    date_label_treeview4.pack(side="left")
    date_entry_treeview4.pack(side="left")

    # 엔터키 바인딩 추가
    date_entry_treeview4.bind("<Return>", lambda event: load_treeview4_data(treeview4, date_var_treeview4.get()))

    # 트리뷰5
    date_label_treeview5 = ttk.Label(date_frame_treeview5, text="조회월:")
    date_entry_treeview5 = ttk.Entry(date_frame_treeview5, textvariable=date_var_treeview5, width=10)
    date_label_treeview5.pack(side="left")
    date_entry_treeview5.pack(side="left")

    # 엔터키 바인딩 추가
    date_entry_treeview5.bind("<Return>", lambda event: load_treeview5_data(treeview5, date_var_treeview5.get()))

    # 트리뷰6
    date_label_treeview6 = ttk.Label(date_frame_treeview6, text="조회월:")
    date_entry_treeview6 = ttk.Entry(date_frame_treeview6, textvariable=date_var_treeview6, width=10)
    date_label_treeview6.pack(side="left")
    date_entry_treeview6.pack(side="left")

    # 엔터키 바인딩 추가
    date_entry_treeview6.bind("<Return>", lambda event: load_treeview6_data(treeview6, date_var_treeview6.get()))

    # 트리뷰7
    date_label_treeview7 = ttk.Label(date_frame_treeview7, text="기준년:")
    date_entry_treeview7 = ttk.Entry(date_frame_treeview7, textvariable=date_var_treeview7, width=10)
    date_label_treeview7.pack(side="left")
    date_entry_treeview7.pack(side="left")    

    # 엔터키 바인딩 추가
    date_entry_treeview7.bind("<Return>", lambda event: load_treeview7_data(treeview7, date_var_treeview7.get()))

    # 총계 프레임 생성 (Treeview0과 Treeview1 제외)
    total_frame_treeview2 = ttk.Frame(treeview_display_frame)
    total_frame_treeview3 = ttk.Frame(treeview_display_frame)
    total_frame_treeview4 = ttk.Frame(treeview_display_frame)
    total_frame_treeview5 = ttk.Frame(treeview_display_frame)
    total_frame_treeview6 = ttk.Frame(treeview_display_frame)
    total_frame_treeview7 = ttk.Frame(treeview_display_frame)
    total_frame_treeview8 = ttk.Frame(treeview_display_frame)

    # 총계 라벨 생성 (Treeview0과 Treeview1 제외)
    total_label_treeview2 = ttk.Label(total_frame_treeview2, text="총계:")
    total_label_treeview2.pack(side="left", padx=5)
    total_label_treeview3 = ttk.Label(total_frame_treeview3, text="총계:")
    total_label_treeview3.pack(side="left", padx=5)
    total_label_treeview4 = ttk.Label(total_frame_treeview4, text="총계:")
    total_label_treeview4.pack(side="left", padx=5)
    total_label_treeview5 = ttk.Label(total_frame_treeview5, text="총계:")
    total_label_treeview5.pack(side="left", padx=5)
    total_label_treeview6 = ttk.Label(total_frame_treeview6, text="총계:")
    total_label_treeview6.pack(side="left", padx=5)
    total_label_treeview7 = ttk.Label(total_frame_treeview7, text="총계:")
    total_label_treeview7.pack(side="left", padx=5)

    # 초기에는 treeview0_frame만 보이도록 설정
    treeview0_frame.pack(fill='both', expand=True)

    # 트리뷰 생성
    treeview0, status_label_for_treeview0 = create_treeview0(treeview0_frame)
    treeview1 = create_treeview1(treeview1_frame)
    treeview2 = create_treeview2(treeview2_frame)
    treeview3 = create_treeview3(treeview3_frame)
    treeview4 = create_treeview4(treeview4_frame)
    treeview5 = create_treeview5(treeview5_frame)
    treeview6 = create_treeview6(treeview6_frame)
    treeview7 = create_treeview7(treeview7_frame)
    treeview8 = create_treeview8(treeview8_frame) 

    # 프로그램 시작 시 필터 설정 로드
    load_filter_settings(filter_vars)

    # 필터 설정을 기반으로 트리뷰에 필터 적용
    apply_column_filters(treeview0)

    # 초기 트리뷰 설정
    current_treeview = None  # 초기화
    show_treeview(treeview0_frame, [], 'treeview0')
    load_treeview0_data(treeview0)

    # 트리뷰 버튼들
    load_treeview1_button = ttk.Button(
        treeview_buttons_frame,
        text="창고별재고현황",
        command=lambda: [
            show_treeview(
                treeview1_frame,
                [treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview1',
                date_frame_treeview1  # ← 트리뷰1 날짜 프레임
            ),
            load_treeview1_data(treeview1, date_var_treeview1.get()),
            # 만약 트리뷰1에 total_frame_treeview1이 있다면:
            # total_frame_treeview1.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview1_button.pack(side='left', padx=5, pady=2)

    load_treeview2_button = ttk.Button(
        treeview_buttons_frame,
        text="계정대체출고현황",
        command=lambda: [
            show_treeview(
                treeview2_frame,
                [treeview1_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview2',
                date_frame_treeview2
            ),
            load_treeview2_data(treeview2, date_var_treeview2.get()),
            total_frame_treeview2.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview2_button.pack(side="left", padx=5, pady=2)

    load_treeview3_button = ttk.Button(
        treeview_buttons_frame,
        text="계정대체입고현황",
        command=lambda: [
            show_treeview(
                treeview3_frame,
                [treeview1_frame, treeview2_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview3',
                date_frame_treeview3
            ),
            load_treeview3_data(treeview3, date_var_treeview3.get()),
            total_frame_treeview3.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview3_button.pack(side="left", padx=5, pady=2)

    load_treeview4_button = ttk.Button(
        treeview_buttons_frame,
        text="출하현황",
        command=lambda: [
            show_treeview(
                treeview4_frame,
                [treeview1_frame, treeview2_frame, treeview3_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview4',
                date_frame_treeview4
            ),
            load_treeview4_data(treeview4, date_var_treeview4.get()),
            total_frame_treeview4.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview4_button.pack(side="left", padx=5, pady=2)

    load_treeview5_button = ttk.Button(
        treeview_buttons_frame,
        text="구매입고현황",
        command=lambda: [
            show_treeview(
                treeview5_frame,
                [treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview6_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview5',
                date_frame_treeview5
            ),
            load_treeview5_data(treeview5, date_var_treeview5.get()),
            total_frame_treeview5.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview5_button.pack(side="left", padx=5, pady=2)

    # 트리뷰6 버튼 추가
    load_treeview6_button = ttk.Button(
        treeview_buttons_frame,
        text="재고평가",
        command=lambda: [
            show_treeview(
                treeview6_frame,
                [treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview7_frame, treeview8_frame, treeview0_frame],
                'treeview6',
                date_frame_treeview6
            ),
            load_treeview6_data(treeview6, date_var_treeview6.get()),
            total_frame_treeview6.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview6_button.pack(side="left", padx=5, pady=2)

    # '기초데이터' 버튼 추가
    load_treeview7_button = ttk.Button(
        treeview_buttons_frame,
        text="기초데이터",
        command=lambda: [
            show_treeview(
                treeview7_frame,
                [treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview8_frame, treeview0_frame],
                'treeview7',
                date_frame_treeview7
            ),
            load_treeview7_data(treeview7, date_var_treeview7.get()),
            total_frame_treeview7.pack(side="top", fill='x')
        ],
        style='Custom.TButton'
    )
    load_treeview7_button.pack(side='left', padx=5, pady=2)    

    # 장려금 버튼 추가
    load_treeview8_button = ttk.Button(
        treeview_buttons_frame,
        text="장려금",
        command=lambda: [
            show_treeview(
                treeview8_frame,
                [
                    treeview0_frame, treeview1_frame, treeview2_frame, treeview3_frame,
                    treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame
                ],
                'treeview8'
            )
        ],
        style='Custom.TButton'
    )
    load_treeview8_button.pack(side='left', padx=5, pady=2)

    # 트리뷰8 상단에 장려금 관련 버튼들 및 현재 기준월 표시 레이블 배치
    incentive_buttons_frame = ttk.Frame(treeview8_frame)
    incentive_buttons_frame.pack(side='top', fill='x', padx=5, pady=5)

    # 현재 작업 기준월 표시 레이블
    label_current_month_8 = ttk.Label(incentive_buttons_frame, text="현재 작업 기준월: -")
    label_current_month_8.pack(side='left', padx=10, pady=2)

    # 장려금배분 버튼
    distribute_incentive_button = ttk.Button(
        incentive_buttons_frame,
        text="장려금 배분",
        command=lambda: distribute_incentive_for_treeview8(treeview8, label_current_month_8),
        style='Custom.TButton'
    )
    distribute_incentive_button.pack(side='left', padx=5, pady=2)

    # 결과 저장하기 버튼
    save_incentive_button = ttk.Button(
        incentive_buttons_frame,
        text="결과 저장하기",
        command=lambda: save_incentive_result_for_treeview8(treeview8, label_current_month_8),
        style='Custom.TButton'
    )
    save_incentive_button.pack(side='left', padx=5, pady=2)


    # 결과 불러오기 버튼
    load_incentive_button = ttk.Button(
        incentive_buttons_frame,
        text="결과 불러오기",
        command=lambda: load_incentive_result_for_treeview8(treeview8, label_current_month_8),
        style='Custom.TButton'
    )
    load_incentive_button.pack(side='left', padx=5, pady=2)




    # 스타일 설정에 새로운 스타일 추가
    style.configure('Important.TButton',
                    background='red',
                    foreground='white',
                    font=('Helvetica', 12, 'bold'))
    style.map('Important.TButton',
              background=[('active', 'darkred')])

    load_treeview0_button = ttk.Button(
        treeview_buttons_frame,
        text="월별수불현황",
        command=lambda: [
            show_treeview(
                treeview0_frame,
                [treeview1_frame, treeview2_frame, treeview3_frame, treeview4_frame, treeview5_frame, treeview6_frame, treeview7_frame, treeview8_frame],
                'treeview0'
            ),
            # load_treeview0_data(treeview0)
            # Treeview0은 합계 프레임이 없으므로 추가 작업 없음
        ],
        style='Important.TButton'  # 중요한 버튼을 위한 새로운 스타일
    )
    load_treeview0_button.pack(side='left', padx=5, pady=2)



    # 업로드 프레임 생성 (트리뷰 하단에 배치)
    upload_frame = ttk.Frame(content_frame, padding=5)
    upload_frame.pack(side="top", fill='x', padx=2, pady=5)

    # ERP-IU DB 프레임 생성
    erp_iu_frame = ttk.LabelFrame(upload_frame, text="ERP-IU DB (xls) Upload", padding=5, style='Gray.TLabelframe')
    erp_iu_frame.pack(fill='x', padx=2, pady=2)

    # 버튼들 배치
    upload_master_button = ttk.Button(
        erp_iu_frame,
        text="공장품목등록",
        command=lambda: [upload_master_data(), load_last_updated_info()],
        style='Custom.TButton'
    )
    upload_master_button.pack(side="left", padx=(5, 2), pady=2)

    # upload_warehouse_button = ttk.Button(
    #     erp_iu_frame,
    #     text="창고별재고현황",
    #     command=lambda: [upload_warehouse_inventory(), load_last_updated_info()],
    #     style='Custom.TButton'
    # )
    # upload_warehouse_button.pack(side="left", padx=(5, 2), pady=2)

    upload_monthly_inventory_button = ttk.Button(
        erp_iu_frame,
        text="월별 기말재고등록",
        command=lambda: [upload_monthly_inventory_status(), load_last_updated_info()],
        style='Custom.TButton'
    )
    upload_monthly_inventory_button.pack(side="left", padx=(5, 2), pady=2)    

    upload_account_substitution_button = ttk.Button(
        erp_iu_frame,
        text="계정대체출고현황",
        command=lambda: upload_account_substitution_data(),
        style='Custom.TButton'
    )
    upload_account_substitution_button.pack(side="left", padx=(5, 2), pady=2)

    upload_account_substitution_input_button = ttk.Button(
        erp_iu_frame,
        text="계정대체입고현황",
        command=lambda: upload_account_substitution_input_data(),
        style='Custom.TButton'
    )
    upload_account_substitution_input_button.pack(side='left', padx=(5, 2), pady=2)

    upload_shipment_status_button = ttk.Button(
        erp_iu_frame,
        text="출하현황",
        command=lambda: upload_shipment_status_data(),
        style='Custom.TButton'
    )
    upload_shipment_status_button.pack(side='left', padx=(5, 2), pady=2)

    upload_purchase_receipt_button = ttk.Button(
        erp_iu_frame,
        text="구매입고현황",
        command=lambda: upload_purchase_receipt_status_data(),
        style='Custom.TButton'
    )
    upload_purchase_receipt_button.pack(side='left', padx=(5, 2), pady=2)

    # 재고평가 업로드 버튼 추가
    upload_inventory_evaluation_button = ttk.Button(
        erp_iu_frame,
        text="재고평가",
        command=lambda: upload_inventory_evaluation_data(),
        style='Custom.TButton'
    )
    upload_inventory_evaluation_button.pack(side='left', padx=(5, 2), pady=2)

    # "계정별원장" 버튼 추가
    upload_account_ledger_button = ttk.Button(
        erp_iu_frame,
        text="계정별원장",
        command=lambda: upload_account_ledger_data(),  # 계정별원장 업로드 함수 호출
        style='Custom.TButton'
    )
    upload_account_ledger_button.pack(side='left', padx=(5, 2), pady=2)

    # 데이터관리 프레임 생성
    data_management_frame = ttk.LabelFrame(upload_frame, text="데이터관리", padding=5, style='Gray.TLabelframe')
    data_management_frame.pack(fill='x', padx=2, pady=2)

    # 비수불 계정관리 버튼 추가
    exclude_item_button = ttk.Button(
        data_management_frame,
        text="비수불 계정관리",
        command=open_exclude_item_codes_window,
        style='Custom.TButton'
    )
    exclude_item_button.pack(side='left', padx=(5, 2), pady=2)


    # '기초데이터 등록' 버튼 추가
    upload_basic_data_button = ttk.Button(
        data_management_frame,
        text="기초데이터 등록",
        command=upload_basic_data,
        style='Custom.TButton'          
    )
    upload_basic_data_button.pack(side="left", padx=5, pady=2)    

    # 정보 전달 영역 생성 (info_frame에 배치)
    create_info_frame(info_frame)  # 수정된 info_frame에 생성

    root.mainloop()




import traceback

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        # 트레이스백을 로그에 포함
        logging.error("프로그램 실행 중 오류가 발생했습니다:", exc_info=True)
        
        # 콘솔에 트레이스백 출력
        print("프로그램 실행 중 오류가 발생했습니다:")
        traceback.print_exc()